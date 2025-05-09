import io
import pandas as pd
from django.db.models import Sum, F, Avg, Q, Count, Case, When, Value, DecimalField
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, TemplateView, DetailView, CreateView, UpdateView, View, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.utils import timezone
from django.db import models
from datetime import datetime, timedelta, date
from django.views.decorators.http import require_POST
import json
from django import forms

from .models import (
    Product, Wagon, Movement, Inventory, Batch, 
    Reservoir, ReservoirMovement, Warehouse, LocalClient, LocalMovement, Placement, Client, Transport, WagonType, 
    AnalyticsReport, AnalyticsData, InventoryAudit, InventoryAuditItem, ProductMinLevel, StockForecast, PurchasePlan, PurchasePlanItem,
    Supplier, SupplierRating, ProductSupplier, OrderPoint, PurchaseNotification, EstokadaOperation
)
from .forms import (
    MovementForm, ProductForm, WagonForm, BatchForm, ReservoirForm, ReservoirMovementForm, 
    WarehouseForm, LocalClientForm, LocalMovementForm, PlacementForm, ClientForm, WagonTypeForm,
    InventoryAuditForm, InventoryAuditItemForm, ProductMinLevelForm, PurchasePlanForm, PurchasePlanItemFormSet, GenerateForecastForm,
    SupplierForm, SupplierRatingForm, ProductSupplierForm, OrderPointForm, PurchaseNotificationForm, EstokadaOperationForm,
    TransportForm
)
from .filters import (
    MovementFilter, ProductFilter, WagonFilter, 
    ReservoirFilter, ReservoirMovementFilter, LocalMovementFilter
)
import json
from django.db import models
from django.urls import reverse_lazy
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import permission_required
from django.contrib import messages
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from apps.accounts.views import estokada_required, sales_required
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import datetime, timedelta, date
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
from scipy import stats

def index(request):
    return render(request, 'warehouse/index.html')

def dashboard(request):
    # Количество продуктов
    product_count = Product.objects.count()
    product_growth = 5  # Заглушка, можно рассчитать реальный рост
    
    # Резервуары
    reservoirs = Reservoir.objects.all()
    total_volume = reservoirs.aggregate(Sum('current_volume'))['current_volume__sum'] or 0
    reservoirs_capacity = reservoirs.aggregate(Sum('capacity'))['capacity__sum'] or 0
    capacity_percentage = (total_volume / reservoirs_capacity * 100) if reservoirs_capacity > 0 else 0
    
    # Вагоны
    wagon_count = Wagon.objects.count()
    active_wagons = Wagon.objects.filter(condition='Active').count()
    
    # Движения за последние 30 дней
    thirty_days_ago = timezone.now().date() - timezone.timedelta(days=30)
    movement_count = Movement.objects.filter(date__gte=thirty_days_ago).count()
    
    context = {
        'product_count': product_count,
        'product_growth': product_growth,
        'total_volume': total_volume,
        'reservoirs_capacity': reservoirs_capacity,
        'capacity_percentage': capacity_percentage,
        'wagon_count': wagon_count,
        'active_wagons': active_wagons,
        'movement_count': movement_count,
    }
    
    return render(request, 'warehouse/dashboard.html', context)


class MovementReportView(TemplateView):
    template_name = "warehouse/movement_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        movements = Movement.objects.all()
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        net = total_in - total_out
        context.update({
            'movements': movements,
            'total_in': total_in,
            'total_out': total_out,
            'net': net,
        })
        return context

class ProductListView(ListView):
    model = Product
    template_name = 'warehouse/product_list.html'
    context_object_name = 'products'
    ordering = ['code']

    def get_queryset(self):
        qs = super().get_queryset()
        self.filter = ProductFilter(self.request.GET, queryset=qs)
        return self.filter.qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filter
        return context

def recalculate_inventory(product):
    """Полностью пересчитывает запасы для продукта на основе всех имеющихся движений."""
    # Получаем все движения по этому продукту
    movements = Movement.objects.filter(product=product)
    
    # Вычисляем сумму всех поступлений и расходов
    total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
    total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
    
    # Рассчитываем фактический запас
    actual_quantity = total_in - total_out
    
    # Обновляем или создаем запись в инвентаре
    inventory, created = Inventory.objects.update_or_create(
        product=product,
        defaults={'quantity': actual_quantity}
    )
    product.in_qty = total_in
    product.out_qty = total_out
    product.save(update_fields=['in_qty', 'out_qty'])
    
    return inventory

class ProductDetailView(DetailView):
    model = Product
    template_name = 'warehouse/product_detail.html'
    context_object_name = 'product'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.get_object()
        
        inventory = recalculate_inventory(product)
        
        # Сначала получаем все движения для подсчета общих сумм
        all_movements = Movement.objects.filter(product=product)
        
        # Вычисляем суммы всех движений
        total_in = all_movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = all_movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        
        # Теперь получаем последние движения для отображения в таблице
        recent_movements = all_movements.order_by('-date')[:10]
        
        context.update({
            'total_in': total_in,
            'total_out': total_out,
            'net_qty': total_in - total_out,
            'current_quantity': inventory.quantity,
            'movements': recent_movements  # Отображаем только последние движения
        })
        return context

class ProductCreateView(CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'warehouse/product_create.html'

    def get_success_url(self):
        return reverse('warehouse:product_list')

class ProductUpdateView(UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'warehouse/product_update.html'

    def get_success_url(self):
        return reverse('warehouse:product_list')
    
class ProductInventoryReportView(TemplateView):
    template_name = 'warehouse/product_inventory_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = Product.objects.all()
        product_data = []
        for prod in products:
            product_data.append({
                'name': prod.name,
                'net_quantity': prod.net_quantity(),
            })
        context['product_data_json'] = json.dumps(product_data)
        return context

class WagonListView(ListView):
    model = Wagon
    template_name = 'warehouse/wagon_list.html'
    context_object_name = 'wagons'
    ordering = ['wagon_number']

    def get_queryset(self):
        qs = super().get_queryset()
        self.filter = WagonFilter(self.request.GET, queryset=qs)
        
        # Пересчитываем количество для каждого вагона
        for wagon in self.filter.qs:
            movements = Movement.objects.filter(wagon=wagon)
            total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
            total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
            current_quantity = total_in - total_out
            
            # Обновляем вагон, если текущее количество отличается
            if wagon.current_quantity != current_quantity:
                wagon.current_quantity = current_quantity
                wagon.save()
        
        return self.filter.qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filter
        return context
    
class WagonDetailView(DetailView):
    model = Wagon
    template_name = 'warehouse/wagon_detail.html'
    context_object_name = 'wagon'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wagon = self.get_object()
        
        # Получаем все движения для этого вагона
        movements = Movement.objects.filter(wagon=wagon).order_by('-date')
        
        # Пересчитываем текущее количество
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        current_quantity = total_in - total_out
        
        # Обновляем вагон с точным количеством
        if wagon.current_quantity != current_quantity:
            wagon.current_quantity = current_quantity
            wagon.save()
        
        context.update({
            'movements': movements[:10],
            'total_in': total_in,
            'total_out': total_out,
            'net_qty': current_quantity
        })
        return context

class MovementListView(ListView):
    model = Movement
    template_name = 'warehouse/movement_list.html'
    context_object_name = 'movements'

    def get_queryset(self):
        queryset = super().get_queryset()
        self.filter = MovementFilter(self.request.GET, queryset=queryset)
        return self.filter.qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filter
        return context

class MovementCreateView(CreateView):
    model = Movement
    form_class = MovementForm
    template_name = 'warehouse/movement_create.html'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def post(self, request, *args, **kwargs):
        try:
            # For AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                form = self.get_form()
                
                if form.is_valid():
                    movement = form.save(commit=False)
                    
                    # Process transports if included
                    transports_json = request.POST.get('transports_json')
                    if transports_json:
                        try:
                            transports_data = json.loads(transports_json)
                            
                            # Validate that at least one transport with quantity exists
                            if not transports_data:
                                return JsonResponse({
                                    'success': False, 
                                    'error': {'quantity': 'At least one transport with quantity is required'}
                                })
                            
                            # Calculate total quantity from transports
                            total_quantity = 0
                            total_doc_ton = 0  # Добавляем подсчет общего doc_ton
                            
                            for transport in transports_data:
                                transport_quantity = float(transport.get('quantity', 0))
                                transport_doc_ton = float(transport.get('doc_ton', 0) or 0)  # Получаем doc_ton из каждого транспорта
                                
                                if transport_quantity <= 0:
                                    return JsonResponse({
                                        'success': False, 
                                        'error': {'quantity': 'Quantity is required and must be greater than 0'}
                                    })
                                total_quantity += transport_quantity
                                total_doc_ton += transport_doc_ton  # Суммируем doc_ton
                            
                            # Set the movement quantity and doc_ton
                            movement.quantity = total_quantity
                            movement.doc_ton = total_doc_ton
                            
                            movement.save()  
                            for transport_data in transports_data:
                                Transport.objects.create(
                                    movement=movement,
                                    transport_number=transport_data.get('transport_number', ''),
                                    density=transport_data.get('density') or None,
                                    temperature=transport_data.get('temperature', 20),
                                    liter=transport_data.get('liter') or None,
                                    quantity=float(transport_data.get('quantity', 0)),  # Ensure it's a float
                                    doc_ton=transport_data.get('doc_ton') or None,
                                    warehouse_id=transport_data.get('warehouse') or None
                                )
                        except json.JSONDecodeError:
                            return JsonResponse({'success': False, 'error': 'Invalid transport data format'})
                    else:
                        # If no transports, just save the movement
                        movement.save()
                    
                    # После сохранения движения, обновляем запасы
                    if movement.movement_type == 'in':
                        # Если это поступление, увеличиваем запас
                        inventory, created = Inventory.objects.get_or_create(
                            product=movement.product,
                            defaults={'quantity': 0}
                        )
                        inventory.quantity += movement.quantity
                        inventory.save()
                    elif movement.movement_type == 'out':
                        # Если это расход, уменьшаем запас
                        inventory, created = Inventory.objects.get_or_create(
                            product=movement.product,
                            defaults={'quantity': 0}
                        )
                        inventory.quantity -= movement.quantity
                        inventory.save()
                    
                    return JsonResponse({
                        'success': True, 
                        'redirect': reverse('warehouse:movement_list')
                    })
                else:
                    # Return form errors
                    errors = {}
                    for field, error_list in form.errors.items():
                        errors[field] = str(error_list[0])
                    
                    return JsonResponse({'success': False, 'error': errors})
            
            # For regular form submissions (fallback)
            return super().post(request, *args, **kwargs)
        
        except Exception as e:
            # Add more detailed error logging
            import traceback
            print(f"Error in MovementCreateView: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({'success': False, 'error': str(e)})

    def get_success_url(self):
        return reverse('warehouse:movement_list')


                

class InventoryListView(ListView):
    model = Inventory
    template_name = 'warehouse/inventory_list.html'
    context_object_name = 'inventory'

class BatchListView(ListView):
    model = Batch
    template_name = 'warehouse/batch_list.html'
    context_object_name = 'batches'
    ordering = ['-id']

class BatchDetailView(DetailView):
    model = Batch
    template_name = 'warehouse/batch_detail.html'
    context_object_name = 'batch'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        batch = self.get_object()
        movements = Movement.objects.filter(batch=batch)
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        net_qty = total_in - total_out
        context.update({
            'movements': movements,
            'total_in': total_in,
            'total_out': total_out,
            'net_qty': net_qty,
        })
        return context

class BatchCreateView(CreateView):
    model = Batch
    form_class = BatchForm
    template_name = 'warehouse/batch_form.html'

    def get_success_url(self):
        return reverse('warehouse:batch_list')

class BatchUpdateView(UpdateView):
    model = Batch
    form_class = BatchForm
    template_name = 'warehouse/batch_form.html'

    def get_success_url(self):
        return reverse('warehouse:batch_list')

def export_products_excel(request):
    # Создаем буфер в памяти для записи Excel-файла
    output = io.BytesIO()
    
    # Получаем все продукты с предварительной загрузкой связанных данных
    products = Product.objects.all().order_by('code')
    
    # Создаем словарь для хранения данных перед экспортом
    data = []
    
    # Заполняем данные о продуктах с актуальными запасами
    for product in products:
        # Используем динамические методы для вычисления текущих запасов
        total_in = product.get_total_in()
        total_out = product.get_total_out()
        current_stock = total_in - total_out
        
        # Добавляем данные в общий список - убраны лишние поля
        data.append({
            'Код': product.code,
            'Наименование': product.name,
            'Категория': product.category or '',
            'Цена (USD)': product.price_usd or 0,
            'Цена (сум)': product.price_sum or 0,
            'Поступления': total_in,
            'Расходы': total_out,
            'Текущий запас': current_stock,
            'Статус': 'Ниже мин. уровня' if current_stock < product.min_stock else 'В норме',
        })
    
    # Создаем DataFrame из данных
    df = pd.DataFrame(data)
    
    # Создаем writer для записи в Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Записываем данные в Excel
        df.to_excel(writer, sheet_name='Продукты', index=False)
        
        # Автоматическая настройка ширины столбцов
        worksheet = writer.sheets['Продукты']
        for idx, col in enumerate(df.columns):
            column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
            column_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[column_letter].width = column_width
        
        # Добавление форматирования
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Применение стилей к заголовкам
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Добавление условного форматирования для статуса запасов
        status_column = df.columns.get_loc('Статус') + 1
        for i, status in enumerate(df['Статус'], start=2):
            cell = worksheet.cell(row=i, column=status_column)
            if status == 'Ниже мин. уровня':
                cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')
    
    # Сбрасываем указатель в начало
    output.seek(0)
    
    # Создаем HTTP-ответ с Excel-файлом
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products_inventory_report.xlsx"'
    return response


def export_movements_excel(request):
    qs = Movement.objects.select_related('product', 'wagon').all().values(
        'document_number', 'product__name', 'wagon__wagon_number',
        'date', 'movement_type', 'quantity', 'price_sum', 'note'
    )
    df = pd.DataFrame(list(qs))
    df.rename(columns={
        'document_number': 'Hujjat #',
        'product__name': 'Mahsulot',
        'wagon__wagon_number': 'Vagon raqami',
        'date': 'Sana',
        'movement_type': 'Harakat turi',
        'quantity': 'Miqdor',
        'price_sum': "Narx (so'm)",
        'note': 'Izoh',
    }, inplace=True)

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Movements')

    writer.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="movements.xlsx"'
    return response


def export_wagons_excel(request):
    qs = Wagon.objects.all().values(
        'wagon_number', 'wagon_type', 'net_weight', 'meter_weight',
        'capacity', 'volume', 'price_sum', 'condition'
    )
    df = pd.DataFrame(list(qs))
    df.rename(columns={
        'wagon_number': 'Vagon raqami',
        'wagon_type': 'Vagon turi',
        'net_weight': "Netto (kg)",
        'meter_weight': "Meter (kg)",
        'capacity': "Sig'im (tonna)",
        'volume': "Hajmi (L)",
        'price_sum': "Umumiy summa (so'm)",
        'condition': "Holati",
    }, inplace=True)

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Wagons')

    writer.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="wagons.xlsx"'
    return response

class ReservoirListView(ListView):
    model = Reservoir
    template_name = 'warehouse/reservoir_list.html'
    context_object_name = 'reservoirs'
    ordering = ['name']

    def get_queryset(self):
        qs = super().get_queryset().select_related('product', 'warehouse')
        self.filter = ReservoirFilter(self.request.GET, queryset=qs)
        
        # Пересчитываем количество для каждого резервуара
        for reservoir in self.filter.qs:
            movements = Movement.objects.filter(reservoir=reservoir)
            total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
            total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
            current_quantity = total_in - total_out
            
            # Обновляем резервуар, если текущее количество отличается
            if reservoir.current_quantity != current_quantity:
                reservoir.current_quantity = current_quantity
                reservoir.save()
        
        return self.filter.qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filter
        return context

class ReservoirDetailView(DetailView):
    model = Reservoir
    template_name = 'warehouse/reservoir_detail.html'
    context_object_name = 'reservoir'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        reservoir = self.get_object()
        
        # Получаем все движения для этого резервуара
        movements = Movement.objects.filter(reservoir=reservoir).order_by('-date')
        
        # Пересчитываем текущее количество, чтобы убедиться, что оно точное
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        current_quantity = total_in - total_out
        
        # Обновляем резервуар с точным количеством
        if reservoir.current_quantity != current_quantity:
            reservoir.current_quantity = current_quantity
            reservoir.save()
        
        context.update({
            'movements': movements[:10],  # Показываем только последние 10 движений
            'total_in': total_in,
            'total_out': total_out,
            'net_qty': current_quantity
        })
        return context


class ReservoirCreateView(CreateView):
    model = Reservoir
    form_class = ReservoirForm
    template_name = 'warehouse/reservoir_form.html'

    def get_success_url(self):
        return reverse('warehouse:reservoir_list')

class ReservoirUpdateView(UpdateView):
    model = Reservoir
    form_class = ReservoirForm
    template_name = 'warehouse/reservoir_form.html'

    def get_success_url(self):
        return reverse('warehouse:reservoir_list')

class ReservoirMovementListView(ListView):
    model = ReservoirMovement
    template_name = 'warehouse/reservoir_movement_list.html'
    context_object_name = 'movements'
    ordering = ['-date']

    def get_queryset(self):
        qs = super().get_queryset().select_related('reservoir', 'product')
        self.filter = ReservoirMovementFilter(self.request.GET, queryset=qs)
        return self.filter.qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filter
        return context

class ReservoirMovementCreateView(CreateView):
    model = ReservoirMovement
    form_class = ReservoirMovementForm
    template_name = 'warehouse/reservoir_movement_form.html'

    def get_success_url(self):
        return reverse('warehouse:reservoir_movement_list')
    
class LocalClientListView(ListView):
    model = LocalClient
    template_name = 'warehouse/localclient_list.html'
    context_object_name = 'clients'

class ClientListView(ListView):
    model = Client
    template_name = 'warehouse/client_list.html'
    context_object_name = 'clients'

class ClientCreateView(CreateView):
    model = Client
    form_class = ClientForm
    template_name = 'warehouse/client_form.html'
    success_url = reverse_lazy('warehouse:client_list')

class LocalClientCreateView(CreateView):
    model = LocalClient
    form_class = LocalClientForm
    template_name = 'warehouse/localclient_form.html'
    success_url = reverse_lazy('warehouse:localclient_list')

class LocalMovementListView(ListView):
    model = LocalMovement
    template_name = 'warehouse/localmovement_list.html'
    context_object_name = 'local_movements'
    ordering = ['-date']

    def get_queryset(self):
        qs = super().get_queryset()
        self.filter = LocalMovementFilter(self.request.GET, queryset=qs)
        return self.filter.qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filter
        return context


class LocalMovementCreateView(CreateView):
    model = LocalMovement
    form_class = LocalMovementForm
    template_name = 'warehouse/localmovement_form.html'
    success_url = reverse_lazy('warehouse:localmovement_list')

    def form_valid(self, form):
        form.instance.mass = form.instance.density * form.instance.liter
        return super().form_valid(form)

class PlacementListView(ListView):
    model = Placement
    template_name = 'warehouse/placement_list.html'
    context_object_name = 'placements'
    ordering = ['-created_at']

class PlacementCreateView(CreateView):
    model = Placement
    form_class = PlacementForm
    template_name = 'warehouse/placement_form.html'

    def get_success_url(self):
        return reverse('warehouse:placement_list')

class PlacementUpdateView(UpdateView):
    model = Placement
    form_class = PlacementForm
    template_name = 'warehouse/placement_form.html'

    def get_success_url(self):
        return reverse('warehouse:placement_list')

class WagonTypeListView(ListView):
    model = WagonType
    template_name = 'warehouse/wagon_type_list.html'
    context_object_name = 'wagon_types'

class WagonTypeCreateView(CreateView):
    model = WagonType
    form_class = WagonTypeForm
    template_name = 'warehouse/wagon_type_form.html'
    
    def get_success_url(self):
        return reverse('warehouse:wagon_type_list')

class WagonTypeUpdateView(UpdateView):
    model = WagonType
    form_class = WagonTypeForm
    template_name = 'warehouse/wagon_type_form.html'
    
    def get_success_url(self):
        return reverse('warehouse:wagon_type_list')

class MovementUpdateView(UpdateView):
    model = Movement
    form_class = MovementForm
    template_name = 'warehouse/movement_create.html'  # Reuse the create template
    
    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_success_url(self):
        return reverse('warehouse:movement_list')

@require_POST
@permission_required('warehouse.delete_movement')
def movement_delete(request, pk):
    try:
        movement = Movement.objects.get(pk=pk)
        movement_number = movement.document_number
        product = movement.product
        quantity = movement.quantity
        movement_type = movement.movement_type
        reservoir = movement.reservoir
        wagon = movement.wagon
        
        # Удаляем связанные транспорты
        Transport.objects.filter(movement=movement).delete()
        
        # Обновляем запасы в резервуаре или вагоне перед удалением
        if reservoir:
            if movement_type == 'in':
                reservoir.current_quantity -= quantity
            elif movement_type == 'out':
                reservoir.current_quantity += quantity
            reservoir.save()
        
        if wagon:
            if movement_type == 'in':
                wagon.current_quantity -= quantity
            elif movement_type == 'out':
                wagon.current_quantity += quantity
            wagon.save()
        
        # Удаляем движение
        movement.delete()
        
        # Пересчитываем запасы продукта
        recalculate_inventory(product)
        
        return JsonResponse({
            'success': True,
            'message': f'Движение {movement_number} успешно удалено'
        })
    except Movement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Движение не найдено'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def recalculate_product_inventory_view(request, pk):
    """Представление для ручного пересчета запасов продукта и перенаправления обратно."""
    product = get_object_or_404(Product, pk=pk)
    recalculate_inventory(product)
    return redirect('warehouse:product_detail', pk=pk)

def recalculate_all_inventories(request):
    """Пересчитывает запасы для всех продуктов."""
    products = Product.objects.all()
    updated_count = 0
    
    for product in products:
        recalculate_inventory(product)
        updated_count += 1
    
    messages.success(request, f"Запасы успешно пересчитаны для {updated_count} продуктов.")
    return redirect('warehouse:product_list')

def recalculate_all_reservoirs(request):
    """Пересчитывает запасы для всех резервуаров."""
    reservoirs = Reservoir.objects.all()
    updated_count = 0
    
    for reservoir in reservoirs:
        movements = Movement.objects.filter(reservoir=reservoir)
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        current_quantity = total_in - total_out
        
        if reservoir.current_quantity != current_quantity:
            reservoir.current_quantity = current_quantity
            reservoir.save()
            updated_count += 1
    
    messages.success(request, f"Запасы успешно пересчитаны для {updated_count} резервуаров.")
    return redirect('warehouse:reservoir_list')

def recalculate_all_wagons(request):
    """Пересчитывает запасы для всех вагонов."""
    wagons = Wagon.objects.all()
    updated_count = 0
    
    for wagon in wagons:
        movements = Movement.objects.filter(wagon=wagon)
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        current_quantity = total_in - total_out
        
        if wagon.current_quantity != current_quantity:
            wagon.current_quantity = current_quantity
            wagon.save()
            updated_count += 1
    
    messages.success(request, f"Запасы успешно пересчитаны для {updated_count} вагонов.")
    return redirect('warehouse:wagon_list')

def export_reservoirs_excel(request):
    output = io.BytesIO()
    
    reservoirs = Reservoir.objects.all().select_related('product', 'warehouse')
    
    data = []
    for reservoir in reservoirs:
        # Пересчитываем текущее количество
        movements = Movement.objects.filter(reservoir=reservoir)
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        current_quantity = total_in - total_out
        
        # Обновляем резервуар если необходимо
        if reservoir.current_quantity != current_quantity:
            reservoir.current_quantity = current_quantity
            reservoir.save()
            
        data.append({
            'Название': reservoir.name,
            'Склад': reservoir.warehouse.name if reservoir.warehouse else '',
            'Емкость': reservoir.capacity,
            'Продукт': reservoir.product.name if reservoir.product else 'Не указан',
            'Текущее количество': current_quantity,
            'Заполнено (%)': round((current_quantity / reservoir.capacity) * 100, 2) if reservoir.capacity else 0,
        })
    
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Резервуары', index=False)
        
        # Форматирование и стили
        worksheet = writer.sheets['Резервуары']
        for idx, col in enumerate(df.columns):
            column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
            column_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[column_letter].width = column_width
        
        # Применение стилей
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
    
    output.seek(0)
    
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reservoirs_report.xlsx"'
    return response

# Представления для Estokada (заводских сотрудников)
# @estokada_required
# def estokada_dashboard(request):
#     """
#     Панель управления для сотрудников эстокады с 4 основными кнопками:
#     1. Приёмка
#     2. Продажа
#     3. Производство
#     4. Перемещение
#     """
#     context = {}
#     return render(request, 'warehouse/estokada/dashboard.html', context)

# @estokada_required
# def estokada_reception_list(request):
#     """Страница с ожидающими приемками на текущий день"""
#     today = timezone.now().date()
    
#     # Получаем все операции приемки на сегодня в статусе 'created'
#     receptions = Movement.objects.filter(
#         movement_type='in', 
#         date=today,
#         status='created'
#     ).order_by('-created_at')
    
#     context = {
#         'receptions': receptions
#     }
    
#     return render(request, 'warehouse/estokada/reception_list.html', context)

# @estokada_required
# def estokada_reception_process(request, pk):
#     """Страница обработки конкретной приемки"""
#     movement = get_object_or_404(Movement, pk=pk, movement_type='in')
    
#     if request.method == 'POST':
#         # Заполняем форму только избранными полями, которые может заполнять сотрудник эстокады
#         form = MovementForm(
#             request.POST, 
#             request.FILES, 
#             instance=movement, 
#             user=request.user
#         )
        
#         # Частичная валидация только тех полей, которые относятся к физическим параметрам
#         if form.is_valid():
#             # Обновляем только определенные поля
#             movement.density = form.cleaned_data['density']
#             movement.temperature = form.cleaned_data['temperature']
#             movement.liter = form.cleaned_data['liter']
#             movement.quantity = form.cleaned_data['quantity']
#             movement.transport_photo = form.cleaned_data.get('transport_photo')
#             movement.note = form.cleaned_data.get('note', movement.note)
#             movement.status = 'processed'  # Изменяем статус на "обработано"
            
#             movement.save()
            
#             messages.success(request, f"Приёмка #{movement.id} успешно обработана")
#             return redirect('warehouse:estokada_reception_list')
#     else:
#         form = MovementForm(instance=movement)
    
#     context = {
#         'movement': movement,
#         'form': form,
#         'is_readonly': True  # Некоторые поля будут только для чтения
#     }
    
#     return render(request, 'warehouse/estokada/reception_process.html', context)

# @estokada_required
# def estokada_sales_list(request):
#     """Страница с ожидающими продажами на текущий день"""
#     today = timezone.now().date()
    
#     # Получаем все операции продажи на сегодня в статусе 'created'
#     sales = Movement.objects.filter(
#         movement_type='out', 
#         date=today,
#         status='created'
#     ).order_by('-created_at')
    
#     context = {
#         'sales': sales
#     }
    
#     return render(request, 'warehouse/estokada/sales_list.html', context)

# @estokada_required
# def estokada_sales_process(request, pk):
#     """Страница обработки конкретной продажи с поддержкой множественных транспортов"""
#     movement = get_object_or_404(Movement, pk=pk, movement_type='out')
    
#     if request.method == 'POST':
#         form = MovementForm(
#             request.POST, 
#             request.FILES, 
#             instance=movement, 
#             user=request.user
#         )
        
#         if form.is_valid():
#             # Получаем данные о транспорте из JSON
#             transports_json = request.POST.get('transports_json', '[]')
            
#             try:
#                 transports_data = json.loads(transports_json)
                
#                 # Проверяем, что есть хотя бы один транспорт с количеством
#                 if not transports_data:
#                     form.add_error(None, 'Необходимо добавить хотя бы один транспорт с данными.')
#                     context = {
#                         'movement': movement,
#                         'form': form,
#                         'available_quantity': get_available_quantity(movement),
#                         'is_readonly': True
#                     }
#                     return render(request, 'warehouse/estokada/sales_process.html', context)
                
#                 # Рассчитываем общие значения из всех транспортов
#                 total_quantity = 0
#                 total_liter = 0
#                 avg_density = 0
#                 avg_temperature = 0
#                 valid_transports = 0
                
#                 for transport_data in transports_data:
#                     quantity = float(transport_data.get('quantity', 0))
#                     total_quantity += quantity
                    
#                     if 'liter' in transport_data and transport_data['liter']:
#                         total_liter += float(transport_data['liter'])
                    
#                     if 'density' in transport_data and transport_data['density'] and quantity > 0:
#                         avg_density += float(transport_data['density'])
#                         valid_transports += 1
                    
#                     if 'temperature' in transport_data and transport_data['temperature'] and quantity > 0:
#                         avg_temperature += float(transport_data['temperature'])
                
#                 # Вычисляем средние значения
#                 if valid_transports > 0:
#                     avg_density = avg_density / valid_transports
#                     avg_temperature = avg_temperature / valid_transports
#                 else:
#                     avg_density = 0
#                     avg_temperature = 20
                
#                 # Обновляем основные поля движения
#                 movement.density = avg_density
#                 movement.temperature = avg_temperature
#                 movement.liter = total_liter
#                 movement.quantity = total_quantity
#                 movement.note = form.cleaned_data.get('note', movement.note)
#                 movement.status = 'processed'
                
#                 # Сохраняем обновленное движение
#                 movement.save()
                
#                 # Удаляем старые записи транспорта, если они есть
#                 Transport.objects.filter(movement=movement).delete()
                
#                 # Создаем новые записи для каждого транспорта
#                 for transport_data in transports_data:
#                     Transport.objects.create(
#                         movement=movement,
#                         transport_number=transport_data.get('transport_number', ''),
#                         density=float(transport_data.get('density', 0)) if transport_data.get('density') else None,
#                         temperature=float(transport_data.get('temperature', 20)),
#                         liter=float(transport_data.get('liter', 0)) if transport_data.get('liter') else None,
#                         quantity=float(transport_data.get('quantity', 0))
#                     )
                
#                 messages.success(request, f"Продажа #{movement.id} успешно обработана")
#                 return redirect('warehouse:estokada_sales_list')
                
#             except json.JSONDecodeError:
#                 form.add_error(None, 'Ошибка в данных о транспорте. Проверьте заполнение формы.')
#                 context = {
#                     'movement': movement,
#                     'form': form,
#                     'available_quantity': get_available_quantity(movement),
#                     'is_readonly': True
#                 }
#                 return render(request, 'warehouse/estokada/sales_process.html', context)
#         else:
#             # Если форма невалидна
#             context = {
#                 'movement': movement,
#                 'form': form,
#                 'available_quantity': get_available_quantity(movement),
#                 'is_readonly': True
#             }
#             return render(request, 'warehouse/estokada/sales_process.html', context)
#     else:
#         form = MovementForm(instance=movement)
    
#     # Получаем доступное количество в источнике
#     available_quantity = get_available_quantity(movement)
    
#     context = {
#         'movement': movement,
#         'form': form,
#         'available_quantity': available_quantity,
#         'is_readonly': True  # Некоторые поля будут только для чтения
#     }
    
#     return render(request, 'warehouse/estokada/sales_process.html', context)

# def get_available_quantity(movement):
#     """Вспомогательная функция для получения доступного количества продукта"""
#     available_quantity = 0
#     if movement.source_reservoir:
#         available_quantity = movement.source_reservoir.current_quantity
#     elif movement.source_wagon:
#         available_quantity = movement.source_wagon.current_quantity
#     elif movement.source_warehouse:
#         # Проверяем наличие продукта на складе
#         inventories = Inventory.objects.filter(
#             warehouse=movement.source_warehouse,
#             product=movement.product
#         )
#         if inventories.exists():
#             available_quantity = inventories.first().quantity
    
#     return available_quantity

# @estokada_required
# def estokada_production_list(request):
#     """Страница с ожидающими производственными операциями на текущий день"""
#     today = timezone.now().date()
    
#     # Получаем все производственные операции на сегодня в статусе 'created'
#     productions = Movement.objects.filter(
#         movement_type='production', 
#         date=today,
#         status='created'
#     ).order_by('-created_at')
    
#     context = {
#         'productions': productions
#     }
    
#     return render(request, 'warehouse/estokada/production_list.html', context)

# @estokada_required
# def estokada_production_process(request, pk):
#     """Страница обработки конкретного производственного процесса"""
#     movement = get_object_or_404(Movement, pk=pk, movement_type='production')
    
#     if request.method == 'POST':
#         form = MovementForm(
#             request.POST, 
#             request.FILES, 
#             instance=movement, 
#             user=request.user
#         )
        
#         if form.is_valid():
#             # Обновляем только определенные поля
#             movement.density = form.cleaned_data['density']
#             movement.temperature = form.cleaned_data['temperature']
#             movement.liter = form.cleaned_data['liter']
#             movement.quantity = form.cleaned_data['quantity']
#             movement.production_loss = form.cleaned_data.get('production_loss', 0)
            
#             # Если потери превышают норму, требуется причина
#             expected_loss = movement.expected_quantity * 0.02  # 2% от ожидаемого количества
#             if movement.production_loss > expected_loss:
#                 production_loss_reason = form.cleaned_data.get('production_loss_reason')
#                 if not production_loss_reason:
#                     form.add_error('production_loss_reason', 'Требуется указать причину превышения допустимых потерь')
#                     context = {
#                         'movement': movement,
#                         'form': form,
#                         'is_readonly': True
#                     }
#                     return render(request, 'warehouse/estokada/production_process.html', context)
                
#                 movement.production_loss_reason = production_loss_reason
            
#             movement.status = 'processed'  # Изменяем статус на "обработано"
#             movement.save()
            
#             messages.success(request, f"Производственный процесс #{movement.id} успешно обработан")
#             return redirect('warehouse:estokada_production_list')
#     else:
#         form = MovementForm(instance=movement)
    
#     context = {
#         'movement': movement,
#         'form': form,
#         'is_readonly': True  # Некоторые поля будут только для чтения
#     }
    
#     return render(request, 'warehouse/estokada/production_process.html', context)

# @estokada_required
# def estokada_transfer_list(request):
#     """Страница с ожидающими перемещениями на текущий день"""
#     today = timezone.now().date()
    
#     # Получаем все перемещения на сегодня в статусе 'created'
#     transfers = Movement.objects.filter(
#         movement_type='transfer', 
#         date=today,
#         status='created'
#     ).order_by('-created_at')
    
#     context = {
#         'transfers': transfers
#     }
    
#     return render(request, 'warehouse/estokada/transfer_list.html', context)

# @estokada_required
# def estokada_transfer_process(request, pk):
#     """Страница обработки конкретного перемещения"""
#     movement = get_object_or_404(Movement, pk=pk, movement_type='transfer')
    
#     if request.method == 'POST':
#         form = MovementForm(
#             request.POST, 
#             request.FILES, 
#             instance=movement, 
#             user=request.user
#         )
        
#         if form.is_valid():
#             # Обновляем только определенные поля
#             movement.density = form.cleaned_data['density']
#             movement.temperature = form.cleaned_data['temperature']
#             movement.liter = form.cleaned_data['liter']
#             movement.quantity = form.cleaned_data['quantity']
#             movement.note = form.cleaned_data.get('note', movement.note)
#             movement.status = 'processed'  # Изменяем статус на "обработано"
            
#             movement.save()
            
#             messages.success(request, f"Перемещение #{movement.id} успешно обработано")
#             return redirect('warehouse:estokada_transfer_list')
#     else:
#         form = MovementForm(instance=movement)
    
#     # Проверка доступного количества в источнике
#     available_quantity = 0
#     if movement.source_reservoir:
#         available_quantity = movement.source_reservoir.current_quantity
#     elif movement.source_wagon:
#         available_quantity = movement.source_wagon.current_quantity
    
#     context = {
#         'movement': movement,
#         'form': form,
#         'available_quantity': available_quantity,
#         'is_readonly': True  # Некоторые поля будут только для чтения
#     }
    
#     return render(request, 'warehouse/estokada/transfer_process.html', context)

# Представления для SalesDepartment (офисных сотрудников)
# @sales_required
# def sales_department_dashboard(request):
#     """Панель управления для сотрудников отдела продаж"""
#     # Получаем статистику по количеству операций разных типов
#     today = timezone.now().date()
#     month_start = today.replace(day=1)
    
#     # Статистика за текущий день
#     today_receptions = Movement.objects.filter(movement_type='in', date=today).count()
#     today_sales = Movement.objects.filter(movement_type='out', date=today).count()
#     today_productions = Movement.objects.filter(movement_type='production', date=today).count()
#     today_transfers = Movement.objects.filter(movement_type='transfer', date=today).count()
    
#     # Статистика за текущий месяц
#     month_receptions = Movement.objects.filter(movement_type='in', date__gte=month_start).count()
#     month_sales = Movement.objects.filter(movement_type='out', date__gte=month_start).count()
#     month_productions = Movement.objects.filter(movement_type='production', date__gte=month_start).count()
#     month_transfers = Movement.objects.filter(movement_type='transfer', date__gte=month_start).count()
    
#     # Последние операции
#     recent_movements = Movement.objects.all().order_by('-created_at')[:10]
    
#     context = {
#         'today_receptions': today_receptions,
#         'today_sales': today_sales,
#         'today_productions': today_productions,
#         'today_transfers': today_transfers,
#         'month_receptions': month_receptions,
#         'month_sales': month_sales,
#         'month_productions': month_productions,
#         'month_transfers': month_transfers,
#         'recent_movements': recent_movements,
#     }
    
#     return render(request, 'warehouse/sales/dashboard.html', context)

# @sales_required
# def sales_movement_create(request):
#     """Создание новой операции движения с поддержкой множественных транспортов и источников сырья"""
#     if request.method == 'POST':
#         form = MovementForm(request.POST, request.FILES, user=request.user, department='sales')
        
#         if form.is_valid():
#             movement = form.save(commit=False)
            
#             # Общая обработка для всех типов операций
#             transports_json = request.POST.get('transports_json', '[]')
            
#             try:
#                 # Парсим данные о транспортах из JSON
#                 transports_data = json.loads(transports_json)
                
#                 # Для приемки и продажи: вычисляем общее количество и средние значения
#                 if movement.movement_type in ['in', 'out']:
#                     if not transports_data:
#                         form.add_error(None, 'Необходимо добавить хотя бы один транспорт с данными.')
#                         context = {
#                             'form': form,
#                             'title': 'Создание новой операции'
#                         }
#                         return render(request, 'warehouse/sales/movement_form.html', context)
                    
#                     # Рассчитываем общие значения из всех транспортов
#                     total_quantity = 0
#                     total_doc_mass = 0
                    
#                     for transport_data in transports_data:
#                         mass = float(transport_data.get('mass', 0) or 0)
#                         doc_mass = float(transport_data.get('doc_mass', 0) or 0)
                        
#                         total_quantity += mass
#                         total_doc_mass += doc_mass
                    
#                     # Переводим кг в тонны для сохранения в БД
#                     movement.quantity = total_quantity / 1000  # кг в тонны
#                     movement.expected_quantity = total_quantity / 1000  # для отслеживания ожидаемого количества
#                     movement.doc_ton = total_doc_mass / 1000  # документальный вес в тоннах
                    
#                     # Разница между фактическим и документальным весом
#                     movement.difference = movement.quantity - movement.doc_ton
                
#                 # Для производства: обработка данных источника и продукта
#                 elif movement.movement_type == 'production':
#                     source_quantity = float(request.POST.get('source_quantity', 0) or 0)
#                     material_quantity = float(request.POST.get('material_quantity', 0) or 0)
                    
#                     # Сохраняем вес продукта (кг в тонны)
#                     movement.quantity = material_quantity / 1000
#                     movement.expected_quantity = material_quantity / 1000
                    
#                     # Сохраняем вес источника (кг в тонны)
#                     movement.source_quantity = source_quantity / 1000
                
#                 # Сохраняем движение для получения ID
#                 movement.save()
                
#                 # Обработка транспортных данных и создание записей (для in/out)
#                 if movement.movement_type in ['in', 'out']:
#                     for transport_data in transports_data:
#                         transport_type = transport_data.get('transport_type', '')
#                         transport_number = transport_data.get('transport_number', '')
#                         mass_kg = float(transport_data.get('mass', 0) or 0)
#                         doc_mass_kg = float(transport_data.get('doc_mass', 0) or 0)
                        
#                         # Создаем запись транспорта
#                         transport = Transport.objects.create(
#                             movement=movement,
#                             transport_number=transport_number,
#                             quantity=mass_kg,  # В кг
#                             doc_ton=doc_mass_kg / 1000  # Преобразуем кг в тонны
#                         )
                        
#                         # Сохраняем дополнительные данные для вагонов
#                         if transport_type == 'wagon' and 'wagon_type' in transport_data:
#                             wagon_type_id = transport_data.get('wagon_type')
#                             if wagon_type_id:
#                                 try:
#                                     # Здесь можно добавить связь с типом вагона, если нужно
#                                     pass
#                                 except Exception as e:
#                                     print(f"Ошибка при обработке типа вагона: {e}")
                
#             except json.JSONDecodeError:
#                 form.add_error(None, 'Ошибка в данных о транспорте. Проверьте заполнение формы.')
#                 context = {
#                     'form': form,
#                     'title': 'Создание новой операции'
#                 }
#                 return render(request, 'warehouse/sales/movement_form.html', context)
            
#             # Перенаправление на соответствующую страницу в зависимости от типа движения
#             if movement.movement_type == 'in':
#                 return redirect('warehouse:sales_reception_list')
#             elif movement.movement_type == 'out':
#                 return redirect('warehouse:sales_sales_list')
#             elif movement.movement_type == 'production':
#                 return redirect('warehouse:sales_production_list')
#             elif movement.movement_type == 'transfer':
#                 return redirect('warehouse:sales_transfer_list')
            
#             return redirect('warehouse:sales_department_dashboard')
#     else:
#         form = MovementForm(user=request.user, department='sales')
    
#     context = {
#         'form': form,
#         'title': 'Создание новой операции'
#     }
    
#     return render(request, 'warehouse/sales/movement_form.html', context)

# @sales_required
# def sales_reception_list(request):
#     """Список операций приемки для отдела продаж"""
#     receptions = Movement.objects.filter(movement_type='in').order_by('-date', '-created_at')
    
#     context = {
#         'receptions': receptions,
#         'title': 'Список приемок'
#     }
    
#     return render(request, 'warehouse/sales/reception_list.html', context)

# @sales_required
# def sales_sales_list(request):
#     """Список операций продажи для отдела продаж"""
#     sales = Movement.objects.filter(movement_type='out').order_by('-date', '-created_at')
    
#     context = {
#         'sales': sales,
#         'title': 'Список продаж'
#     }
    
#     return render(request, 'warehouse/sales/sales_list.html', context)

# @sales_required
# def sales_production_list(request):
#     """Список производственных операций для отдела продаж"""
#     productions = Movement.objects.filter(movement_type='production').order_by('-date', '-created_at')
    
#     context = {
#         'productions': productions,
#         'title': 'Список производственных операций'
#     }
    
#     return render(request, 'warehouse/sales/production_list.html', context)

# @sales_required
# def sales_transfer_list(request):
#     """Список операций перемещения для отдела продаж"""
#     transfers = Movement.objects.filter(movement_type='transfer').order_by('-date', '-created_at')
    
#     context = {
#         'transfers': transfers,
#         'title': 'Список перемещений'
#     }
    
#     return render(request, 'warehouse/sales/transfer_list.html', context)

# @sales_required
# def sales_movement_detail(request, pk):
#     """Детальная информация о движении"""
#     movement = get_object_or_404(Movement, pk=pk)
    
#     context = {
#         'movement': movement,
#         'title': f'Детали {movement.get_movement_type_display()} #{movement.id}'
#     }
    
#     return render(request, 'warehouse/sales/movement_detail.html', context)

# @sales_required
# def sales_movement_edit(request, pk):
#     """Редактирование операции движения"""
#     movement = get_object_or_404(Movement, pk=pk)
    
#     # Проверка статуса - нельзя редактировать подтвержденные или отмененные операции
#     if movement.status in ['confirmed', 'cancelled']:
#         messages.error(request, "Невозможно редактировать подтвержденную или отмененную операцию")
#         return redirect('warehouse:sales_movement_detail', pk=movement.id)
    
#     if request.method == 'POST':
#         form = MovementForm(request.POST, request.FILES, instance=movement, user=request.user, department='sales')
        
#         if form.is_valid():
#             updated_movement = form.save(commit=False)
            
#             # Обработка транспортных данных
#             try:
#                 transports_json = request.POST.get('transports_json', '[]')
#                 transports_data = json.loads(transports_json)
                
#                 if movement.movement_type in ['in', 'out'] and not transports_data:
#                     form.add_error(None, 'Необходимо добавить хотя бы один транспорт')
#                     context = {
#                         'form': form,
#                         'movement': movement,
#                         'title': f'Редактирование {movement.get_movement_type_display()}'
#                     }
#                     return render(request, 'warehouse/sales/movement_form.html', context)
                
#                 # Для приемки и продажи пересчитываем общие значения
#                 if movement.movement_type in ['in', 'out']:
#                     total_quantity = 0
#                     total_doc_mass = 0
                    
#                     for transport_data in transports_data:
#                         mass = float(transport_data.get('mass', 0) or 0)
#                         doc_mass = float(transport_data.get('doc_mass', 0) or 0)
                        
#                         total_quantity += mass
#                         total_doc_mass += doc_mass
                    
#                     # Обновляем поля движения (кг в тонны)
#                     updated_movement.quantity = total_quantity / 1000
#                     updated_movement.expected_quantity = total_quantity / 1000
#                     updated_movement.doc_ton = total_doc_mass / 1000
#                     updated_movement.difference = updated_movement.quantity - updated_movement.doc_ton
                
#                 # Сохраняем обновленные данные движения
#                 updated_movement.save()
                
#                 # Удаляем существующие записи транспорта
#                 Transport.objects.filter(movement=movement).delete()
                
#                 # Создаем новые записи транспорта
#                 for transport_data in transports_data:
#                     transport_type = transport_data.get('transport_type', '')
#                     transport_number = transport_data.get('transport_number', '')
#                     mass_kg = float(transport_data.get('mass', 0) or 0)
#                     doc_mass_kg = float(transport_data.get('doc_mass', 0) or 0)
                    
#                     Transport.objects.create(
#                         movement=movement,
#                         transport_number=transport_number,
#                         quantity=mass_kg,  # В кг
#                         doc_ton=doc_mass_kg / 1000  # Преобразуем кг в тонны
#                     )
                    
#                 messages.success(request, f"{movement.get_movement_type_display()} #{movement.id} успешно обновлена")
#                 return redirect('warehouse:sales_movement_detail', pk=movement.id)
                
#             except json.JSONDecodeError:
#                 form.add_error(None, 'Ошибка в данных о транспорте')
#                 context = {
#                     'form': form,
#                     'movement': movement,
#                     'title': f'Редактирование {movement.get_movement_type_display()}'
#                 }
#                 return render(request, 'warehouse/sales/movement_form.html', context)
#         else:
#             # Если форма невалидна
#             context = {
#                 'form': form,
#                 'movement': movement,
#                 'title': f'Редактирование {movement.get_movement_type_display()}'
#             }
#             return render(request, 'warehouse/sales/movement_form.html', context)
#     else:
#         form = MovementForm(instance=movement, user=request.user, department='sales')
    
#     context = {
#         'form': form,
#         'movement': movement,
#         'title': f'Редактирование {movement.get_movement_type_display()}'
#     }
    
#     return render(request, 'warehouse/sales/movement_form.html', context)

# @sales_required
# def sales_movement_confirm(request, pk):
#     """Подтверждение операции после обработки специалистом эстокады"""
#     movement = get_object_or_404(Movement, pk=pk)
    
#     # Подтверждать можно только обработанные операции
#     if movement.status != 'processed':
#         messages.error(request, f"Операция #{movement.id} не может быть подтверждена, так как не обработана")
#         return redirect('warehouse:sales_movement_detail', pk=movement.pk)
    
#     if request.method == 'POST':
#         # Обновление статуса операции на 'completed'
#         movement.status = 'completed'
#         movement.confirmed_by = request.user
#         movement.save()
        
#         messages.success(request, f"{movement.get_movement_type_display()} #{movement.id} успешно подтверждена и завершена")
#         return redirect('warehouse:sales_movement_detail', pk=movement.pk)
    
#     context = {
#         'movement': movement,
#         'title': f'Подтверждение {movement.get_movement_type_display()} #{movement.id}'
#     }
    
#     return render(request, 'warehouse/sales/movement_confirm.html', context)

# @sales_required
# def sales_movement_cancel(request, pk):
#     """Отмена операции"""
#     movement = get_object_or_404(Movement, pk=pk)
    
#     # Нельзя отменить завершенную операцию
#     if movement.status == 'completed':
#         messages.error(request, f"Операция #{movement.id} уже завершена и не может быть отменена")
#         return redirect('warehouse:sales_movement_detail', pk=movement.pk)
    
#     if request.method == 'POST':
#         cancel_reason = request.POST.get('cancel_reason')
        
#         if not cancel_reason:
#             messages.error(request, "Необходимо указать причину отмены")
#             return redirect('warehouse:sales_movement_cancel', pk=movement.pk)
        
#         # Обновление статуса операции на 'cancelled'
#         movement.status = 'cancelled'
#         movement.note = f"{movement.note or ''}\n\nПричина отмены: {cancel_reason}"
#         movement.save()
        
#         messages.success(request, f"{movement.get_movement_type_display()} #{movement.id} отменена")
#         return redirect('warehouse:sales_movement_detail', pk=movement.pk)
    
#     context = {
#         'movement': movement,
#         'title': f'Отмена {movement.get_movement_type_display()} #{movement.id}'
#     }
    
#     return render(request, 'warehouse/sales/movement_cancel.html', context)

class StatisticsView(LoginRequiredMixin, TemplateView):
    """Представление для отображения статистики"""
    template_name = 'warehouse/statistics.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем параметры из запроса
        period = self.request.GET.get('period', 'day')
        movement_type = self.request.GET.get('movement_type', '')
        selected_product = self.request.GET.get('product', '')
        
        # Определяем период
        today = timezone.now().date()
        if period == 'day':
            start_date = today
        elif period == 'week':
            start_date = today - timedelta(days=7)
        elif period == 'month':
            start_date = today - timedelta(days=30)
        elif period == 'year':
            start_date = today - timedelta(days=365)
        else:
            start_date = today
        
        # Базовый запрос
        movements = Movement.objects.filter(date__gte=start_date, date__lte=today)
        
        # Применяем фильтры
        if movement_type:
            movements = movements.filter(movement_type=movement_type)
        if selected_product:
            movements = movements.filter(product_id=selected_product)
        
        # Статистика по всем типам операций
        operations_stats = {
            'reception': movements.filter(movement_type='in').count(),
            'sale': movements.filter(movement_type='out').count(),
            'transfer': movements.filter(movement_type='transfer').count(),
            'production': movements.filter(movement_type='production').count()
        }
        
        # Объемы по типам операций
        volume_stats = {
            'reception': movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0,
            'sale': movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0,
            'transfer': movements.filter(movement_type='transfer').aggregate(total=Sum('quantity'))['total'] or 0,
            'production': movements.filter(movement_type='production').aggregate(total=Sum('quantity'))['total'] or 0
        }
        
        # Статистика по транспорту
        transport_data = Transport.objects.filter(movement__in=movements)
        transport_stats = {
            'truck': transport_data.filter(transport_type='truck').count(),
            'wagon': transport_data.filter(transport_type='wagon').count(),
            'truck_volume': transport_data.filter(transport_type='truck').aggregate(
                total=Sum('quantity'))['total'] or 0,
            'wagon_volume': transport_data.filter(transport_type='wagon').aggregate(
                total=Sum('quantity'))['total'] or 0
        }
        
        # Динамика по дням
        date_range = [(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
        date_range.reverse()  # От старых к новым
        
        daily_data = []
        for date_str in date_range:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            day_movements = movements.filter(date=date_obj)
            
            day_data = {
                'date': date_str,
                'reception': day_movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0,
                'sale': day_movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0,
                'transfer': day_movements.filter(movement_type='transfer').aggregate(total=Sum('quantity'))['total'] or 0,
                'production': day_movements.filter(movement_type='production').aggregate(total=Sum('quantity'))['total'] or 0
            }
            daily_data.append(day_data)
        
        # Топ продуктов
        products = Product.objects.all()
        product_stats = []
        
        for product in products:
            product_movements = movements.filter(product=product)
            total_in = product_movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
            total_out = product_movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
            
            stats = {
                'id': product.id,
                'name': product.name,
                'quantity': total_in - total_out,
                'operations_count': product_movements.count(),
                'last_movement': product_movements.order_by('-date').first()
            }
            
            product_stats.append(stats)
        
        # Сортируем по количеству операций и берем топ 10
        product_stats.sort(key=lambda x: x['operations_count'], reverse=True)
        top_products = product_stats[:10]
        
        # Подсчёт общих показателей
        total_operations = movements.count()
        total_quantity = movements.aggregate(total=Sum('quantity'))['total'] or 0
        
        # Считаем прибыль/убыток (разница между продажами и приемкой)
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        profit_loss = total_in - total_out
        
        # Средняя разница между документальным и фактическим весом
        transports = Transport.objects.filter(movement__in=movements)
        difference_sum = transports.aggregate(total=Sum('difference'))['total'] or 0
        transports_count = transports.count()
        average_difference = difference_sum / transports_count if transports_count > 0 else 0
        
        context.update({
            'period': period,
            'movement_type': movement_type,
            'selected_product': selected_product,
            'products': Product.objects.all(),
            'total_operations': total_operations,
            'total_quantity': total_quantity,
            'profit_loss': profit_loss,
            'average_difference': average_difference,
            'operations_stats': operations_stats,
            'volume_stats': volume_stats,
            'transport_stats': transport_stats,
            'daily_data': daily_data,
            'top_products': top_products,
            'chart_dates': [day['date'] for day in daily_data],
            'chart_reception': [day['reception'] for day in daily_data],
            'chart_sale': [day['sale'] for day in daily_data],
            'chart_transfer': [day['transfer'] for day in daily_data],
            'chart_production': [day['production'] for day in daily_data],
        })
        
        return context

class AnalyticsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'warehouse/analytics/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем данные за последние 30 дней
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
        
        analytics_data = AnalyticsData.objects.filter(
            date__range=[start_date, end_date]
        ).order_by('date')
        
        # Подготовка данных для графиков
        dates = [data.date.strftime('%Y-%m-%d') for data in analytics_data]
        received = [float(data.total_received) for data in analytics_data]
        shipped = [float(data.total_shipped) for data in analytics_data]
        produced = [float(data.total_produced) for data in analytics_data]
        transferred = [float(data.total_transferred) for data in analytics_data]
        
        # График движения товаров
        movement_fig = go.Figure()
        movement_fig.add_trace(go.Scatter(x=dates, y=received, name='Приход'))
        movement_fig.add_trace(go.Scatter(x=dates, y=shipped, name='Отгрузка'))
        movement_fig.add_trace(go.Scatter(x=dates, y=produced, name='Производство'))
        movement_fig.add_trace(go.Scatter(x=dates, y=transferred, name='Перемещение'))
        
        # Настройка внешнего вида графика
        movement_fig.update_layout(
            title='Динамика движения товаров',
            xaxis_title='Дата',
            yaxis_title='Количество (тонн)',
            legend=dict(
                yanchor="top",
                y=0.99,
                xanchor="left",
                x=0.01
            ),
            hovermode="x unified"
        )
        
        context['movement_chart'] = movement_fig.to_html()
        
        # Статистика
        context['total_received'] = sum(received)
        context['total_shipped'] = sum(shipped)
        context['total_produced'] = sum(produced)
        context['total_transferred'] = sum(transferred)
        
        # Список последних отчетов
        context['reports'] = AnalyticsReport.objects.all().order_by('-created_at')[:5]
        
        return context

class GenerateReportView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        report_type = request.POST.get('report_type')
        date_from = datetime.strptime(request.POST.get('date_from'), '%Y-%m-%d').date()
        date_to = datetime.strptime(request.POST.get('date_to'), '%Y-%m-%d').date()
        
        # Получаем данные для отчета
        data = AnalyticsData.objects.filter(
            date__range=[date_from, date_to]
        ).order_by('date')
        
        # Создаем DataFrame
        df = pd.DataFrame(list(data.values()))
        
        # Создаем Excel файл
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename=report_{date_from}_{date_to}.xlsx'
        
        df.to_excel(response, index=False)
        
        # Сохраняем отчет в базе
        AnalyticsReport.objects.create(
            report_type=report_type,
            date_from=date_from,
            date_to=date_to,
            file=response
        )
        
        return response

class UpdateAnalyticsDataView(View):
    def post(self, request, *args, **kwargs):
        date = datetime.now().date()
        
        # Обновляем аналитические данные
        for product in Product.objects.all():
            analytics_data, created = AnalyticsData.objects.get_or_create(
                date=date,
                product=product
            )
            
            # Получаем данные о движениях за день
            movements = Movement.objects.filter(
                product=product,
                created_at__date=date
            )
            
            analytics_data.total_received = movements.filter(
                type='reception'
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            analytics_data.total_shipped = movements.filter(
                type='shipment'
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            analytics_data.total_produced = movements.filter(
                type='production'
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            analytics_data.total_transferred = movements.filter(
                type='transfer'
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            analytics_data.save()
        
        return JsonResponse({'status': 'success'})

# Представления для инвентаризации (периодического аудита)
class InventoryAuditListView(LoginRequiredMixin, ListView):
    model = InventoryAudit
    template_name = 'warehouse/inventory/audit_list.html'
    context_object_name = 'audits'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем статистику по инвентаризациям
        context['audits_completed'] = InventoryAudit.objects.filter(status='completed').count()
        context['audits_in_progress'] = InventoryAudit.objects.filter(status='in_progress').count()
        context['audits_planned'] = InventoryAudit.objects.filter(status='planned').count()
        
        # Последняя завершенная инвентаризация
        last_audit = InventoryAudit.objects.filter(status='completed').order_by('-end_date').first()
        if last_audit:
            context['last_audit'] = last_audit
            context['last_audit_discrepancies'] = last_audit.discrepancy_items()
        
        return context


class InventoryAuditCreateView(LoginRequiredMixin, CreateView):
    model = InventoryAudit
    form_class = InventoryAuditForm
    template_name = 'warehouse/inventory/audit_form.html'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.status = 'planned'
        response = super().form_valid(form)
        
        # Получаем текущие остатки для всех продуктов на всех местах хранения
        # и создаем элементы аудита
        
        # 1. Товары на складах
        inventories = Inventory.objects.all()
        for inventory in inventories:
            if inventory.quantity > 0:
                InventoryAuditItem.objects.create(
                    audit=self.object,
                    product=inventory.product,
                    location=f"Склад: {inventory.warehouse.name}",
                    expected_quantity=inventory.quantity
                )
        
        # 2. Товары в резервуарах
        reservoirs = Reservoir.objects.all()
        for reservoir in reservoirs:
            if reservoir.current_volume > 0 and reservoir.product:
                InventoryAuditItem.objects.create(
                    audit=self.object,
                    product=reservoir.product,
                    location=f"Резервуар: {reservoir.name}",
                    expected_quantity=reservoir.current_volume
                )
        
        # 3. Товары в вагонах
        wagons = Wagon.objects.filter(status='loaded')
        for wagon in wagons:
            if wagon.current_volume > 0 and wagon.product:
                InventoryAuditItem.objects.create(
                    audit=self.object,
                    product=wagon.product,
                    location=f"Вагон: {wagon.wagon_number}",
                    expected_quantity=wagon.current_volume
                )
        
        messages.success(self.request, f'Инвентаризация "{self.object.name}" создана. Добавлено {self.object.items.count()} позиций.')
        return response
    
    def get_success_url(self):
        return reverse('warehouse:audit_detail', kwargs={'pk': self.object.pk})


class InventoryAuditDetailView(LoginRequiredMixin, DetailView):
    model = InventoryAudit
    template_name = 'warehouse/inventory/audit_detail.html'
    context_object_name = 'audit'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        audit = self.get_object()
        
        # Добавляем элементы аудита, сгруппированные по местоположению
        items = audit.items.all().order_by('location', 'product__name')
        locations = items.values_list('location', flat=True).distinct()
        
        items_by_location = {}
        for location in locations:
            items_by_location[location] = items.filter(location=location)
        
        context['items_by_location'] = items_by_location
        
        # Общая статистика
        total_expected = items.aggregate(total=Sum('expected_quantity'))['total'] or 0
        total_actual = items.filter(actual_quantity__isnull=False).aggregate(total=Sum('actual_quantity'))['total'] or 0
        
        context['total_expected'] = total_expected
        context['total_actual'] = total_actual
        context['total_checked'] = items.filter(actual_quantity__isnull=False).count()
        context['total_items'] = items.count()
        
        # Проверяем, все ли элементы проверены
        context['all_checked'] = context['total_checked'] == context['total_items']
        
        # Расхождения
        if audit.status == 'completed':
            discrepancies = audit.calculate_discrepancies()
            context['discrepancies'] = discrepancies
            
            # Строим диаграмму расхождений
            if discrepancies:
                products = [d['item'].product.name for d in discrepancies]
                percentages = [float(d['percentage']) for d in discrepancies]
                
                fig = px.bar(
                    x=products, 
                    y=percentages,
                    labels={'x': 'Продукт', 'y': 'Расхождение (%)'},
                    title='Расхождения по продуктам (%)'
                )
                context['discrepancy_chart'] = fig.to_html()
        
        return context


class InventoryAuditUpdateView(LoginRequiredMixin, UpdateView):
    model = InventoryAudit
    form_class = InventoryAuditForm
    template_name = 'warehouse/inventory/audit_form.html'
    
    def get_success_url(self):
        return reverse('warehouse:audit_detail', kwargs={'pk': self.object.pk})


class InventoryAuditStartView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        audit = get_object_or_404(InventoryAudit, pk=kwargs['pk'])
        
        if audit.status != 'planned':
            messages.error(request, 'Инвентаризация не может быть запущена, так как она не находится в статусе "Запланирована"')
            return redirect('warehouse:audit_detail', pk=audit.pk)
        
        audit.status = 'in_progress'
        audit.save()
        
        messages.success(request, f'Инвентаризация "{audit.name}" запущена')
        return redirect('warehouse:audit_detail', pk=audit.pk)


class InventoryAuditCompleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        audit = get_object_or_404(InventoryAudit, pk=kwargs['pk'])
        
        if audit.status != 'in_progress':
            messages.error(request, 'Инвентаризация не может быть завершена, так как она не находится в статусе "В процессе"')
            return redirect('warehouse:audit_detail', pk=audit.pk)
        
        # Проверяем, все ли элементы проверены
        unchecked_items = audit.items.filter(actual_quantity__isnull=True)
        if unchecked_items.exists():
            messages.error(request, f'Инвентаризация не может быть завершена, так как не все элементы проверены. Осталось проверить: {unchecked_items.count()}')
            return redirect('warehouse:audit_detail', pk=audit.pk)
        
        audit.status = 'completed'
        audit.end_date = timezone.now()
        audit.save()
        
        messages.success(request, f'Инвентаризация "{audit.name}" завершена')
        return redirect('warehouse:audit_detail', pk=audit.pk)


class InventoryAuditCancelView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        audit = get_object_or_404(InventoryAudit, pk=kwargs['pk'])
        
        if audit.status == 'completed':
            messages.error(request, 'Инвентаризация не может быть отменена, так как она уже завершена')
            return redirect('warehouse:audit_detail', pk=audit.pk)
        
        audit.status = 'cancelled'
        audit.save()
        
        messages.success(request, f'Инвентаризация "{audit.name}" отменена')
        return redirect('warehouse:audit_detail', pk=audit.pk)


class InventoryAuditItemUpdateView(LoginRequiredMixin, UpdateView):
    model = InventoryAuditItem
    form_class = InventoryAuditItemForm
    template_name = 'warehouse/inventory/audit_item_form.html'
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Позиция "{self.object}" обновлена')
        return response
    
    def get_success_url(self):
        return reverse('warehouse:audit_detail', kwargs={'pk': self.object.audit.pk})


class InventoryAuditReportView(LoginRequiredMixin, DetailView):
    model = InventoryAudit
    template_name = 'warehouse/inventory/audit_report.html'
    context_object_name = 'audit'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        audit = self.get_object()
        
        if audit.status != 'completed':
            messages.error(self.request, 'Отчет доступен только для завершенных инвентаризаций')
            return context
        
        # Расхождения
        discrepancies = audit.calculate_discrepancies()
        context['discrepancies'] = discrepancies
        
        # Статистика
        items = audit.items.all()
        total_expected = items.aggregate(total=Sum('expected_quantity'))['total'] or 0
        total_actual = items.aggregate(total=Sum('actual_quantity'))['total'] or 0
        
        context['total_expected'] = total_expected
        context['total_actual'] = total_actual
        context['total_discrepancy'] = total_actual - total_expected
        context['total_discrepancy_percent'] = (context['total_discrepancy'] / total_expected * 100) if total_expected > 0 else 0
        
        # Диаграммы
        if discrepancies:
            # Диаграмма расхождений по продуктам
            products = [d['item'].product.name for d in discrepancies]
            actual = [float(d['item'].actual_quantity) for d in discrepancies]
            expected = [float(d['item'].expected_quantity) for d in discrepancies]
            
            fig = go.Figure()
            fig.add_trace(go.Bar(name='Фактическое количество', x=products, y=actual))
            fig.add_trace(go.Bar(name='Учетное количество', x=products, y=expected))
            
            fig.update_layout(
                title='Сравнение фактического и учетного количества',
                xaxis_title='Продукт',
                yaxis_title='Количество',
                barmode='group'
            )
            
            context['comparison_chart'] = fig.to_html()
            
            # Круговая диаграмма расхождений по местам хранения
            location_discrepancies = {}
            for d in discrepancies:
                location = d['item'].location
                if location not in location_discrepancies:
                    location_discrepancies[location] = 0
                location_discrepancies[location] += abs(float(d['discrepancy']))
            
            locations = list(location_discrepancies.keys())
            discrepancy_values = list(location_discrepancies.values())
            
            pie_fig = px.pie(
                values=discrepancy_values,
                names=locations,
                title='Распределение расхождений по местам хранения'
            )
            
            context['location_chart'] = pie_fig.to_html()
        
        return context


def export_audit_to_excel(request, pk):
    audit = get_object_or_404(InventoryAudit, pk=pk)
    
    if audit.status != 'completed':
        messages.error(request, 'Экспорт доступен только для завершенных инвентаризаций')
        return redirect('warehouse:audit_detail', pk=audit.pk)
    
    # Создаем DataFrame из элементов аудита
    items = audit.items.all()
    data = []
    
    for item in items:
        discrepancy = item.discrepancy() if item.actual_quantity is not None else None
        discrepancy_percent = item.discrepancy_percentage() if item.actual_quantity is not None else None
        
        data.append({
            'Продукт': item.product.name,
            'Местоположение': item.location,
            'Учетное количество': float(item.expected_quantity),
            'Фактическое количество': float(item.actual_quantity) if item.actual_quantity is not None else None,
            'Расхождение': float(discrepancy) if discrepancy is not None else None,
            'Расхождение (%)': float(discrepancy_percent) if discrepancy_percent is not None else None,
            'Примечания': item.notes
        })
    
    df = pd.DataFrame(data)
    
    # Создаем Excel файл
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Инвентаризация', index=False)
    
    # Получаем объект workbook и worksheet
    workbook = writer.book
    worksheet = writer.sheets['Инвентаризация']
    
    # Форматы для выделения ячеек
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#D9EAD3',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    negative_format = workbook.add_format({
        'bg_color': '#F4CCCC',
        'num_format': '0.00'
    })
    
    positive_format = workbook.add_format({
        'bg_color': '#D9EAD3',
        'num_format': '0.00'
    })
    
    # Применяем форматы
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    
    # Форматирование расхождений
    for row_num, row in enumerate(df.values):
        if pd.notna(row[4]):  # Расхождение
            if row[4] < 0:
                worksheet.write(row_num + 1, 4, row[4], negative_format)
                worksheet.write(row_num + 1, 5, row[5], negative_format)
            elif row[4] > 0:
                worksheet.write(row_num + 1, 4, row[4], positive_format)
                worksheet.write(row_num + 1, 5, row[5], positive_format)
    
    # Устанавливаем ширину столбцов
    worksheet.set_column('A:A', 30)
    worksheet.set_column('B:B', 30)
    worksheet.set_column('C:F', 15)
    worksheet.set_column('G:G', 40)
    
    # Добавляем информацию об инвентаризации в начало файла
    summary_sheet = workbook.add_worksheet('Общая информация')
    
    summary_sheet.write('A1', 'Инвентаризация:', header_format)
    summary_sheet.write('B1', audit.name)
    summary_sheet.write('A2', 'Дата начала:', header_format)
    summary_sheet.write('B2', audit.start_date.strftime('%d.%m.%Y %H:%M'))
    summary_sheet.write('A3', 'Дата завершения:', header_format)
    summary_sheet.write('B3', audit.end_date.strftime('%d.%m.%Y %H:%M') if audit.end_date else '')
    summary_sheet.write('A4', 'Статус:', header_format)
    summary_sheet.write('B4', audit.get_status_display())
    summary_sheet.write('A5', 'Создал:', header_format)
    summary_sheet.write('B5', audit.created_by.username)
    
    summary_sheet.write('A7', 'Общее количество позиций:', header_format)
    summary_sheet.write('B7', audit.items.count())
    
    summary_sheet.write('A8', 'Позиции с расхождениями:', header_format)
    summary_sheet.write('B8', audit.discrepancy_items())
    
    # Общая статистика
    total_expected = audit.items.aggregate(total=Sum('expected_quantity'))['total'] or 0
    total_actual = audit.items.aggregate(total=Sum('actual_quantity'))['total'] or 0
    total_discrepancy = total_actual - total_expected
    total_discrepancy_percent = (total_discrepancy / total_expected * 100) if total_expected > 0 else 0
    
    summary_sheet.write('A10', 'Общее учетное количество:', header_format)
    summary_sheet.write('B10', float(total_expected))
    summary_sheet.write('A11', 'Общее фактическое количество:', header_format)
    summary_sheet.write('B11', float(total_actual))
    summary_sheet.write('A12', 'Общее расхождение:', header_format)
    summary_sheet.write('B12', float(total_discrepancy))
    summary_sheet.write('A13', 'Общее расхождение (%):', header_format)
    summary_sheet.write('B13', float(total_discrepancy_percent))
    
    summary_sheet.set_column('A:A', 30)
    summary_sheet.set_column('B:B', 20)
    
    writer.save()
    
    # Создаем HTTP-ответ с Excel файлом
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=audit_{audit.id}_{audit.name}.xlsx'
    
    return response

# Представления для минимальных уровней запасов и уведомлений
class ProductMinLevelListView(LoginRequiredMixin, ListView):
    model = ProductMinLevel
    template_name = 'warehouse/inventory/min_level_list.html'
    context_object_name = 'min_levels'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем список продуктов без установленных минимальных уровней
        configured_products = ProductMinLevel.objects.values_list('product_id', flat=True)
        context['unconfigured_products'] = Product.objects.exclude(id__in=configured_products)
        
        # Добавляем уведомления о низком уровне запасов
        below_min_levels = []
        for min_level in context['min_levels']:
            if min_level.is_below_min() and min_level.alert_enabled:
                below_min_levels.append({
                    'product': min_level.product,
                    'current': min_level.product.current_quantity(),
                    'min_level': min_level.min_quantity,
                    'percentage': min_level.get_percentage()
                })
        
        context['below_min_levels'] = below_min_levels
        
        # График текущих уровней запасов относительно минимальных
        products = [ml.product.name for ml in context['min_levels']]
        current_quantities = [float(ml.product.current_quantity()) for ml in context['min_levels']]
        min_quantities = [float(ml.min_quantity) for ml in context['min_levels']]
        optimal_quantities = [float(ml.optimal_quantity) for ml in context['min_levels']]
        
        if products:
            fig = go.Figure()
            fig.add_trace(go.Bar(name='Текущее количество', x=products, y=current_quantities))
            fig.add_trace(go.Bar(name='Минимальный уровень', x=products, y=min_quantities))
            fig.add_trace(go.Bar(name='Оптимальный уровень', x=products, y=optimal_quantities))
            
            fig.update_layout(
                title='Сравнение текущего, минимального и оптимального уровня запасов',
                xaxis_title='Продукт',
                yaxis_title='Количество',
                barmode='group'
            )
            
            context['inventory_chart'] = fig.to_html()
        
        return context


class ProductMinLevelCreateView(LoginRequiredMixin, CreateView):
    model = ProductMinLevel
    form_class = ProductMinLevelForm
    template_name = 'warehouse/inventory/min_level_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = get_object_or_404(Product, pk=self.kwargs['product_id'])
        return context
    
    def form_valid(self, form):
        form.instance.product = get_object_or_404(Product, pk=self.kwargs['product_id'])
        response = super().form_valid(form)
        messages.success(self.request, f'Минимальный уровень для {form.instance.product.name} установлен')
        return response
    
    def get_success_url(self):
        return reverse('warehouse:min_level_list')


class ProductMinLevelUpdateView(LoginRequiredMixin, UpdateView):
    model = ProductMinLevel
    form_class = ProductMinLevelForm
    template_name = 'warehouse/inventory/min_level_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.object.product
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Минимальный уровень для {form.instance.product.name} обновлен')
        return response
    
    def get_success_url(self):
        return reverse('warehouse:min_level_list')


class ProductMinLevelDeleteView(LoginRequiredMixin, DeleteView):
    model = ProductMinLevel
    template_name = 'warehouse/inventory/min_level_confirm_delete.html'
    success_url = reverse_lazy('warehouse:min_level_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, f'Минимальный уровень для {self.get_object().product.name} удален')
        return super().delete(request, *args, **kwargs)


# Представления для прогнозирования и планирования запасов
class StockForecastView(LoginRequiredMixin, FormView):
    form_class = GenerateForecastForm
    template_name = 'warehouse/inventory/forecast_form.html'
    
    def form_valid(self, form):
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        products = form.cleaned_data['products']
        
        if not products:
            products = Product.objects.all()
        
        # Очищаем существующие прогнозы за указанный период
        StockForecast.objects.filter(
            date__range=[start_date, end_date],
            product__in=products
        ).delete()
        
        forecasts_created = 0
        
        for product in products:
            # Получаем исторические данные о движениях продукта
            movements = Movement.objects.filter(
                product=product,
                created_at__lt=start_date,
                status='completed'
            ).order_by('created_at')
            
            if movements.count() < 10:
                continue  # Недостаточно данных для прогнозирования
            
            # Агрегируем данные по дням
            daily_data = {}
            for movement in movements:
                day = movement.created_at.date()
                if day not in daily_data:
                    daily_data[day] = 0
                
                if movement.type in ['shipment', 'transfer']:
                    daily_data[day] += float(movement.quantity)
            
            if not daily_data:
                continue
            
            # Создаем временной ряд для анализа
            dates = sorted(daily_data.keys())
            quantities = [daily_data[date] for date in dates]
            
            # Проверяем, что у нас есть непрерывный временной ряд
            all_dates = []
            current_date = dates[0]
            while current_date <= dates[-1]:
                all_dates.append(current_date)
                current_date += timedelta(days=1)
            
            # Заполняем пропущенные даты нулями
            filled_quantities = []
            for date in all_dates:
                if date in daily_data:
                    filled_quantities.append(daily_data[date])
                else:
                    filled_quantities.append(0)
            
            # Простое прогнозирование на основе скользящего среднего
            # и экспоненциального сглаживания
            try:
                # Рассчитываем средний расход за неделю, месяц
                week_avg = np.mean(filled_quantities[-7:])
                month_avg = np.mean(filled_quantities[-30:]) if len(filled_quantities) >= 30 else week_avg
                
                # Для предсказания используем экспоненциальное сглаживание
                alpha = 0.3  # Параметр сглаживания
                smoothed = filled_quantities[-1]
                
                current_date = start_date
                day_count = (end_date - start_date).days + 1
                
                for i in range(day_count):
                    # Используем взвешенную комбинацию недельного и месячного среднего
                    # с добавлением тренда и сезонности
                    forecast = 0.7 * week_avg + 0.3 * month_avg
                    
                    # Добавляем коррекцию на день недели
                    day_of_week = current_date.weekday()
                    if day_of_week == 5 or day_of_week == 6:  # Выходные
                        forecast *= 0.5
                    
                    # Сглаживаем прогноз
                    smoothed = alpha * forecast + (1 - alpha) * smoothed
                    
                    # Создаем запись прогноза
                    StockForecast.objects.create(
                        product=product,
                        date=current_date,
                        forecasted_quantity=smoothed,
                        confidence_level=70.0  # Примерный уровень уверенности
                    )
                    
                    forecasts_created += 1
                    current_date += timedelta(days=1)
            
            except Exception as e:
                messages.error(self.request, f'Ошибка при прогнозировании для {product.name}: {str(e)}')
        
        messages.success(self.request, f'Создано {forecasts_created} прогнозов для {len(products)} продуктов')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('warehouse:forecast_list')


class StockForecastListView(LoginRequiredMixin, ListView):
    model = StockForecast
    template_name = 'warehouse/inventory/forecast_list.html'
    context_object_name = 'forecasts'
    
    def get_queryset(self):
        # Группируем прогнозы по продуктам, отображаем только последние даты
        products = Product.objects.all()
        latest_forecasts = []
        
        for product in products:
            forecast = StockForecast.objects.filter(product=product).order_by('-date').first()
            if forecast:
                latest_forecasts.append(forecast)
        
        return latest_forecasts
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем график прогнозов для первых 5 продуктов
        top_products = context['forecasts'][:5]
        
        if top_products:
            # Получаем прогнозы на ближайшие 30 дней
            today = date.today()
            end_date = today + timedelta(days=30)
            
            product_data = {}
            for product in [forecast.product for forecast in top_products]:
                forecasts = StockForecast.objects.filter(
                    product=product,
                    date__range=[today, end_date]
                ).order_by('date')
                
                dates = [forecast.date.strftime('%Y-%m-%d') for forecast in forecasts]
                quantities = [float(forecast.forecasted_quantity) for forecast in forecasts]
                
                product_data[product.name] = {
                    'dates': dates,
                    'quantities': quantities
                }
            
            # Создаем график прогнозов
            fig = go.Figure()
            
            for product_name, data in product_data.items():
                fig.add_trace(go.Scatter(
                    x=data['dates'],
                    y=data['quantities'],
                    mode='lines+markers',
                    name=product_name
                ))
            
            fig.update_layout(
                title='Прогноз расхода продуктов на следующие 30 дней',
                xaxis_title='Дата',
                yaxis_title='Прогнозируемый расход'
            )
            
            context['forecast_chart'] = fig.to_html()
        
        return context


class ProductForecastDetailView(LoginRequiredMixin, DetailView):
    model = Product
    template_name = 'warehouse/inventory/product_forecast_detail.html'
    context_object_name = 'product'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.get_object()
        
        # Получаем прогнозы на ближайшие 90 дней
        today = date.today()
        end_date = today + timedelta(days=90)
        
        forecasts = StockForecast.objects.filter(
            product=product,
            date__range=[today, end_date]
        ).order_by('date')
        
        context['forecasts'] = forecasts
        
        if forecasts:
            # Рассчитываем прогнозируемые остатки
            current_quantity = product.current_quantity()
            
            dates = [forecast.date.strftime('%Y-%m-%d') for forecast in forecasts]
            daily_usage = [float(forecast.forecasted_quantity) for forecast in forecasts]
            
            # Прогноз оставшегося количества
            remaining = [current_quantity]
            for usage in daily_usage:
                remaining.append(max(0, remaining[-1] - usage))
            remaining = remaining[1:]  # Удаляем первое значение (текущее количество)
            
            # Прогноз исчерпания запаса
            try:
                depletion_index = next(i for i, r in enumerate(remaining) if r == 0)
                depletion_date = dates[depletion_index]
                context['depletion_date'] = depletion_date
                context['days_to_depletion'] = depletion_index + 1
            except (StopIteration, IndexError):
                context['depletion_date'] = None
                context['days_to_depletion'] = "90+"
            
            # Минимальный уровень, если установлен
            try:
                min_level = product.min_level
                min_quantity = float(min_level.min_quantity)
                optimal_quantity = float(min_level.optimal_quantity)
                
                context['min_level'] = min_level
                
                try:
                    # Дата, когда запас упадет ниже минимального уровня
                    min_level_index = next(i for i, r in enumerate(remaining) if r < min_quantity)
                    min_level_date = dates[min_level_index]
                    context['min_level_date'] = min_level_date
                    context['days_to_min_level'] = min_level_index + 1
                except (StopIteration, IndexError):
                    context['min_level_date'] = None
                    context['days_to_min_level'] = "90+"
                
                # Рассчитываем рекомендуемый объем закупки
                if context['days_to_min_level'] != "90+":
                    # Закупка для пополнения до оптимального уровня
                    recommended_purchase = optimal_quantity - remaining[min_level_index]
                    context['recommended_purchase'] = recommended_purchase
            except ProductMinLevel.DoesNotExist:
                context['min_level'] = None
            
            # График прогноза
            fig = go.Figure()
            
            # Текущее количество
            fig.add_trace(go.Scatter(
                x=[today.strftime('%Y-%m-%d')] + dates,
                y=[current_quantity] + remaining,
                mode='lines',
                name='Прогнозируемый остаток'
            ))
            
            # График ежедневного расхода
            fig.add_trace(go.Bar(
                x=dates,
                y=daily_usage,
                name='Прогнозируемый расход',
                yaxis='y2'
            ))
            
            # Добавляем минимальный уровень, если установлен
            if 'min_level' in context and context['min_level']:
                fig.add_trace(go.Scatter(
                    x=[today.strftime('%Y-%m-%d')] + dates,
                    y=[min_quantity] * (len(dates) + 1),
                    mode='lines',
                    name='Минимальный уровень',
                    line=dict(color='red', dash='dash')
                ))
                
                fig.add_trace(go.Scatter(
                    x=[today.strftime('%Y-%m-%d')] + dates,
                    y=[optimal_quantity] * (len(dates) + 1),
                    mode='lines',
                    name='Оптимальный уровень',
                    line=dict(color='green', dash='dash')
                ))
            
            fig.update_layout(
                title=f'Прогноз запасов для {product.name}',
                xaxis_title='Дата',
                yaxis_title='Остаток',
                yaxis2=dict(
                    title='Ежедневный расход',
                    overlaying='y',
                    side='right'
                ),
                hovermode='x unified'
            )
            
            context['forecast_chart'] = fig.to_html()
        
        return context

# Представления для планирования закупок
class PurchasePlanListView(LoginRequiredMixin, ListView):
    model = PurchasePlan
    template_name = 'warehouse/inventory/purchase_plan_list.html'
    context_object_name = 'plans'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем счетчики планов по статусам
        context['draft_plans'] = PurchasePlan.objects.filter(status='draft').count()
        context['approved_plans'] = PurchasePlan.objects.filter(status='approved').count()
        context['in_progress_plans'] = PurchasePlan.objects.filter(status='in_progress').count()
        context['completed_plans'] = PurchasePlan.objects.filter(status='completed').count()
        
        # Добавляем продукты, требующие закупки (ниже минимального уровня)
        below_min_products = []
        for min_level in ProductMinLevel.objects.all():
            if min_level.is_below_min():
                current = min_level.product.current_quantity()
                below_min_products.append({
                    'product': min_level.product,
                    'current': current,
                    'min_level': min_level.min_quantity,
                    'optimal_level': min_level.optimal_quantity,
                    'to_purchase': min_level.optimal_quantity - current
                })
        
        context['below_min_products'] = below_min_products
        
        return context


class PurchasePlanCreateView(LoginRequiredMixin, CreateView):
    model = PurchasePlan
    form_class = PurchasePlanForm
    template_name = 'warehouse/inventory/purchase_plan_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.POST:
            context['formset'] = PurchasePlanItemFormSet(self.request.POST)
        else:
            context['formset'] = PurchasePlanItemFormSet()
            
            # Автоматически добавляем продукты ниже минимального уровня
            if 'auto_fill' in self.request.GET and self.request.GET['auto_fill'] == '1':
                for min_level in ProductMinLevel.objects.all():
                    if min_level.is_below_min():
                        current = min_level.product.current_quantity()
                        optimal = min_level.optimal_quantity
                        to_purchase = optimal - current
                        
                        if to_purchase > 0:
                            # Добавляем элемент в formset
                            form = context['formset'].forms[0]
                            form.initial = {
                                'product': min_level.product.id,
                                'quantity': to_purchase,
                                'is_automatic': True,
                                'notes': 'Автоматически добавлено системой'
                            }
        
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            form.instance.created_by = self.request.user
            self.object = form.save()
            
            formset.instance = self.object
            formset.save()
            
            messages.success(self.request, f'План закупок "{self.object.name}" создан')
            return redirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))
    
    def get_success_url(self):
        return reverse('warehouse:purchase_plan_detail', kwargs={'pk': self.object.pk})


class PurchasePlanDetailView(LoginRequiredMixin, DetailView):
    model = PurchasePlan
    template_name = 'warehouse/inventory/purchase_plan_detail.html'
    context_object_name = 'plan'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plan = self.get_object()
        
        # Добавляем общую статистику
        total_quantity = plan.total_quantity()
        context['total_quantity'] = total_quantity
        
        # График распределения по продуктам
        items = plan.items.all()
        products = [item.product.name for item in items]
        quantities = [float(item.quantity) for item in items]
        
        if products:
            fig = px.pie(
                values=quantities,
                names=products,
                title='Распределение закупок по продуктам'
            )
            
            context['distribution_chart'] = fig.to_html()
        
        return context


class PurchasePlanUpdateView(LoginRequiredMixin, UpdateView):
    model = PurchasePlan
    form_class = PurchasePlanForm
    template_name = 'warehouse/inventory/purchase_plan_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.POST:
            context['formset'] = PurchasePlanItemFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = PurchasePlanItemFormSet(instance=self.object)
        
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            
            messages.success(self.request, f'План закупок "{self.object.name}" обновлен')
            return redirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))
    
    def get_success_url(self):
        return reverse('warehouse:purchase_plan_detail', kwargs={'pk': self.object.pk})


class PurchasePlanApproveView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        plan = get_object_or_404(PurchasePlan, pk=kwargs['pk'])
        
        if plan.status != 'draft':
            messages.error(request, 'План закупок не может быть утвержден, так как он не находится в статусе "Черновик"')
            return redirect('warehouse:purchase_plan_detail', pk=plan.pk)
        
        plan.status = 'approved'
        plan.approved_by = request.user
        plan.approved_at = timezone.now()
        plan.save()
        
        messages.success(request, f'План закупок "{plan.name}" утвержден')
        return redirect('warehouse:purchase_plan_detail', pk=plan.pk)


class PurchasePlanStartView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        plan = get_object_or_404(PurchasePlan, pk=kwargs['pk'])
        
        if plan.status != 'approved':
            messages.error(request, 'План закупок не может быть запущен, так как он не находится в статусе "Утвержден"')
            return redirect('warehouse:purchase_plan_detail', pk=plan.pk)
        
        plan.status = 'in_progress'
        plan.save()
        
        messages.success(request, f'План закупок "{plan.name}" запущен в работу')
        return redirect('warehouse:purchase_plan_detail', pk=plan.pk)


class PurchasePlanCompleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        plan = get_object_or_404(PurchasePlan, pk=kwargs['pk'])
        
        if plan.status != 'in_progress':
            messages.error(request, 'План закупок не может быть завершен, так как он не находится в статусе "В процессе"')
            return redirect('warehouse:purchase_plan_detail', pk=plan.pk)
        
        plan.status = 'completed'
        plan.save()
        
        messages.success(request, f'План закупок "{plan.name}" завершен')
        return redirect('warehouse:purchase_plan_detail', pk=plan.pk)


class PurchasePlanCancelView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        plan = get_object_or_404(PurchasePlan, pk=kwargs['pk'])
        
        if plan.status in ['completed']:
            messages.error(request, 'План закупок не может быть отменен, так как он уже завершен')
            return redirect('warehouse:purchase_plan_detail', pk=plan.pk)
        
        plan.status = 'cancelled'
        plan.save()
        
        messages.success(request, f'План закупок "{plan.name}" отменен')
        return redirect('warehouse:purchase_plan_detail', pk=plan.pk)


def export_purchase_plan_to_excel(request, pk):
    """Экспорт плана закупок в Excel"""
    plan = get_object_or_404(PurchasePlan, pk=pk)
    # Реализация экспорта Excel
    # Добавьте функционал экспорта Excel здесь
    return HttpResponse("Export functionality will be added soon.")

# Точки заказа и оптимизация закупок

class OrderPointListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Представление для списка точек заказа"""
    model = OrderPoint
    template_name = 'warehouse/purchasing/order_point_list.html'
    context_object_name = 'order_points'
    permission_required = 'warehouse.view_orderpoint'
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('product')
        
        # Фильтрация по продукту
        product_id = self.request.GET.get('product')
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        # Фильтрация по статусу (ниже точки заказа)
        below_reorder = self.request.GET.get('below_reorder')
        if below_reorder:
            below_points = []
            for point in queryset:
                if point.is_below_reorder_point():
                    below_points.append(point.pk)
            queryset = queryset.filter(pk__in=below_points)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем дополнительные данные
        order_points_with_data = []
        for point in context['order_points']:
            current_quantity = point.product.current_quantity()
            optimal_time = point.calculate_optimal_order_time()
            optimal_supplier = point.get_optimal_supplier()
            
            order_points_with_data.append({
                'point': point,
                'current_quantity': current_quantity,
                'optimal_time': optimal_time,
                'optimal_supplier': optimal_supplier,
                'is_below': current_quantity <= point.reorder_point,
                'percentage': (current_quantity / point.reorder_point * 100) if point.reorder_point > 0 else 0
            })
        
        context['order_points_with_data'] = order_points_with_data
        context['products'] = Product.objects.all()
        context['product_filter'] = self.request.GET.get('product', '')
        context['below_reorder_filter'] = self.request.GET.get('below_reorder', '')
        return context


class OrderPointCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Представление для создания точки заказа"""
    model = OrderPoint
    form_class = OrderPointForm
    template_name = 'warehouse/purchasing/order_point_form.html'
    permission_required = 'warehouse.add_orderpoint'
    
    def get_initial(self):
        initial = super().get_initial()
        product_id = self.request.GET.get('product')
        
        if product_id:
            try:
                product = Product.objects.get(pk=product_id)
                initial['product'] = product
                
                # Автоматически предлагаем точку заказа на основе исторических данных
                thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
                consumption = Movement.objects.filter(
                    product=product,
                    movement_type__in=['sale', 'production_usage'],
                    created_at__gte=thirty_days_ago
                ).aggregate(total=models.Sum('quantity'))['total'] or 0
                
                # Среднее дневное потребление
                daily_consumption = consumption / 30
                
                # Получаем среднее время поставки от поставщиков
                suppliers = ProductSupplier.objects.filter(product=product, lead_time__isnull=False)
                avg_lead_time = suppliers.aggregate(avg=models.Avg('lead_time'))['avg'] or 7
                
                # Расчет точки заказа и запаса безопасности
                lead_time_demand = daily_consumption * avg_lead_time
                safety_stock = lead_time_demand * 0.5  # 50% от потребления за время поставки
                reorder_point = lead_time_demand + safety_stock
                
                initial['lead_time_demand'] = lead_time_demand
                initial['safety_stock'] = safety_stock
                initial['reorder_point'] = reorder_point
                initial['order_quantity'] = reorder_point * 2  # Предлагаем заказать в 2 раза больше точки заказа
            except Product.DoesNotExist:
                pass
                
        return initial
    
    def get_success_url(self):
        return reverse('warehouse:order_point_list')


class OrderPointUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Представление для редактирования точки заказа"""
    model = OrderPoint
    form_class = OrderPointForm
    template_name = 'warehouse/purchasing/order_point_form.html'
    permission_required = 'warehouse.change_orderpoint'
    
    def get_success_url(self):
        return reverse('warehouse:order_point_list')


class OrderPointDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Представление для просмотра детальной информации о точке заказа"""
    model = OrderPoint
    template_name = 'warehouse/purchasing/order_point_detail.html'
    context_object_name = 'order_point'
    permission_required = 'warehouse.view_orderpoint'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем текущее количество
        current_quantity = self.object.product.current_quantity()
        context['current_quantity'] = current_quantity
        
        # Рассчитываем оптимальное время заказа
        context['optimal_time'] = self.object.calculate_optimal_order_time()
        
        # Получаем оптимального поставщика
        context['optimal_supplier'] = self.object.get_optimal_supplier()
        
        # Получаем историю потребления
        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
        consumption_history = Movement.objects.filter(
            product=self.object.product,
            movement_type__in=['sale', 'production_usage'],
            created_at__gte=thirty_days_ago
        ).values('created_at__date').annotate(
            total=models.Sum('quantity')
        ).order_by('created_at__date')
        
        context['consumption_history'] = consumption_history
        
        # Получаем поставщиков продукта
        context['suppliers'] = ProductSupplier.objects.filter(
            product=self.object.product
        ).select_related('supplier').order_by('-is_preferred', 'price')
        
        # Проверяем, находимся ли мы ниже точки заказа
        context['is_below_reorder'] = current_quantity <= self.object.reorder_point
        context['percentage'] = (current_quantity / self.object.reorder_point * 100) if self.object.reorder_point > 0 else 0
        
        return context


class OrderPointDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Представление для удаления точки заказа"""
    model = OrderPoint
    template_name = 'warehouse/purchasing/order_point_confirm_delete.html'
    permission_required = 'warehouse.delete_orderpoint'
    success_url = reverse_lazy('warehouse:order_point_list')


class PurchaseNotificationListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Представление для списка уведомлений о закупке"""
    model = PurchaseNotification
    template_name = 'warehouse/purchasing/purchase_notification_list.html'
    context_object_name = 'notifications'
    permission_required = 'warehouse.view_purchasenotification'
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('product', 'recommended_supplier')
        
        # Фильтрация по статусу
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        else:
            # По умолчанию показываем только активные
            queryset = queryset.exclude(status__in=['ordered', 'canceled'])
        
        # Фильтрация по продукту
        product_id = self.request.GET.get('product')
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        # Фильтрация по поставщику
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            queryset = queryset.filter(recommended_supplier_id=supplier_id)
        
        # Фильтрация по дате
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(recommended_order_date__gte=date_from)
        
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(recommended_order_date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем дополнительные данные для шаблона
        context['products'] = Product.objects.all()
        context['suppliers'] = Supplier.objects.filter(is_active=True)
        
        # Сохраняем параметры фильтрации
        context['status_filter'] = self.request.GET.get('status', '')
        context['product_filter'] = self.request.GET.get('product', '')
        context['supplier_filter'] = self.request.GET.get('supplier', '')
        context['date_from_filter'] = self.request.GET.get('date_from', '')
        context['date_to_filter'] = self.request.GET.get('date_to', '')
        
        # Группируем уведомления по продуктам для отображения статистики
        product_stats = {}
        for notification in self.object_list:
            product_id = notification.product_id
            if product_id not in product_stats:
                product_stats[product_id] = {
                    'product': notification.product,
                    'count': 0,
                    'quantity_total': 0
                }
            product_stats[product_id]['count'] += 1
            product_stats[product_id]['quantity_total'] += notification.quantity_needed
        
        context['product_stats'] = product_stats.values()
        
        return context


class PurchaseNotificationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Представление для создания уведомления о закупке"""
    model = PurchaseNotification
    form_class = PurchaseNotificationForm
    template_name = 'warehouse/purchasing/purchase_notification_form.html'
    permission_required = 'warehouse.add_purchasenotification'
    
    def get_initial(self):
        initial = super().get_initial()
        
        # Если указан продукт в параметрах
        product_id = self.request.GET.get('product')
        if product_id:
            try:
                product = Product.objects.get(pk=product_id)
                initial['product'] = product
                
                # Определяем рекомендуемое количество заказа
                try:
                    order_point = OrderPoint.objects.get(product=product)
                    initial['quantity_needed'] = order_point.order_quantity
                except OrderPoint.DoesNotExist:
                    # Если нет точки заказа, предлагаем заказать минимальный уровень продукта
                    try:
                        min_level = ProductMinLevel.objects.get(product=product)
                        initial['quantity_needed'] = min_level.min_quantity
                    except ProductMinLevel.DoesNotExist:
                        # Если и минимального уровня нет, используем запас на 30 дней
                        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
                        consumption = Movement.objects.filter(
                            product=product,
                            movement_type__in=['sale', 'production_usage'],
                            created_at__gte=thirty_days_ago
                        ).aggregate(total=models.Sum('quantity'))['total'] or 0
                        initial['quantity_needed'] = consumption
                
                # Определяем оптимального поставщика
                try:
                    order_point = OrderPoint.objects.get(product=product)
                    optimal_supplier = order_point.get_optimal_supplier()
                    if optimal_supplier:
                        initial['recommended_supplier'] = optimal_supplier.supplier
                        
                        # Определяем даты заказа и доставки
                        today = timezone.now().date()
                        initial['recommended_order_date'] = today
                        lead_time = optimal_supplier.lead_time or 7
                        initial['expected_delivery_date'] = today + timezone.timedelta(days=lead_time)
                except OrderPoint.DoesNotExist:
                    # Если нет точки заказа, предлагаем любого поставщика продукта
                    preferred_supplier = ProductSupplier.objects.filter(
                        product=product, is_preferred=True
                    ).first()
                    if preferred_supplier:
                        initial['recommended_supplier'] = preferred_supplier.supplier
                        
                        # Определяем даты заказа и доставки
                        today = timezone.now().date()
                        initial['recommended_order_date'] = today
                        lead_time = preferred_supplier.lead_time or 7
                        initial['expected_delivery_date'] = today + timezone.timedelta(days=lead_time)
            except Product.DoesNotExist:
                pass
        else:
            # Если продукт не указан, устанавливаем текущую дату для заказа
            initial['recommended_order_date'] = timezone.now().date()
            initial['expected_delivery_date'] = timezone.now().date() + timezone.timedelta(days=7)
            
        return initial
    
    def form_valid(self, form):
        # Сохраняем уведомление
        notification = form.save(commit=False)
        
        # Уведомляем соответствующих пользователей
        product = notification.product
        emails = []
        
        # Получаем email из точки заказа
        try:
            order_point = OrderPoint.objects.get(product=product)
            if order_point.notify_emails:
                emails.extend([e.strip() for e in order_point.notify_emails.split(',')])
        except OrderPoint.DoesNotExist:
            pass
        
        # Отправляем уведомления
        if emails:
            self.send_notification_emails(emails)
        
        notification.save()
        
        messages.success(self.request, 'Уведомление о закупке успешно создано.')
        return super().form_valid(form)
    
    def send_notification_emails(self, emails):
        """Отправка email-уведомлений"""
        # Реализация отправки email
        pass
    
    def get_success_url(self):
        return reverse_lazy('warehouse:purchase_notification_list')


class PurchaseNotificationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Представление для редактирования уведомления о закупке"""
    model = PurchaseNotification
    form_class = PurchaseNotificationForm
    template_name = 'warehouse/purchasing/purchase_notification_form.html'
    permission_required = 'warehouse.change_purchasenotification'
    
    def get_success_url(self):
        return reverse_lazy('warehouse:purchase_notification_detail', kwargs={'pk': self.object.pk})


class PurchaseNotificationDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Представление для просмотра детальной информации о уведомлении"""
    model = PurchaseNotification
    template_name = 'warehouse/purchasing/purchase_notification_detail.html'
    context_object_name = 'notification'
    permission_required = 'warehouse.view_purchasenotification'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем историю использования продукта
        product = self.object.product
        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
        usage_history = Movement.objects.filter(
            product=product,
            movement_type__in=['sale', 'production_usage'],
            created_at__gte=thirty_days_ago
        ).values('created_at__date').annotate(
            quantity=models.Sum('quantity')
        ).order_by('created_at__date')
        
        context['usage_history'] = usage_history
        
        # Добавляем поставщиков продукта
        context['suppliers'] = ProductSupplier.objects.filter(
            product=product
        ).select_related('supplier').order_by('-is_preferred', 'price')
        
        # Добавляем текущий остаток продукта
        context['current_quantity'] = product.current_quantity()
        
        return context


class PurchaseNotificationStatusUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Представление для обновления статуса уведомления о закупке"""
    permission_required = 'warehouse.change_purchasenotification'
    
    def post(self, request, pk, status):
        notification = get_object_or_404(PurchaseNotification, pk=pk)
        
        # Проверка допустимости статуса
        valid_statuses = ['pending', 'acknowledged', 'ordered', 'canceled']
        if status not in valid_statuses:
            messages.error(request, 'Неверный статус.')
            return redirect('warehouse:purchase_notification_detail', pk=pk)
        
        notification.status = status
        notification.save()
        
        status_messages = {
            'pending': 'Уведомление переведено в статус "Ожидает обработки".',
            'acknowledged': 'Уведомление принято к сведению.',
            'ordered': 'Уведомление помечено как заказанное.',
            'canceled': 'Уведомление отменено.'
        }
        
        messages.success(request, status_messages[status])
        return redirect('warehouse:purchase_notification_detail', pk=pk)


class GenerateNotificationsView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Представление для автоматической генерации уведомлений о закупке"""
    template_name = 'warehouse/purchasing/generate_notifications.html'
    form_class = forms.Form
    permission_required = 'warehouse.add_purchasenotification'
    
    def form_valid(self, form):
        # Получаем все точки заказа
        order_points = OrderPoint.objects.select_related('product').all()
        notifications_created = 0
        
        for point in order_points:
            # Проверяем, находится ли продукт ниже точки заказа
            if point.is_below_reorder_point():
                # Проверяем, нет ли уже активного уведомления для этого продукта
                existing_notification = PurchaseNotification.objects.filter(
                    product=point.product,
                    status__in=['pending', 'acknowledged']
                ).exists()
                
                if not existing_notification:
                    # Определяем оптимального поставщика
                    optimal_supplier_link = point.get_optimal_supplier()
                    optimal_supplier = optimal_supplier_link.supplier if optimal_supplier_link else None
                    
                    # Определяем оптимальное время заказа
                    optimal_time_info = point.calculate_optimal_order_time()
                    order_date = optimal_time_info.get('date') if optimal_time_info else timezone.now().date()
                    lead_time = optimal_time_info.get('lead_time', 7)
                    
                    # Создаем уведомление
                    notification = PurchaseNotification.objects.create(
                        product=point.product,
                        quantity_needed=point.order_quantity,
                        recommended_supplier=optimal_supplier,
                        recommended_order_date=order_date,
                        expected_delivery_date=order_date + timezone.timedelta(days=lead_time),
                        status='pending',
                        note=f'Автоматически создано {timezone.now().strftime("%d.%m.%Y %H:%M")}. '
                             f'Текущий остаток: {point.product.current_quantity()}, '
                             f'точка заказа: {point.reorder_point}.'
                    )
                    
                    # Отправляем уведомления
                    if point.notify_emails:
                        emails = [e.strip() for e in point.notify_emails.split(',')]
                        self.send_notification_emails(notification, emails)
                    
                    notifications_created += 1
        
        if notifications_created > 0:
            messages.success(self.request, f'Создано {notifications_created} уведомлений о закупке.')
        else:
            messages.info(self.request, 'Новые уведомления о закупке не требуются.')
        
        return redirect('warehouse:purchase_notification_list')
    
    def send_notification_emails(self, notification, emails):
        """Отправка email-уведомлений"""
        # Реализация отправки email
        pass
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем все точки заказа
        order_points = OrderPoint.objects.select_related('product').all()
        points_below_reorder = []
        
        for point in order_points:
            current_quantity = point.product.current_quantity()
            if current_quantity <= point.reorder_point:
                # Проверяем, нет ли уже активного уведомления
                existing_notification = PurchaseNotification.objects.filter(
                    product=point.product,
                    status__in=['pending', 'acknowledged']
                ).exists()
                
                points_below_reorder.append({
                    'point': point,
                    'current_quantity': current_quantity,
                    'deficit': point.reorder_point - current_quantity,
                    'has_notification': existing_notification
                })
        
        context['points_below_reorder'] = points_below_reorder
        context['total_points'] = order_points.count()
        context['below_count'] = len(points_below_reorder)
        context['auto_order_count'] = len([p for p in points_below_reorder if p['point'].auto_order])
        
        # Получаем информацию о существующих уведомлениях
        active_notifications = PurchaseNotification.objects.filter(
            status__in=['pending', 'acknowledged']
        ).count()
        
        context['active_notifications'] = active_notifications
        
        return context

class SupplierListView(LoginRequiredMixin, ListView):
    """Представление для списка поставщиков"""
    model = Supplier
    template_name = 'warehouse/purchasing/supplier_list.html'
    context_object_name = 'suppliers'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        active_filter = self.request.GET.get('active')

        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(contact_person__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(email__icontains=search_query)
            )

        if active_filter == 'active':
            queryset = queryset.filter(is_active=True)
        elif active_filter == 'inactive':
            queryset = queryset.filter(is_active=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        suppliers = self.get_queryset()
        suppliers_with_data = []

        for supplier in suppliers:
            avg_rating = supplier.get_average_rating()
            avg_lead_time = supplier.get_average_lead_time()
            products_count = supplier.product_suppliers.count()

            suppliers_with_data.append({
                'supplier': supplier,
                'avg_rating': avg_rating,
                'avg_lead_time': avg_lead_time,
                'products_count': products_count
            })

        context['suppliers_with_data'] = suppliers_with_data
        context['search_query'] = self.request.GET.get('search', '')
        context['active_filter'] = self.request.GET.get('active', '')
        return context

class SupplierCreateView(LoginRequiredMixin, CreateView):
    """Представление для создания поставщика"""
    model = Supplier
    form_class = SupplierForm
    template_name = 'warehouse/purchasing/supplier_form.html'
    success_url = reverse_lazy('warehouse:supplier_list')

    def form_valid(self, form):
        messages.success(self.request, 'Поставщик успешно создан.')
        return super().form_valid(form)

class SupplierDetailView(LoginRequiredMixin, DetailView):
    """Представление для просмотра детальной информации о поставщике"""
    model = Supplier
    template_name = 'warehouse/purchasing/supplier_detail.html'
    context_object_name = 'supplier'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier = self.get_object()
        
        # Получаем все рейтинги поставщика
        ratings = supplier.ratings.all()
        ratings_by_category = {}
        for category, _ in SupplierRating.RATING_CHOICES:
            category_ratings = ratings.filter(category=category)
            if category_ratings.exists():
                avg_rating = category_ratings.aggregate(Avg('rating'))['rating__avg']
                ratings_by_category[category] = {
                    'avg': avg_rating,
                    'count': category_ratings.count()
                }
        
        # Получаем связанные продукты
        product_suppliers = supplier.product_suppliers.select_related('product').all()
        
        context.update({
            'ratings_by_category': ratings_by_category,
            'product_suppliers': product_suppliers,
            'avg_rating': supplier.get_average_rating(),
            'avg_lead_time': supplier.get_average_lead_time(),
            'products_count': product_suppliers.count(),
            'ratings': ratings
        })
        return context

class SupplierUpdateView(LoginRequiredMixin, UpdateView):
    """Представление для редактирования поставщика"""
    model = Supplier
    form_class = SupplierForm
    template_name = 'warehouse/purchasing/supplier_form.html'

    def get_success_url(self):
        return reverse_lazy('warehouse:supplier_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Данные поставщика успешно обновлены.')
        return super().form_valid(form)

class SupplierDeleteView(LoginRequiredMixin, DeleteView):
    """Представление для удаления поставщика"""
    model = Supplier
    template_name = 'warehouse/purchasing/supplier_confirm_delete.html'
    success_url = reverse_lazy('warehouse:supplier_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Поставщик успешно удален.')
        return super().delete(request, *args, **kwargs)

class SupplierRatingCreateView(LoginRequiredMixin, CreateView):
    """Представление для создания оценки поставщика"""
    model = SupplierRating
    form_class = SupplierRatingForm
    template_name = 'warehouse/purchasing/supplier_rating_form.html'

    def get_initial(self):
        initial = super().get_initial()
        supplier_id = self.kwargs.get('supplier_id')
        if supplier_id:
            initial['supplier'] = supplier_id
        return initial

    def get_success_url(self):
        return reverse_lazy('warehouse:supplier_detail', kwargs={'pk': self.object.supplier.pk})

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Оценка поставщика успешно добавлена.')
        return super().form_valid(form)

class ProductSupplierListView(LoginRequiredMixin, ListView):
    """Представление для списка связей продуктов с поставщиками"""
    model = ProductSupplier
    template_name = 'warehouse/purchasing/product_supplier_list.html'
    context_object_name = 'product_suppliers'

class ProductSupplierCreateView(LoginRequiredMixin, CreateView):
    """Представление для создания связи продукта с поставщиком"""
    model = ProductSupplier
    form_class = ProductSupplierForm
    template_name = 'warehouse/purchasing/product_supplier_form.html'
    success_url = reverse_lazy('warehouse:product_supplier_list')

    def form_valid(self, form):
        messages.success(self.request, 'Связь продукта с поставщиком успешно создана.')
        return super().form_valid(form)

class ProductSupplierUpdateView(LoginRequiredMixin, UpdateView):
    """Представление для редактирования связи продукта с поставщиком"""
    model = ProductSupplier
    form_class = ProductSupplierForm
    template_name = 'warehouse/purchasing/product_supplier_form.html'
    success_url = reverse_lazy('warehouse:product_supplier_list')

    def form_valid(self, form):
        messages.success(self.request, 'Связь продукта с поставщиком успешно обновлена.')
        return super().form_valid(form)

class ProductSupplierDeleteView(LoginRequiredMixin, DeleteView):
    """Представление для удаления связи продукта с поставщиком"""
    model = ProductSupplier
    template_name = 'warehouse/purchasing/product_supplier_confirm_delete.html'
    success_url = reverse_lazy('warehouse:product_supplier_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Связь продукта с поставщиком успешно удалена.')
        return super().delete(request, *args, **kwargs)

class EstokadaListView(LoginRequiredMixin, ListView):
    model = EstokadaOperation
    template_name = 'warehouse/estokada_list.html'
    context_object_name = 'operations'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Фильтры
        status = self.request.GET.get('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        result = self.request.GET.get('result')
        
        if status:
            queryset = queryset.filter(status=status)
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        if result:
            queryset = queryset.filter(result=result)
        
        return queryset.select_related('movement', 'movement__product', 'processed_by')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика
        total_operations = EstokadaOperation.objects.count()
        in_progress = EstokadaOperation.objects.filter(status__in=['pending', 'measuring', 'processing']).count()
        completed = EstokadaOperation.objects.filter(status='completed').count()
        cancelled = EstokadaOperation.objects.filter(status='cancelled').count()
        
        context.update({
            'statistics': {
                'total': total_operations,
                'in_progress': in_progress,
                'completed': completed,
                'cancelled': cancelled
            },
            'status_choices': EstokadaOperation.STATUS_CHOICES,
            'result_choices': EstokadaOperation.OPERATION_RESULT,
            'selected_status': self.request.GET.get('status', ''),
            'selected_result': self.request.GET.get('result', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', '')
        })
        
        return context

class EstokadaDetailView(LoginRequiredMixin, DetailView):
    model = EstokadaOperation
    template_name = 'warehouse/estokada_detail.html'
    context_object_name = 'operation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        operation = self.get_object()
        
        # Получаем связанные данные
        context.update({
            'movement': operation.movement,
            'product': operation.movement.product,
        })
        
        return context


class EstokadaProcessView(LoginRequiredMixin, UpdateView):
    model = EstokadaOperation
    form_class = EstokadaOperationForm
    template_name = 'warehouse/estokada_process.html'
    context_object_name = 'operation'

    def get_success_url(self):
        return reverse_lazy('warehouse:estokada_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        operation = self.get_object()
        
        context.update({
            'movement': operation.movement,
            'product': operation.movement.product,
        })
        
        return context

    def form_valid(self, form):
        operation = form.save(commit=False)
        
        # Если статус изменился на completed или cancelled
        if operation.status in ['completed', 'cancelled']:
            operation.processed_by = self.request.user
            operation.processed_at = timezone.now()
        
        # Расчет разницы
        if operation.actual_quantity:
            operation.difference_quantity = operation.actual_quantity - operation.movement.quantity
            if operation.movement.quantity != 0:
                operation.difference_percentage = (operation.difference_quantity / operation.movement.quantity) * 100
            
            # Определение результата
            if abs(operation.difference_percentage) <= 0.1:
                operation.result = 'match'
            else:
                operation.result = 'positive_diff' if operation.difference_quantity > 0 else 'negative_diff'
        
        operation.save()
        return super().form_valid(form)

class TransportCreateView(LoginRequiredMixin, CreateView):
    model = Transport
    template_name = 'warehouse/transport_form.html'
    fields = ['transport_type', 'transport_number', 'wagon', 'wagon_type', 
              'density', 'temperature', 'volume', 'quantity', 'doc_quantity', 
              'warehouse', 'notes']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['movement'] = get_object_or_404(Movement, pk=self.kwargs['movement_pk'])
        return context
    
    def form_valid(self, form):
        transport = form.save(commit=False)
        transport.movement = get_object_or_404(Movement, pk=self.kwargs['movement_pk'])
        transport.save()
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('warehouse:movement_detail', 
                          kwargs={'pk': self.kwargs['movement_pk']})

class ReceptionCreateView(LoginRequiredMixin, CreateView):
    """Представление для создания операции приемки товара"""
    model = Movement
    template_name = 'warehouse/reception_form.html'
    form_class = MovementForm
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем в контекст список резервуаров и вагонов
        context['reservoirs'] = Reservoir.objects.all()
        context['wagons'] = Wagon.objects.all()
        context['wagon_types'] = WagonType.objects.all()
        return context
    
    def form_valid(self, form):
        # Устанавливаем тип движения как "in" (приемка)
        form.instance.movement_type = 'in'
        
        # Сохраняем движение, но не коммитим пока в базу
        movement = form.save(commit=False)
        
        # Устанавливаем куда положить товар (резервуар или вагон)
        placement_type = self.request.POST.get('placement_type')
        if placement_type == 'reservoir':
            reservoir_id = self.request.POST.get('reservoir')
            if reservoir_id:
                movement.reservoir_id = reservoir_id
        elif placement_type == 'wagon':
            wagon_id = self.request.POST.get('wagon')
            if wagon_id:
                movement.wagon_id = wagon_id
        
        # Обрабатываем данные о транспорте
        try:
            # Получаем JSON с данными о транспорте
            transports_json = self.request.POST.get('transports_json')
            if transports_json:
                transports_data = json.loads(transports_json)
                
                # Проверяем, что есть хотя бы один транспорт
                if not transports_data:
                    form.add_error(None, 'Необходимо добавить хотя бы один транспорт')
                    return self.form_invalid(form)
                
                # Рассчитываем общее количество из всех транспортов
                total_quantity = 0
                total_doc_ton = 0
                
                for transport_data in transports_data:
                    quantity = float(transport_data.get('quantity', 0))
                    doc_ton = float(transport_data.get('doc_ton', 0) or 0)
                    
                    if quantity <= 0:
                        form.add_error(None, 'Количество должно быть больше 0')
                        return self.form_invalid(form)
                    
                    total_quantity += quantity
                    total_doc_ton += doc_ton
                
                # Переводим кг в тонны
                movement.quantity = total_quantity / 1000
                movement.doc_ton = total_doc_ton / 1000
                
                # Сохраняем движение, чтобы получить ID
                movement.save()
                
                # Создаем записи транспорта
                for transport_data in transports_data:
                    transport_type = transport_data.get('transport_type')
                    transport_number = transport_data.get('transport_number')
                    quantity_kg = float(transport_data.get('quantity', 0))
                    doc_ton_kg = float(transport_data.get('doc_ton', 0) or 0)
                    
                    transport = Transport(
                        movement=movement,
                        transport_type=transport_type,
                        transport_number=transport_number,
                        quantity=quantity_kg,  # В кг
                        doc_ton=doc_ton_kg / 1000  # В тоннах
                    )
                    
                    # Добавляем специфические данные в зависимости от типа транспорта
                    if transport_type == 'truck':
                        transport.density = transport_data.get('density')
                        transport.temperature = transport_data.get('temperature')
                        transport.liter = transport_data.get('liter')
                    elif transport_type == 'wagon':
                        if 'wagon_type' in transport_data and transport_data['wagon_type']:
                            transport.wagon_type_id = transport_data['wagon_type']
                        transport.capacity = transport_data.get('capacity')
                        transport.tare_weight = transport_data.get('tare_weight')
                    
                    transport.save()
                
                # Обновляем запасы в резервуаре или вагоне
                if movement.reservoir:
                    reservoir = movement.reservoir
                    reservoir.current_quantity += movement.quantity
                    reservoir.save()
                elif movement.wagon:
                    wagon = movement.wagon
                    wagon.current_quantity += movement.quantity
                    wagon.save()
                
                # Обновляем инвентарь продукта
                inventory, created = Inventory.objects.get_or_create(
                    product=movement.product,
                    defaults={'quantity': 0}
                )
                inventory.quantity += movement.quantity
                inventory.save()
                
                # Обновляем счетчики продукта
                product = movement.product
                product.in_qty = (product.in_qty or 0) + movement.quantity
                product.save(update_fields=['in_qty'])
                
                messages.success(self.request, 'Операция приемки успешно создана')
                return redirect('warehouse:movement_list')
            else:
                form.add_error(None, 'Необходимо добавить данные о транспорте')
                return self.form_invalid(form)
        except json.JSONDecodeError:
            form.add_error(None, 'Ошибка в данных о транспорте')
            return self.form_invalid(form)
        except Exception as e:
            form.add_error(None, f'Произошла ошибка: {str(e)}')
            return self.form_invalid(form)


class SaleCreateView(LoginRequiredMixin, CreateView):
    """Представление для создания операции продажи товара"""
    model = Movement
    template_name = 'warehouse/sale_form.html'
    form_class = MovementForm
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем в контекст список резервуаров и вагонов
        context['reservoirs'] = Reservoir.objects.all()
        context['wagons'] = Wagon.objects.all()
        context['wagon_types'] = WagonType.objects.all()
        return context
    
    def form_valid(self, form):
        # Устанавливаем тип движения как "out" (продажа)
        form.instance.movement_type = 'out'
        
        # Сохраняем движение, но не коммитим пока в базу
        movement = form.save(commit=False)
        
        # Получаем источник товара (резервуар или вагон)
        source_type = self.request.POST.get('source_type')
        source_id = self.request.POST.get('source_id')
        
        # Проверяем доступное количество
        available_quantity = 0
        
        if source_type == 'reservoir' and source_id:
            try:
                reservoir = Reservoir.objects.get(pk=source_id)
                available_quantity = reservoir.current_quantity
                movement.reservoir_id = source_id
            except Reservoir.DoesNotExist:
                form.add_error(None, 'Указанный резервуар не найден')
                return self.form_invalid(form)
        elif source_type == 'wagon' and source_id:
            try:
                wagon = Wagon.objects.get(pk=source_id)
                available_quantity = wagon.current_quantity
                movement.wagon_id = source_id
            except Wagon.DoesNotExist:
                form.add_error(None, 'Указанный вагон не найден')
                return self.form_invalid(form)
        else:
            form.add_error(None, 'Необходимо указать источник продукта')
            return self.form_invalid(form)
        
        # Обрабатываем данные о транспорте
        try:
            # Получаем JSON с данными о транспорте
            transports_json = self.request.POST.get('transports_json')
            if transports_json:
                transports_data = json.loads(transports_json)
                
                # Проверяем, что есть хотя бы один транспорт
                if not transports_data:
                    form.add_error(None, 'Необходимо добавить хотя бы один транспорт')
                    return self.form_invalid(form)
                
                # Рассчитываем общее количество из всех транспортов
                total_quantity = 0
                total_doc_ton = 0
                
                for transport_data in transports_data:
                    quantity = float(transport_data.get('quantity', 0))
                    doc_ton = float(transport_data.get('doc_ton', 0) or 0)
                    
                    if quantity <= 0:
                        form.add_error(None, 'Количество должно быть больше 0')
                        return self.form_invalid(form)
                    
                    total_quantity += quantity
                    total_doc_ton += doc_ton
                
                # Переводим кг в тонны
                total_quantity_tons = total_quantity / 1000
                
                # Проверяем, не превышает ли общее количество доступное количество
                if total_quantity_tons > available_quantity:
                    form.add_error(None, f'Общее количество ({total_quantity_tons:.2f} тонн) превышает доступное количество ({available_quantity:.2f} тонн)')
                    return self.form_invalid(form)
                
                # Устанавливаем количество и документальный вес
                movement.quantity = total_quantity_tons
                movement.doc_ton = total_doc_ton / 1000
                
                # Сохраняем движение, чтобы получить ID
                movement.save()
                
                # Создаем записи транспорта
                for transport_data in transports_data:
                    transport_type = transport_data.get('transport_type')
                    transport_number = transport_data.get('transport_number')
                    quantity_kg = float(transport_data.get('quantity', 0))
                    doc_ton_kg = float(transport_data.get('doc_ton', 0) or 0)
                    
                    transport = Transport(
                        movement=movement,
                        transport_type=transport_type,
                        transport_number=transport_number,
                        quantity=quantity_kg,  # В кг
                        doc_ton=doc_ton_kg / 1000  # В тоннах
                    )
                    
                    # Добавляем специфические данные в зависимости от типа транспорта
                    if transport_type == 'truck':
                        transport.density = transport_data.get('density')
                        transport.temperature = transport_data.get('temperature')
                        transport.liter = transport_data.get('liter')
                    elif transport_type == 'wagon':
                        if 'wagon_type' in transport_data and transport_data['wagon_type']:
                            transport.wagon_type_id = transport_data['wagon_type']
                        transport.capacity = transport_data.get('capacity')
                        transport.tare_weight = transport_data.get('tare_weight')
                    
                    transport.save()
                
                # Обновляем запасы в резервуаре или вагоне
                if source_type == 'reservoir':
                    reservoir.current_quantity -= movement.quantity
                    reservoir.save()
                elif source_type == 'wagon':
                    wagon.current_quantity -= movement.quantity
                    wagon.save()
                
                # Обновляем инвентарь продукта
                inventory, created = Inventory.objects.get_or_create(
                    product=movement.product,
                    defaults={'quantity': 0}
                )
                inventory.quantity -= movement.quantity
                inventory.save()
                
        except json.JSONDecodeError:
            form.add_error(None, 'Ошибка в данных о транспорте')
            return self.form_invalid(form)
        except Exception as e:
            form.add_error(None, f'Произошла ошибка: {str(e)}')
            return self.form_invalid(form)

class TransferCreateView(LoginRequiredMixin, CreateView):
    """Представление для операции перемещения товара между резервуарами и вагонами"""
    model = Movement
    template_name = 'warehouse/transfer_form.html'
    form_class = MovementForm
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем в контекст список резервуаров и вагонов
        context['reservoirs'] = Reservoir.objects.all()
        context['wagons'] = Wagon.objects.all()
        return context
    
    def form_valid(self, form):
        # Устанавливаем тип движения как "transfer" (перемещение)
        form.instance.movement_type = 'transfer'
        
        # Получаем данные формы
        movement = form.save(commit=False)
        
        # Получаем источник товара (резервуар или вагон)
        source_type = self.request.POST.get('source_type')
        source_id = self.request.POST.get('source_id')
        
        # Получаем назначение товара (резервуар или вагон)
        destination_type = self.request.POST.get('destination_type')
        destination_id = self.request.POST.get('destination_id')
        
        # Получаем количество для перемещения
        quantity = float(self.request.POST.get('transfer_quantity') or 0)
        
        if quantity <= 0:
            form.add_error(None, 'Количество должно быть больше 0')
            return self.form_invalid(form)
        
        # Устанавливаем количество в движении
        movement.quantity = quantity
        
        # Проверяем доступное количество в источнике
        source_obj = None
        available_quantity = 0
        
        if source_type == 'reservoir' and source_id:
            try:
                source_obj = Reservoir.objects.get(pk=source_id)
                available_quantity = source_obj.current_quantity
                movement.source_reservoir_id = source_id
            except Reservoir.DoesNotExist:
                form.add_error(None, 'Указанный резервуар-источник не найден')
                return self.form_invalid(form)
        elif source_type == 'wagon' and source_id:
            try:
                source_obj = Wagon.objects.get(pk=source_id)
                available_quantity = source_obj.current_quantity
                movement.source_wagon_id = source_id
            except Wagon.DoesNotExist:
                form.add_error(None, 'Указанный вагон-источник не найден')
                return self.form_invalid(form)
        else:
            form.add_error(None, 'Необходимо указать источник продукта')
            return self.form_invalid(form)
        
        # Проверяем, не превышает ли количество доступное в источнике
        if quantity > available_quantity:
            form.add_error(None, f'Количество ({quantity:.2f} тонн) превышает доступное в источнике ({available_quantity:.2f} тонн)')
            return self.form_invalid(form)
        
        # Проверяем доступное пространство в назначении
        destination_obj = None
        available_space = 0
        
        if destination_type == 'reservoir' and destination_id:
            try:
                destination_obj = Reservoir.objects.get(pk=destination_id)
                available_space = destination_obj.capacity - destination_obj.current_quantity
                movement.destination_reservoir_id = destination_id
            except Reservoir.DoesNotExist:
                form.add_error(None, 'Указанный резервуар-назначение не найден')
                return self.form_invalid(form)
        elif destination_type == 'wagon' and destination_id:
            try:
                destination_obj = Wagon.objects.get(pk=destination_id)
                available_space = destination_obj.capacity - destination_obj.current_quantity
                movement.destination_wagon_id = destination_id
            except Wagon.DoesNotExist:
                form.add_error(None, 'Указанный вагон-назначение не найден')
                return self.form_invalid(form)
        else:
            form.add_error(None, 'Необходимо указать назначение продукта')
            return self.form_invalid(form)
        
        # Проверяем, не превышает ли количество доступное пространство в назначении
        if quantity > available_space:
            form.add_error(None, f'Количество ({quantity:.2f} тонн) превышает доступное пространство в назначении ({available_space:.2f} тонн)')
            return self.form_invalid(form)
        
        # Проверяем, что источник и назначение не одинаковые
        if source_type == destination_type and source_id == destination_id:
            form.add_error(None, 'Источник и назначение не могут быть одинаковыми')
            return self.form_invalid(form)
        
        try:
            # Сохраняем движение
            movement.save()
            
            # Обновляем количество в источнике (уменьшаем)
            if source_type == 'reservoir':
                source_obj.current_quantity -= quantity
                source_obj.save()
            elif source_type == 'wagon':
                source_obj.current_quantity -= quantity
                source_obj.save()
            
            # Обновляем количество в назначении (увеличиваем)
            if destination_type == 'reservoir':
                destination_obj.current_quantity += quantity
                destination_obj.save()
            elif destination_type == 'wagon':
                destination_obj.current_quantity += quantity
                destination_obj.save()
            
            messages.success(self.request, f'Операция перемещения {quantity:.2f} тонн успешно выполнена')
            return redirect('warehouse:movement_list')
        except Exception as e:
            form.add_error(None, f'Произошла ошибка: {str(e)}')
            return self.form_invalid(form)

class ProductionCreateView(LoginRequiredMixin, CreateView):
    """Представление для операции производства товара"""
    model = Movement
    template_name = 'warehouse/production_form.html'
    form_class = MovementForm
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем в контекст список продуктов, резервуаров и вагонов
        context['products'] = Product.objects.all()
        context['reservoirs'] = Reservoir.objects.all()
        context['wagons'] = Wagon.objects.all()
        return context
    
    def form_valid(self, form):
        # Устанавливаем тип движения как "production" (производство)
        form.instance.movement_type = 'production'
        
        # Получаем данные формы
        movement = form.save(commit=False)
        
        # Получаем количество произведенного продукта
        quantity = float(self.request.POST.get('quantity') or 0)
        
        if quantity <= 0:
            form.add_error(None, 'Количество должно быть больше 0')
            return self.form_invalid(form)
        
        # Устанавливаем количество в движении
        movement.quantity = quantity
        
        # Получаем место размещения готового продукта
        placement_type = self.request.POST.get('placement_type')
        placement_id = self.request.POST.get('placement_id')
        
        placement_obj = None
        available_space = 0
        
        if placement_type == 'reservoir' and placement_id:
            try:
                placement_obj = Reservoir.objects.get(pk=placement_id)
                available_space = placement_obj.capacity - placement_obj.current_quantity
                movement.reservoir_id = placement_id
            except Reservoir.DoesNotExist:
                form.add_error(None, 'Указанный резервуар не найден')
                return self.form_invalid(form)
        elif placement_type == 'wagon' and placement_id:
            try:
                placement_obj = Wagon.objects.get(pk=placement_id)
                available_space = placement_obj.capacity - placement_obj.current_quantity
                movement.wagon_id = placement_id
            except Wagon.DoesNotExist:
                form.add_error(None, 'Указанный вагон не найден')
                return self.form_invalid(form)
        else:
            form.add_error(None, 'Необходимо указать место размещения продукта')
            return self.form_invalid(form)
        
        # Проверяем, не превышает ли количество доступное пространство
        if quantity > available_space:
            form.add_error(None, f'Количество ({quantity:.2f} тонн) превышает доступное пространство ({available_space:.2f} тонн)')
            return self.form_invalid(form)
        
        # Получаем данные об ингредиентах
        try:
            ingredients_json = self.request.POST.get('ingredients_json')
            if not ingredients_json:
                form.add_error(None, 'Не указаны ингредиенты для производства')
                return self.form_invalid(form)
            
            ingredients_data = json.loads(ingredients_json)
            
            if not ingredients_data:
                form.add_error(None, 'Необходимо добавить хотя бы один ингредиент')
                return self.form_invalid(form)
            
            # Проверяем ингредиенты
            for i, ingredient in enumerate(ingredients_data):
                product_id = ingredient.get('product_id')
                ingredient_quantity = float(ingredient.get('quantity') or 0)
                source_type = ingredient.get('source_type')
                source_id = ingredient.get('source_id')
                
                if not product_id:
                    form.add_error(None, f'Не указан продукт для ингредиента #{i + 1}')
                    return self.form_invalid(form)
                
                if ingredient_quantity <= 0:
                    form.add_error(None, f'Количество должно быть больше 0 для ингредиента #{i + 1}')
                    return self.form_invalid(form)
                
                # Проверяем доступное количество ингредиента
                source_obj = None
                available_quantity = 0
                
                if source_type == 'reservoir' and source_id:
                    try:
                        source_obj = Reservoir.objects.get(pk=source_id)
                        if source_obj.product_id != int(product_id):
                            form.add_error(None, f'Резервуар содержит другой продукт для ингредиента #{i + 1}')
                            return self.form_invalid(form)
                        available_quantity = source_obj.current_quantity
                    except Reservoir.DoesNotExist:
                        form.add_error(None, f'Указанный резервуар не найден для ингредиента #{i + 1}')
                        return self.form_invalid(form)
                elif source_type == 'wagon' and source_id:
                    try:
                        source_obj = Wagon.objects.get(pk=source_id)
                        if source_obj.product_id != int(product_id):
                            form.add_error(None, f'Вагон содержит другой продукт для ингредиента #{i + 1}')
                            return self.form_invalid(form)
                        available_quantity = source_obj.current_quantity
                    except Wagon.DoesNotExist:
                        form.add_error(None, f'Указанный вагон не найден для ингредиента #{i + 1}')
                        return self.form_invalid(form)
                else:
                    form.add_error(None, f'Необходимо указать источник продукта для ингредиента #{i + 1}')
                    return self.form_invalid(form)
                
                # Проверяем, не превышает ли количество доступное в источнике
                if ingredient_quantity > available_quantity:
                    form.add_error(None, f'Количество ({ingredient_quantity:.2f} тонн) превышает доступное в источнике ({available_quantity:.2f} тонн) для ингредиента #{i + 1}')
                    return self.form_invalid(form)
            
            # Сохраняем движение
            movement.save()
            
            # Обрабатываем ингредиенты и создаем связанные движения
            for ingredient in ingredients_data:
                product_id = ingredient.get('product_id')
                ingredient_quantity = float(ingredient.get('quantity') or 0)
                source_type = ingredient.get('source_type')
                source_id = ingredient.get('source_id')
                
                # Создаем движение для вычитания ингредиента
                ingredient_movement = Movement(
                    movement_type='production_ingredient',
                    product_id=product_id,
                    quantity=ingredient_quantity,
                    date=movement.date,
                    document_number=f"{movement.document_number}-{product_id}",
                    parent_movement=movement,
                    note=f"Ингредиент для производства {movement.document_number}"
                )
                
                # Устанавливаем источник ингредиента
                if source_type == 'reservoir':
                    source_obj = Reservoir.objects.get(pk=source_id)
                    ingredient_movement.reservoir_id = source_id
                    
                    # Обновляем количество в резервуаре
                    source_obj.current_quantity -= ingredient_quantity
                    source_obj.save()
                elif source_type == 'wagon':
                    source_obj = Wagon.objects.get(pk=source_id)
                    ingredient_movement.wagon_id = source_id
                    
                    # Обновляем количество в вагоне
                    source_obj.current_quantity -= ingredient_quantity
                    source_obj.save()
                
                ingredient_movement.save()
                
                # Обновляем инвентарь ингредиента
                ingredient_inventory, created = Inventory.objects.get_or_create(
                    product_id=product_id,
                    defaults={'quantity': 0}
                )
                ingredient_inventory.quantity -= ingredient_quantity
                ingredient_inventory.save()
            
            # Обновляем место размещения готового продукта
            if placement_type == 'reservoir':
                placement_obj.current_quantity += quantity
                placement_obj.save()
            elif placement_type == 'wagon':
                placement_obj.current_quantity += quantity
                placement_obj.save()
            
            # Обновляем инвентарь готового продукта
            product_inventory, created = Inventory.objects.get_or_create(
                product=movement.product,
                defaults={'quantity': 0}
            )
            product_inventory.quantity += quantity
            product_inventory.save()
            
            # Обновляем счетчики продукта
            product = movement.product
            product.production_qty = (product.production_qty or 0) + quantity
            product.save(update_fields=['production_qty'])
            
            messages.success(self.request, f'Операция производства {quantity:.2f} тонн продукта успешно выполнена')
            return redirect('warehouse:movement_list')
        except json.JSONDecodeError:
            form.add_error(None, 'Ошибка в данных об ингредиентах')
            return self.form_invalid(form)
        except Exception as e:
            form.add_error(None, f'Произошла ошибка: {str(e)}')
            return self.form_invalid(form)
