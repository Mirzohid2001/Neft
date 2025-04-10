import io
import pandas as pd
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.http import HttpResponse, JsonResponse
from .services import create_financial_records, calculate_summary
from .filters import FinancialRecordFilter
from .models import FinancialRecord, Account, Transaction, Category, JournalEntry, Budget
from django.db.models import Sum, Q, Count
from django.http import JsonResponse
from apps.warehouse.models import Product, Warehouse
from .utils import send_telegram_message
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils import timezone
import json
import random
import csv
from openpyxl import Workbook
from .forms import (
    AccountForm, 
    TransactionForm, 
    CategoryForm
)
from .templatetags.accounting_tags import get_date_range_options

class FinancialDashboardView(View):
    template_name = 'accounting/dashboard.html'

    def get(self, request):
        filter_type = request.GET.get('filter_type', 'all')

        today = datetime.today()
        date_from, date_to = None, None

        if filter_type == 'today':
            date_from = today
            date_to = today

        elif filter_type == 'week':
            date_from = today - timedelta(days=7)
            date_to = today

        elif filter_type == 'month':
            date_from = today - timedelta(days=30)
            date_to = today

        elif filter_type == 'custom':
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')

        create_financial_records(date_from, date_to)
        summary = calculate_summary()

        limit = 1000000 
        net_result = summary['net_profit_loss']
        status = summary['status']

        if abs(net_result) >= limit:
            message = f"⚠️ <b>{status}</b>\nSumma: {net_result:,} UZS\nChegara: {limit:,} UZS dan oshdi."
            send_telegram_message(message)

        records = FinancialRecord.objects.all()
        filters = FinancialRecordFilter(request.GET, queryset=records)

        if request.GET.get('export') == 'excel':
            return self.export_excel(filters.qs)

        context = {
            'summary': summary,
            'filter': filters,
            'records': filters.qs,
            'filter_type': filter_type,
            'date_from': date_from,
            'date_to': date_to,
        }
        return render(request, self.template_name, context)

    def export_excel(self, queryset):
        df = pd.DataFrame(list(queryset.values(
            'date', 'product__name', 'warehouse__name',
            'movement_type', 'quantity', 'total_price_sum', 'total_price_usd'
        )))

        df.columns = [
            'Sana', 'Mahsulot', 'Ombor', 'Harakat turi',
            'Miqdor', 'Summa (UZS)', 'Summa (USD)'
        ]

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Moliyaviy Hisobot')

        buffer.seek(0)
        response = HttpResponse(
            buffer, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="hisobot.xlsx"'
        return response
    
class ChartDataView(View):

    def get(self, request):
        products = Product.objects.all()
        warehouses = Warehouse.objects.all()

        product_data = []
        for product in products:
            in_sum = FinancialRecord.objects.filter(product=product, movement_type='in').aggregate(Sum('total_price_sum'))['total_price_sum__sum'] or 0
            out_sum = FinancialRecord.objects.filter(product=product, movement_type='out').aggregate(Sum('total_price_sum'))['total_price_sum__sum'] or 0
            product_data.append({
                'product': product.name,
                'in_sum': in_sum,
                'out_sum': out_sum
            })

        warehouse_data = []
        for warehouse in warehouses:
            total = FinancialRecord.objects.filter(warehouse=warehouse).aggregate(Sum('total_price_sum'))['total_price_sum__sum'] or 0
            warehouse_data.append({
                'warehouse': warehouse.name,
                'total_sum': total
            })

        return JsonResponse({'product_data': product_data, 'warehouse_data': warehouse_data})

# Вспомогательные функции
def get_date_range(period='month'):
    today = timezone.now().date()
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = today.replace(day=1)
    elif period == 'quarter':
        month = today.month - 3
        year = today.year
        if month <= 0:
            month += 12
            year -= 1
        start_date = today.replace(year=year, month=month, day=1)
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today - timedelta(days=30)
    
    return start_date, today

# Главная страница - панель управления
@login_required
def dashboard(request):
    # Получаем недавние транзакции
    recent_transactions = Transaction.objects.all().order_by('-date', '-id')[:5]
    
    # Получаем активные счета с балансами
    accounts = Account.objects.filter(is_active=True)
    
    # Общая статистика
    total_income = Transaction.objects.filter(transaction_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = Transaction.objects.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    net_flow = total_income - total_expense
    
    # Категории с наибольшими расходами
    top_expense_categories = Category.objects.filter(
        transactions__transaction_type='expense'
    ).annotate(
        total=Sum('transactions__amount')
    ).order_by('-total')[:5]
    
    # Категории с наибольшими доходами
    top_income_categories = Category.objects.filter(
        transactions__transaction_type='income'
    ).annotate(
        total=Sum('transactions__amount')
    ).order_by('-total')[:5]
    
    context = {
        'recent_transactions': recent_transactions,
        'accounts': accounts,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_flow': net_flow,
        'top_expense_categories': top_expense_categories,
        'top_income_categories': top_income_categories,
    }
    return render(request, 'accounting/dashboard.html', context)

# Представления для транзакций
@login_required
def transaction_list(request):
    transactions = Transaction.objects.all().order_by('-date', '-id')
    
    # Фильтрация
    transaction_type = request.GET.get('type')
    account_id = request.GET.get('account')
    category_id = request.GET.get('category')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    if account_id:
        transactions = transactions.filter(
            Q(account_id=account_id) | Q(to_account_id=account_id)
        )
    
    if category_id:
        transactions = transactions.filter(category_id=category_id)
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    # Данные для фильтров
    accounts = Account.objects.filter(is_active=True)
    categories = Category.objects.all()
    
    context = {
        'transactions': transactions,
        'accounts': accounts,
        'categories': categories,
        'date_ranges': get_date_range_options(),
    }
    return render(request, 'accounting/transaction_list.html', context)

@login_required
def transaction_detail(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    context = {
        'transaction': transaction
    }
    return render(request, 'accounting/transaction_detail.html', context)

@login_required
def transaction_create(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save()
            messages.success(request, 'Транзакция успешно создана')
            return redirect('accounting:transaction_detail', pk=transaction.pk)
    else:
        form = TransactionForm()
    
    context = {
        'form': form,
        'is_new': True
    }
    return render(request, 'accounting/transaction_form.html', context)

@login_required
def transaction_edit(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            transaction = form.save()
            messages.success(request, 'Транзакция успешно обновлена')
            return redirect('accounting:transaction_detail', pk=transaction.pk)
    else:
        form = TransactionForm(instance=transaction)
    
    context = {
        'form': form,
        'transaction': transaction,
        'is_new': False
    }
    return render(request, 'accounting/transaction_form.html', context)

@login_required
def transaction_delete(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    if request.method == 'POST':
        transaction.delete()
        messages.success(request, 'Транзакция успешно удалена')
        return redirect('accounting:transaction_list')
    
    context = {
        'transaction': transaction
    }
    return render(request, 'accounting/transaction_confirm_delete.html', context)

# Представления для счетов
@login_required
def account_list(request):
    accounts = Account.objects.all()
    context = {
        'accounts': accounts
    }
    return render(request, 'accounting/account_list.html', context)

@login_required
def account_detail(request, pk):
    account = get_object_or_404(Account, pk=pk)
    transactions = Transaction.objects.filter(
        Q(account=account) | Q(to_account=account)
    ).order_by('-date', '-id')
    
    context = {
        'account': account,
        'transactions': transactions
    }
    return render(request, 'accounting/account_detail.html', context)

@login_required
def account_create(request):
    if request.method == 'POST':
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save()
            messages.success(request, 'Счет успешно создан')
            return redirect('accounting:account_detail', pk=account.pk)
    else:
        form = AccountForm()
    
    context = {
        'form': form,
        'is_new': True
    }
    return render(request, 'accounting/account_form.html', context)

@login_required
def account_edit(request, pk):
    account = get_object_or_404(Account, pk=pk)
    if request.method == 'POST':
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            account = form.save()
            messages.success(request, 'Счет успешно обновлен')
            return redirect('accounting:account_detail', pk=account.pk)
    else:
        form = AccountForm(instance=account)
    
    context = {
        'form': form,
        'account': account,
        'is_new': False
    }
    return render(request, 'accounting/account_form.html', context)

@login_required
def account_delete(request, pk):
    account = get_object_or_404(Account, pk=pk)
    if request.method == 'POST':
        account.delete()
        messages.success(request, 'Счет успешно удален')
        return redirect('accounting:account_list')
    
    context = {
        'account': account
    }
    return render(request, 'accounting/account_confirm_delete.html', context)

# Представления для категорий
@login_required
def category_list(request):
    categories = Category.objects.all()
    context = {
        'categories': categories
    }
    return render(request, 'accounting/category_list.html', context)

@login_required
def category_detail(request, pk):
    category = get_object_or_404(Category, pk=pk)
    transactions = Transaction.objects.filter(
        category=category
    ).order_by('-date', '-id')
    
    context = {
        'category': category,
        'transactions': transactions
    }
    return render(request, 'accounting/category_detail.html', context)

@login_required
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, 'Категория успешно создана')
            return redirect('accounting:category_detail', pk=category.pk)
    else:
        form = CategoryForm()
    
    context = {
        'form': form,
        'is_new': True
    }
    return render(request, 'accounting/category_form.html', context)

@login_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, 'Категория успешно обновлена')
            return redirect('accounting:category_detail', pk=category.pk)
    else:
        form = CategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
        'is_new': False
    }
    return render(request, 'accounting/category_form.html', context)

@login_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Категория успешно удалена')
        return redirect('accounting:category_list')
    
    context = {
        'category': category
    }
    return render(request, 'accounting/category_confirm_delete.html', context)

# Представления для отчетов
@login_required
def expense_report(request):
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    account_id = request.GET.get('account')
    
    # Базовый запрос
    transactions = Transaction.objects.filter(transaction_type='expense')
    
    # Применяем фильтры
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    if account_id:
        transactions = transactions.filter(account_id=account_id)
    
    # Группировка по категориям
    categories = Category.objects.filter(
        transactions__in=transactions
    ).annotate(
        total=Sum('transactions__amount'),
        count=Count('transactions')
    ).order_by('-total')
    
    total_expense = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Данные для фильтров
    accounts = Account.objects.filter(is_active=True)
    
    context = {
        'categories': categories,
        'total_expense': total_expense,
        'accounts': accounts,
        'date_ranges': get_date_range_options(),
    }
    return render(request, 'accounting/expense_report.html', context)

@login_required
def income_report(request):
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    account_id = request.GET.get('account')
    
    # Базовый запрос
    transactions = Transaction.objects.filter(transaction_type='income')
    
    # Применяем фильтры
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    if account_id:
        transactions = transactions.filter(account_id=account_id)
    
    # Группировка по категориям
    categories = Category.objects.filter(
        transactions__in=transactions
    ).annotate(
        total=Sum('transactions__amount'),
        count=Count('transactions')
    ).order_by('-total')
    
    total_income = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Данные для фильтров
    accounts = Account.objects.filter(is_active=True)
    
    context = {
        'categories': categories,
        'total_income': total_income,
        'accounts': accounts,
        'date_ranges': get_date_range_options(),
    }
    return render(request, 'accounting/income_report.html', context)

@login_required
def cash_flow(request):
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    account_id = request.GET.get('account')
    period = request.GET.get('period', 'month')  # По умолчанию группировка по месяцам
    
    # Базовый запрос для доходов и расходов
    income_query = Transaction.objects.filter(transaction_type='income')
    expense_query = Transaction.objects.filter(transaction_type='expense')
    
    # Применяем фильтры
    if date_from:
        income_query = income_query.filter(date__gte=date_from)
        expense_query = expense_query.filter(date__gte=date_from)
    
    if date_to:
        income_query = income_query.filter(date__lte=date_to)
        expense_query = expense_query.filter(date__lte=date_to)
    
    if account_id:
        income_query = income_query.filter(account_id=account_id)
        expense_query = expense_query.filter(account_id=account_id)
    
    # Расчет общих сумм для статистики
    total_income = income_query.aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = expense_query.aggregate(Sum('amount'))['amount__sum'] or 0
    net_flow = total_income - total_expense
    
    # Получаем все счета для фильтра
    accounts = Account.objects.filter(is_active=True)
    
    # Данные для графика
    # Логика группировки по периодам реализуется на стороне клиента через API
    
    context = {
        'accounts': accounts,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_flow': net_flow,
        'date_ranges': get_date_range_options(),
        'period': period,
    }
    return render(request, 'accounting/cash_flow.html', context)

@login_required
def cash_flow_report(request):
    """Представление для отчета о движении денежных средств"""
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    account_id = request.GET.get('account')
    period = request.GET.get('period', 'month')  # По умолчанию группировка по месяцам
    
    # Устанавливаем значения по умолчанию, если они не указаны
    today = timezone.now().date()
    if not date_to:
        date_to = today
    if not date_from:
        # По умолчанию последние 3 месяца
        date_from = today - timezone.timedelta(days=90)
    
    # Базовый запрос для доходов и расходов
    income_query = Transaction.objects.filter(transaction_type='income')
    expense_query = Transaction.objects.filter(transaction_type='expense')
    
    # Применяем фильтры
    if date_from:
        income_query = income_query.filter(date__gte=date_from)
        expense_query = expense_query.filter(date__gte=date_from)
    
    if date_to:
        income_query = income_query.filter(date__lte=date_to)
        expense_query = expense_query.filter(date__lte=date_to)
    
    if account_id:
        income_query = income_query.filter(account_id=account_id)
        expense_query = expense_query.filter(account_id=account_id)
    
    # Расчет общих сумм для статистики
    total_income = income_query.aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = expense_query.aggregate(Sum('amount'))['amount__sum'] or 0
    net_flow = total_income - total_expense
    
    # Получаем все счета для фильтра
    accounts = Account.objects.all()
    
    # Рассчитываем начальный и конечный баланс
    opening_balance = 0
    if account_id:
        # Получаем начальный баланс указанного счета
        try:
            account = Account.objects.get(id=account_id)
            # Начальный баланс - это баланс до начала периода отчета
            opening_balance = account.initial_balance
            
            # Добавляем транзакции до начала периода
            if date_from:
                prev_income = Transaction.objects.filter(
                    transaction_type='income',
                    account_id=account_id,
                    date__lt=date_from
                ).aggregate(Sum('amount'))['amount__sum'] or 0
                
                prev_expense = Transaction.objects.filter(
                    transaction_type='expense',
                    account_id=account_id,
                    date__lt=date_from
                ).aggregate(Sum('amount'))['amount__sum'] or 0
                
                opening_balance += (prev_income - prev_expense)
        except Account.DoesNotExist:
            pass
    else:
        # Получаем общий начальный баланс всех счетов
        opening_balance = Account.objects.aggregate(Sum('initial_balance'))['initial_balance__sum'] or 0
        
        # Добавляем транзакции до начала периода
        if date_from:
            prev_income = Transaction.objects.filter(
                transaction_type='income',
                date__lt=date_from
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            prev_expense = Transaction.objects.filter(
                transaction_type='expense',
                date__lt=date_from
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            opening_balance += (prev_income - prev_expense)
    
    # Конечный баланс = начальный баланс + чистый поток
    ending_balance = opening_balance + net_flow
    
    # Подготовка данных для графика по периодам
    chart_data = {
        'labels': [],
        'income': [],
        'expense': [],
        'balance': []
    }
    
    # Группировка данных по периодам
    if period == 'month':
        # Группировка по месяцам
        income_by_period = income_query.annotate(
            period=TruncMonth('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
        
        expense_by_period = expense_query.annotate(
            period=TruncMonth('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
    elif period == 'week':
        # Группировка по неделям
        income_by_period = income_query.annotate(
            period=TruncWeek('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
        
        expense_by_period = expense_query.annotate(
            period=TruncWeek('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
    elif period == 'day':
        # Группировка по дням
        income_by_period = income_query.annotate(
            period=TruncDay('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
        
        expense_by_period = expense_query.annotate(
            period=TruncDay('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
    else:  # year
        # Группировка по годам
        income_by_period = income_query.annotate(
            period=TruncYear('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
        
        expense_by_period = expense_query.annotate(
            period=TruncYear('date')
        ).values('period').annotate(total=Sum('amount')).order_by('period')
    
    # Преобразование QuerySet в словари для удобства
    income_dict = {item['period']: item['total'] for item in income_by_period}
    expense_dict = {item['period']: item['total'] for item in expense_by_period}
    
    # Получаем все уникальные периоды
    all_periods = sorted(set(list(income_dict.keys()) + list(expense_dict.keys())))
    
    # Заполняем данные для графика
    running_balance = opening_balance
    for period_date in all_periods:
        # Форматируем метку в зависимости от выбранного периода
        if period == 'month':
            label = period_date.strftime('%b %Y')
        elif period == 'week':
            label = f"{period_date.strftime('%d/%m/%Y')} - {(period_date + timezone.timedelta(days=6)).strftime('%d/%m/%Y')}"
        elif period == 'day':
            label = period_date.strftime('%d/%m/%Y')
        else:  # year
            label = period_date.strftime('%Y')
        
        income_amount = income_dict.get(period_date, 0)
        expense_amount = expense_dict.get(period_date, 0)
        net_amount = income_amount - expense_amount
        running_balance += net_amount
        
        chart_data['labels'].append(label)
        chart_data['income'].append(float(income_amount))
        chart_data['expense'].append(float(expense_amount))
        chart_data['balance'].append(float(running_balance))
    
    # Подготовка данных для таблицы по периодам
    table_data = []
    running_balance = opening_balance
    for period_date in all_periods:
        income_amount = income_dict.get(period_date, 0)
        expense_amount = expense_dict.get(period_date, 0)
        net_amount = income_amount - expense_amount
        
        opening_balance_for_period = running_balance
        running_balance += net_amount
        closing_balance = running_balance
        
        # Форматируем метку в зависимости от выбранного периода
        if period == 'month':
            label = period_date.strftime('%b %Y')
        elif period == 'week':
            label = f"{period_date.strftime('%d/%m/%Y')} - {(period_date + timezone.timedelta(days=6)).strftime('%d/%m/%Y')}"
        elif period == 'day':
            label = period_date.strftime('%d/%m/%Y')
        else:  # year
            label = period_date.strftime('%Y')
        
        table_data.append({
            'period': label,
            'opening_balance': opening_balance_for_period,
            'income': income_amount,
            'expense': expense_amount,
            'net_flow': net_amount,
            'closing_balance': closing_balance
        })
    
    context = {
        'accounts': accounts,
        'selected_account': int(account_id) if account_id else None,
        'date_from': date_from,
        'date_to': date_to,
        'period': period,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_flow': net_flow,
        'opening_balance': opening_balance,
        'ending_balance': ending_balance,
        'chart_data': json.dumps(chart_data),
        'table_data': table_data,
    }
    
    return render(request, 'accounting/cash_flow.html', context)

@login_required
def budget(request):
    # Получаем все бюджеты
    budgets = Budget.objects.all()
    
    # Общая статистика
    total_budget = budgets.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Расчет процентов использования для каждого бюджета
    for budget in budgets:
        budget.spent_percentage = (budget.spent_amount / budget.amount * 100) if budget.amount > 0 else 0
    
    context = {
        'budgets': budgets,
        'total_budget': total_budget,
    }
    return render(request, 'accounting/budget.html', context)

# Представления для профиля и настроек
@login_required
def profile(request):
    return render(request, 'accounting/profile.html')

@login_required
def settings(request):
    return render(request, 'accounting/settings.html')

# Представление для экспорта данных
@login_required
def export_data(request):
    # Здесь можно получить список резервных копий, если они есть
    backups = []  # Здесь должна быть модель для хранения резервных копий
    
    context = {
        'backups': backups
    }
    return render(request, 'accounting/export.html', context)

# API для данных графиков
@login_required
def chart_data(request):
    """API для получения данных для графиков"""
    chart_type = request.GET.get('type', 'cash_flow')
    period = request.GET.get('period', 'month')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    account_id = request.GET.get('account')
    
    # Здесь реализуем логику подготовки данных в зависимости от типа графика
    # и запрошенных параметров
    data = {
        'labels': [],
        'datasets': []
    }
    
    # Пример: данные о денежных потоках по месяцам
    if chart_type == 'cash_flow':
        # Логика группировки данных по периодам (месяц, неделя, год)
        # и подготовка данных для графика
        pass
    
    return JsonResponse(data)

@login_required
def category_data(request):
    """API для получения данных по категориям для диаграмм"""
    transaction_type = request.GET.get('type', 'expense')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Базовый запрос
    transactions = Transaction.objects.filter(transaction_type=transaction_type)
    
    # Применяем фильтры
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    # Группировка по категориям
    categories = Category.objects.filter(
        transactions__in=transactions
    ).annotate(
        total=Sum('transactions__amount')
    ).order_by('-total')
    
    # Подготовка данных для диаграммы
    data = {
        'labels': [category.name for category in categories],
        'datasets': [{
            'data': [float(category.total) for category in categories],
            'backgroundColor': [category.color or '#' + ''.join([format(c, '02x') for c in range(100, 255, 50)[:3]]) for category in categories]
        }]
    }
    
    return JsonResponse(data)

# Функции экспорта данных
@login_required
def export_transactions(request):
    """Экспорт транзакций в выбранном формате"""
    # Получаем параметры
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    export_format = request.GET.get('format', 'excel')
    
    # Фильтруем транзакции по датам
    transactions = Transaction.objects.all()
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    # Сортируем по дате
    transactions = transactions.order_by('-date')
    
    # Экспорт в зависимости от формата
    if export_format == 'excel':
        return export_transactions_excel(transactions)
    elif export_format == 'csv':
        return export_transactions_csv(transactions)
    elif export_format == 'pdf':
        return export_transactions_pdf(transactions)
    else:
        messages.error(request, 'Неподдерживаемый формат экспорта')
        return redirect('accounting:export')

def export_transactions_excel(transactions):
    """Экспорт транзакций в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    
    # Заголовки
    headers = [
        'ID', 'Дата', 'Тип', 'Счет', 'Счет получателя', 'Категория', 
        'Сумма', 'Описание', 'Теги', 'Дата создания'
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Данные
    for row_num, transaction in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=transaction.id)
        ws.cell(row=row_num, column=2, value=transaction.date)
        ws.cell(row=row_num, column=3, value=transaction.get_transaction_type_display())
        ws.cell(row=row_num, column=4, value=transaction.account.name if transaction.account else '')
        ws.cell(row=row_num, column=5, value=transaction.to_account.name if transaction.to_account else '')
        ws.cell(row=row_num, column=6, value=transaction.category.name if transaction.category else '')
        ws.cell(row=row_num, column=7, value=float(transaction.amount))
        ws.cell(row=row_num, column=8, value=transaction.description)
        ws.cell(row=row_num, column=9, value=transaction.tags)
        ws.cell(row=row_num, column=10, value=transaction.created_at)
    
    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
    
    # Сохраняем книгу в response
    wb.save(response)
    
    return response

def export_transactions_csv(transactions):
    """Экспорт транзакций в CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=transactions.csv'
    
    writer = csv.writer(response)
    
    # Заголовки
    headers = [
        'ID', 'Дата', 'Тип', 'Счет', 'Счет получателя', 'Категория', 
        'Сумма', 'Описание', 'Теги', 'Дата создания'
    ]
    writer.writerow(headers)
    
    # Данные
    for transaction in transactions:
        writer.writerow([
            transaction.id,
            transaction.date,
            transaction.get_transaction_type_display(),
            transaction.account.name if transaction.account else '',
            transaction.to_account.name if transaction.to_account else '',
            transaction.category.name if transaction.category else '',
            float(transaction.amount),
            transaction.description,
            transaction.tags,
            transaction.created_at
        ])
    
    return response

def export_transactions_pdf(transactions):
    """Экспорт транзакций в PDF"""
    # Здесь должна быть логика для создания PDF с использованием библиотеки
    # Например, reportlab, weasyprint, xhtml2pdf или pdfkit
    
    # Это заглушка, возвращаем Excel вместо PDF
    # Сделаем сообщение внутри Excel файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    
    ws.cell(row=1, column=1, value="ВНИМАНИЕ: Экспорт в PDF временно недоступен. Данные экспортированы в Excel.")
    
    # Заголовки
    headers = [
        'ID', 'Дата', 'Тип', 'Счет', 'Счет получателя', 'Категория', 
        'Сумма', 'Описание', 'Теги', 'Дата создания'
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_num, value=header)
    
    # Данные
    for row_num, transaction in enumerate(transactions, 4):
        ws.cell(row=row_num, column=1, value=transaction.id)
        ws.cell(row=row_num, column=2, value=transaction.date)
        ws.cell(row=row_num, column=3, value=transaction.get_transaction_type_display())
        ws.cell(row=row_num, column=4, value=transaction.account.name if transaction.account else '')
        ws.cell(row=row_num, column=5, value=transaction.to_account.name if transaction.to_account else '')
        ws.cell(row=row_num, column=6, value=transaction.category.name if transaction.category else '')
        ws.cell(row=row_num, column=7, value=float(transaction.amount))
        ws.cell(row=row_num, column=8, value=transaction.description)
        ws.cell(row=row_num, column=9, value=transaction.tags)
        ws.cell(row=row_num, column=10, value=transaction.created_at)
    
    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
    
    # Сохраняем книгу в response
    wb.save(response)
    
    return response

@login_required
def export_accounts(request):
    """Экспорт счетов в выбранном формате"""
    export_format = request.GET.get('format', 'excel')
    include_inactive = request.GET.get('include_inactive') == '1'
    
    # Получаем счета
    accounts = Account.objects.all()
    
    if not include_inactive:
        accounts = accounts.filter(is_active=True)
    
    # Экспорт в зависимости от формата
    if export_format == 'excel':
        return export_accounts_excel(accounts)
    elif export_format == 'csv':
        return export_accounts_csv(accounts)
    elif export_format == 'pdf':
        return export_accounts_pdf(accounts)
    else:
        messages.error(request, 'Неподдерживаемый формат экспорта')
        return redirect('accounting:export')

def export_accounts_excel(accounts):
    """Экспорт счетов в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Счета"
    
    # Заголовки
    headers = [
        'ID', 'Код', 'Название', 'Категория', 'Описание', 'Активен', 
        'Текущий баланс', 'Дата создания', 'Дата обновления'
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Данные
    for row_num, account in enumerate(accounts, 2):
        ws.cell(row=row_num, column=1, value=account.id)
        ws.cell(row=row_num, column=2, value=account.code)
        ws.cell(row=row_num, column=3, value=account.name)
        ws.cell(row=row_num, column=4, value=account.category.name if account.category else '')
        ws.cell(row=row_num, column=5, value=account.description)
        ws.cell(row=row_num, column=6, value='Да' if account.is_active else 'Нет')
        ws.cell(row=row_num, column=7, value=float(account.get_balance()))
        ws.cell(row=row_num, column=8, value=account.created_at)
        ws.cell(row=row_num, column=9, value=account.updated_at)
    
    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=accounts.xlsx'
    
    # Сохраняем книгу в response
    wb.save(response)
    
    return response

def export_accounts_csv(accounts):
    """Экспорт счетов в CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=accounts.csv'
    
    writer = csv.writer(response)
    
    # Заголовки
    headers = [
        'ID', 'Код', 'Название', 'Категория', 'Описание', 'Активен', 
        'Текущий баланс', 'Дата создания', 'Дата обновления'
    ]
    writer.writerow(headers)
    
    # Данные
    for account in accounts:
        writer.writerow([
            account.id,
            account.code,
            account.name,
            account.category.name if account.category else '',
            account.description,
            'Да' if account.is_active else 'Нет',
            float(account.get_balance()),
            account.created_at,
            account.updated_at
        ])
    
    return response

def export_accounts_pdf(accounts):
    """Экспорт счетов в PDF"""
    # Аналогично - заглушка вместо PDF
    wb = Workbook()
    ws = wb.active
    ws.title = "Счета"
    
    ws.cell(row=1, column=1, value="ВНИМАНИЕ: Экспорт в PDF временно недоступен. Данные экспортированы в Excel.")
    
    # Заголовки
    headers = [
        'ID', 'Код', 'Название', 'Категория', 'Описание', 'Активен', 
        'Текущий баланс', 'Дата создания', 'Дата обновления'
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_num, value=header)
    
    # Данные
    for row_num, account in enumerate(accounts, 4):
        ws.cell(row=row_num, column=1, value=account.id)
        ws.cell(row=row_num, column=2, value=account.code)
        ws.cell(row=row_num, column=3, value=account.name)
        ws.cell(row=row_num, column=4, value=account.category.name if account.category else '')
        ws.cell(row=row_num, column=5, value=account.description)
        ws.cell(row=row_num, column=6, value='Да' if account.is_active else 'Нет')
        ws.cell(row=row_num, column=7, value=float(account.get_balance()))
        ws.cell(row=row_num, column=8, value=account.created_at)
        ws.cell(row=row_num, column=9, value=account.updated_at)
    
    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=accounts.xlsx'
    
    # Сохраняем книгу в response
    wb.save(response)
    
    return response

@login_required
def export_report(request):
    """Экспорт отчета в выбранном формате"""
    report_type = request.GET.get('type', 'income')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    export_format = request.GET.get('format', 'excel')
    
    # В зависимости от типа отчета вызываем разные функции
    if report_type == 'income':
        return export_income_report(date_from, date_to, export_format)
    elif report_type == 'expense':
        return export_expense_report(date_from, date_to, export_format)
    elif report_type == 'cash_flow':
        return export_cash_flow_report(date_from, date_to, export_format)
    elif report_type == 'budget':
        return export_budget_report(export_format)
    elif report_type == 'balance':
        return export_balance_report(date_to, export_format)
    else:
        messages.error(request, 'Неподдерживаемый тип отчета')
        return redirect('accounting:export')

def export_income_report(date_from, date_to, format):
    """Экспорт отчета о доходах"""
    # Получаем данные для отчета
    transactions = Transaction.objects.filter(transaction_type='income')
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    # Группировка по категориям
    categories = Category.objects.filter(
        transactions__in=transactions
    ).annotate(
        total=Sum('transactions__amount'),
        count=Count('transactions')
    ).order_by('-total')
    
    # Экспорт в зависимости от формата
    if format == 'excel':
        return export_income_report_excel(categories, transactions, date_from, date_to)
    elif format == 'csv':
        return export_income_report_csv(categories, transactions, date_from, date_to)
    else:
        # Возвращаем Excel как запасной вариант
        return export_income_report_excel(categories, transactions, date_from, date_to)

def export_income_report_excel(categories, transactions, date_from, date_to):
    """Экспорт отчета о доходах в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет о доходах"
    
    # Заголовок отчета
    ws.cell(row=1, column=1, value="Отчет о доходах")
    ws.cell(row=2, column=1, value=f"Период: {date_from or 'Все время'} - {date_to or 'Все время'}")
    
    # Общая сумма
    total_income = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    ws.cell(row=3, column=1, value=f"Общая сумма доходов: {total_income}")
    
    # Заголовки для таблицы категорий
    headers = ['Категория', 'Количество транзакций', 'Сумма', '% от общего']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)
    
    # Данные по категориям
    for row_num, category in enumerate(categories, 6):
        percentage = (category.total / total_income * 100) if total_income > 0 else 0
        
        ws.cell(row=row_num, column=1, value=category.name)
        ws.cell(row=row_num, column=2, value=category.count)
        ws.cell(row=row_num, column=3, value=float(category.total))
        ws.cell(row=row_num, column=4, value=f"{percentage:.2f}%")
    
    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=income_report.xlsx'
    
    # Сохраняем книгу в response
    wb.save(response)
    
    return response

def export_income_report_csv(categories, transactions, date_from, date_to):
    """Экспорт отчета о доходах в CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=income_report.csv'
    
    writer = csv.writer(response)
    
    # Заголовок отчета
    writer.writerow(["Отчет о доходах"])
    writer.writerow([f"Период: {date_from or 'Все время'} - {date_to or 'Все время'}"])
    
    # Общая сумма
    total_income = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    writer.writerow([f"Общая сумма доходов: {total_income}"])
    writer.writerow([])  # Пустая строка
    
    # Заголовки для таблицы категорий
    writer.writerow(['Категория', 'Количество транзакций', 'Сумма', '% от общего'])
    
    # Данные по категориям
    for category in categories:
        percentage = (category.total / total_income * 100) if total_income > 0 else 0
        writer.writerow([
            category.name,
            category.count,
            float(category.total),
            f"{percentage:.2f}%"
        ])
    
    return response

def export_expense_report(date_from, date_to, format):
    """Экспорт отчета о расходах"""
    # Аналогично отчету о доходах, но для расходов
    transactions = Transaction.objects.filter(transaction_type='expense')
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    # Группировка по категориям
    categories = Category.objects.filter(
        transactions__in=transactions
    ).annotate(
        total=Sum('transactions__amount'),
        count=Count('transactions')
    ).order_by('-total')
    
    # Экспорт в зависимости от формата
    if format == 'excel':
        return export_expense_report_excel(categories, transactions, date_from, date_to)
    elif format == 'csv':
        return export_expense_report_csv(categories, transactions, date_from, date_to)
    else:
        return export_expense_report_excel(categories, transactions, date_from, date_to)

def export_expense_report_excel(categories, transactions, date_from, date_to):
    """Экспорт отчета о расходах в Excel"""
    # Аналогично отчету о доходах
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет о расходах"
    
    # Заголовок отчета
    ws.cell(row=1, column=1, value="Отчет о расходах")
    ws.cell(row=2, column=1, value=f"Период: {date_from or 'Все время'} - {date_to or 'Все время'}")
    
    # Общая сумма
    total_expense = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    ws.cell(row=3, column=1, value=f"Общая сумма расходов: {total_expense}")
    
    # Заголовки для таблицы категорий
    headers = ['Категория', 'Количество транзакций', 'Сумма', '% от общего']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=5, column=col_num, value=header)
    
    # Данные по категориям
    for row_num, category in enumerate(categories, 6):
        percentage = (category.total / total_expense * 100) if total_expense > 0 else 0
        
        ws.cell(row=row_num, column=1, value=category.name)
        ws.cell(row=row_num, column=2, value=category.count)
        ws.cell(row=row_num, column=3, value=float(category.total))
        ws.cell(row=row_num, column=4, value=f"{percentage:.2f}%")
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expense_report.xlsx'
    wb.save(response)
    
    return response

def export_expense_report_csv(categories, transactions, date_from, date_to):
    """Экспорт отчета о расходах в CSV"""
    # Аналогично отчету о доходах
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=expense_report.csv'
    
    writer = csv.writer(response)
    
    writer.writerow(["Отчет о расходах"])
    writer.writerow([f"Период: {date_from or 'Все время'} - {date_to or 'Все время'}"])
    
    total_expense = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    writer.writerow([f"Общая сумма расходов: {total_expense}"])
    writer.writerow([])
    
    writer.writerow(['Категория', 'Количество транзакций', 'Сумма', '% от общего'])
    
    for category in categories:
        percentage = (category.total / total_expense * 100) if total_expense > 0 else 0
        writer.writerow([
            category.name,
            category.count,
            float(category.total),
            f"{percentage:.2f}%"
        ])
    
    return response

def export_cash_flow_report(date_from, date_to, format):
    """Экспорт отчета о движении денежных средств"""
    # Заглушка для демонстрации
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cash_flow_report.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Движение денежных средств"
    
    ws.cell(row=1, column=1, value="Отчет о движении денежных средств")
    ws.cell(row=2, column=1, value=f"Период: {date_from or 'Все время'} - {date_to or 'Все время'}")
    
    wb.save(response)
    return response

def export_budget_report(format):
    """Экспорт отчета о бюджете"""
    # Заглушка для демонстрации
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=budget_report.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет о бюджете"
    
    ws.cell(row=1, column=1, value="Отчет о бюджете")
    
    wb.save(response)
    return response

def export_balance_report(date_to, format):
    """Экспорт отчета о балансе"""
    # Заглушка для демонстрации
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=balance_report.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет о балансе"
    
    ws.cell(row=1, column=1, value="Отчет о балансе")
    ws.cell(row=2, column=1, value=f"На дату: {date_to or timezone.now().date()}")
    
    wb.save(response)
    return response

# Функции для работы с резервными копиями
@login_required
@require_POST
def create_backup(request):
    """Создание резервной копии данных"""
    try:
        # Здесь должна быть логика создания резервной копии
        # Это заглушка для демонстрации
        messages.success(request, 'Резервная копия успешно создана')
    except Exception as e:
        messages.error(request, f'Ошибка при создании резервной копии: {e}')
    
    return redirect('accounting:export')

@login_required
def download_backup(request, pk):
    """Скачивание резервной копии"""
    # Заглушка для демонстрации
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename=backup_{pk}.zip'
    
    # Здесь должна быть логика для получения файла резервной копии
    
    return response

@login_required
def restore_backup(request, pk):
    """Восстановление из резервной копии"""
    try:
        # Здесь должна быть логика восстановления из резервной копии
        # Это заглушка для демонстрации
        messages.success(request, 'Данные успешно восстановлены из резервной копии')
    except Exception as e:
        messages.error(request, f'Ошибка при восстановлении данных: {e}')
    
    return redirect('accounting:export')

@login_required
def delete_backup(request, pk):
    """Удаление резервной копии"""
    try:
        # Здесь должна быть логика удаления резервной копии
        # Это заглушка для демонстрации
        messages.success(request, 'Резервная копия успешно удалена')
    except Exception as e:
        messages.error(request, f'Ошибка при удалении резервной копии: {e}')
    
    return redirect('accounting:export')