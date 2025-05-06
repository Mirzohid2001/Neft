from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, DecimalField
from django.utils import timezone
from datetime import datetime, timedelta
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from apps.accounts.views import infrastructure_required
from .models import (
    Product, Receiving, Giving, Stock,
    CanteenExpense, Project, ProjectItem, ReceivingItem, ProjectProduct,
    Order, OrderItem, TelegramGroup, OrderProduct
)
from .forms import (
    ProductForm, ReceivingForm, GivingForm,
    CanteenExpenseForm, ProjectForm, ProjectItemForm, ProjectProductForm,
    OrderForm, OrderItemFormSet
)
import json
from django.db.models.deletion import ProtectedError
from django.db.models.fields.json import JSONField
from django.db.models.fields.related import (
    ForeignKey, ManyToManyField, OneToOneField
)
from django.db import models
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import requests
from django.forms import modelformset_factory
from django.urls import reverse_lazy
from django.views.decorators.http import require_POST
from django.conf import settings

@login_required
@infrastructure_required
def dashboard(request):
    # Get today's date
    today = timezone.now().date()
    
    # Get daily expenses for receiving and giving
    today_receivings = Receiving.objects.filter(date=today)
    today_givings = Giving.objects.filter(date=today)
    
    # Calculate totals for today's receivings
    total_received = 0
    for receiving in today_receivings:
        for item in receiving.items.all():
            total_received += item.quantity * item.unit_price
    
    # Calculate totals for today's givings
    total_given = 0
    for giving in Giving.objects.filter(date=today):
        total_given += giving.quantity * giving.product.unit_price
    
    # Get the current month data for canteen expenses
    current_month = today.month
    current_year = today.year
    start_date = datetime(current_year, current_month, 1).date()
    if current_month == 12:
        end_date = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(current_year, current_month + 1, 1).date() - timedelta(days=1)
    
    # Get monthly canteen expenses total
    canteen_monthly_expenses = CanteenExpense.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).aggregate(monthly_total=Sum(models.F('quantity') * models.F('unit_price')))['monthly_total'] or 0
    
    # Get total stock value and recent stocks
    stocks = Stock.objects.select_related('product').all()
    total_stock_value = sum(stock.quantity * stock.product.unit_price for stock in stocks)
    recent_stocks = stocks.order_by('-last_updated')[:10] if hasattr(Stock, 'last_updated') else stocks[:10]
    
    # Get active projects and their costs
    projects = Project.objects.filter(status='in_progress')
    
    # Calculate costs for each project
    project_costs = {}
    for project in projects:
        project_total = sum(pp.total_cost for pp in ProjectProduct.objects.filter(project=project))
        project_costs[project.id] = project_total
    
    # Recent activities (combined from receiving, giving, and canteen expenses)
    recent_activities = []
    
    # Add receivings to activities
    for receiving in Receiving.objects.order_by('-date', '-created_at')[:5]:
        total_amount = sum(item.quantity * item.unit_price for item in receiving.items.all())
        recent_activities.append({
            'date': receiving.date,
            'description': f'–ü–æ—Å—Ç—É–ø–ª–µ–Ω–∏–µ #{receiving.id}',
            'amount': total_amount
        })
    
    # Add givings to activities
    for giving in Giving.objects.order_by('-date', '-created_at')[:5]:
        recent_activities.append({
            'date': giving.date,
            'description': f'–í—ã–¥–∞—á–∞: {giving.product.name}',
            'amount': giving.quantity * giving.product.unit_price
        })
    
    # Add canteen expenses to activities
    for expense in CanteenExpense.objects.order_by('-date', '-created_at')[:5]:
        recent_activities.append({
            'date': expense.date,
            'description': f'–†–∞—Å—Ö–æ–¥ —Å—Ç–æ–ª–æ–≤–æ–π: {expense.product}',
            'amount': expense.quantity * expense.unit_price
        })
    
    # Sort activities by date (newest first)
    recent_activities.sort(key=lambda x: x['date'], reverse=True)
    recent_activities = recent_activities[:10]  # Keep only the 10 most recent
    
    # Generate monthly statistics for the last 6 months
    monthly_stats = []
    
    for i in range(5, -1, -1):  # From 5 months ago to current month
        # Calculate month and year
        stats_month = current_month - i
        stats_year = current_year
        
        while stats_month <= 0:
            stats_month += 12
            stats_year -= 1
        
        # Get month start and end dates
        month_start = datetime(stats_year, stats_month, 1).date()
        if stats_month == 12:
            month_end = datetime(stats_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(stats_year, stats_month + 1, 1).date() - timedelta(days=1)
        
        # Calculate receiving total for the month
        month_receiving_total = 0
        for receiving in Receiving.objects.filter(date__gte=month_start, date__lte=month_end):
            for item in receiving.items.all():
                month_receiving_total += item.quantity * item.unit_price
        
        # Calculate giving total for the month
        month_giving_total = 0
        for giving in Giving.objects.filter(date__gte=month_start, date__lte=month_end):
            month_giving_total += giving.quantity * giving.product.unit_price
        
        # Calculate canteen total for the month
        month_canteen_total = CanteenExpense.objects.filter(
            date__gte=month_start, 
            date__lte=month_end
        ).aggregate(total=Sum(models.F('quantity') * models.F('unit_price')))['total'] or 0
        
        # Calculate month balance
        month_balance = month_receiving_total - month_giving_total - month_canteen_total
        
        # Get month name
        month_name = month_start.strftime('%B %Y')
        
        # Add to stats
        monthly_stats.append({
            'month_name': month_name,
            'receiving_total': month_receiving_total,
            'giving_total': month_giving_total,
            'canteen_total': month_canteen_total,
            'balance': month_balance
        })
    
    return render(request, 'infrastruction/dashboard.html', {
        'total_received': total_received,
        'total_given': total_given,
        'canteen_monthly_total': canteen_monthly_expenses,
        'total_stock_value': total_stock_value,
        'projects': projects,
        'project_costs': project_costs,
        'recent_stocks': recent_stocks,
        'recent_activities': recent_activities,
        'monthly_stats': monthly_stats
    })

@login_required
@infrastructure_required
def product_list(request):
    # Get all products with their related stock information
    products = Product.objects.all().prefetch_related('stock')
    
    # Create dictionaries for easier access to stock information and totals
    stock_dict = {}
    total_value_dict = {}
    comments_dict = {}
    
    for stock in Stock.objects.select_related('product').all():
        stock_dict[stock.product_id] = stock.quantity
        total_value_dict[stock.product_id] = stock.quantity * stock.product.unit_price
    
    # Get the most recent comments for each product from receiving transactions
    for product in products:
        # Get the most recent receiving items with comments for this product
        recent_items = ReceivingItem.objects.filter(
            product_id=product.id, 
            comment__isnull=False
        ).exclude(
            comment__exact=''
        ).order_by('-receiving__date', '-created_at')[:3]
        
        if recent_items:
            comments_dict[product.id] = recent_items
    
    # Calculate the grand total
    grand_total = sum(total_value_dict.values())
    
    return render(request, 'infrastruction/product_list.html', {
        'products': products,
        'stock_dict': stock_dict,
        'total_value_dict': total_value_dict,
        'comments_dict': comments_dict,
        'grand_total': grand_total,
        'title': '–°–ø–∏—Å–æ–∫ –ø—Ä–æ–¥—É–∫—Ç–æ–≤'
    })

@login_required
@infrastructure_required
def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '–ü—Ä–æ–¥—É–∫—Ç —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω.')
            return redirect('infrastruction:product_list')
    else:
        form = ProductForm()
    
    return render(request, 'infrastruction/product_form.html', {
        'form': form,
        'title': '–î–æ–±–∞–≤–∏—Ç—å –ø—Ä–æ–¥—É–∫—Ç'
    })

@login_required
@infrastructure_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, '–ü—Ä–æ–¥—É–∫—Ç —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω.')
            return redirect('infrastruction:product_list')
    else:
        form = ProductForm(instance=product)
    
    return render(request, 'infrastruction/product_form.html', {
        'form': form,
        'title': '–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å –ø—Ä–æ–¥—É–∫—Ç'
    })

@login_required
@infrastructure_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    try:
        product.delete()
        messages.success(request, '–ü—Ä–æ–¥—É–∫—Ç —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω.')
    except ProtectedError:
        messages.error(request, '–ù–µ–≤–æ–∑–º–æ–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å –ø—Ä–æ–¥—É–∫—Ç, —Ç–∞–∫ –∫–∞–∫ –æ–Ω –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –≤ –¥—Ä—É–≥–∏—Ö –∑–∞–ø–∏—Å—è—Ö.')
    return redirect('infrastruction:product_list')

@login_required
@infrastructure_required
def receiving_list(request):
    receivings = Receiving.objects.all().order_by('-date')
    return render(request, 'infrastruction/receiving_list.html', {
        'receivings': receivings,
        'title': '–°–ø–∏—Å–æ–∫ –ø—Ä–∏—Ö–æ–¥–æ–≤'
    })

@login_required
@infrastructure_required
def receiving_form(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Extract data from the form
                date_str = request.POST.get('date')
                date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.now().date()
                notes = request.POST.get('notes', '')
                photo = request.FILES.get('photo')
                
                # Create the receiving record
                receiving = Receiving.objects.create(
                    date=date,
                    notes=notes,
                    photo=photo,
                    created_by=request.user
                )
                
                # Get products data from hidden input
                products_data = json.loads(request.POST.get('products_data', '[]'))
                
                if not products_data:
                    raise ValueError('–ù–µ–æ–±—Ö–æ–¥–∏–º–æ –¥–æ–±–∞–≤–∏—Ç—å —Ö–æ—Ç—è –±—ã –æ–¥–∏–Ω —Ç–æ–≤–∞—Ä')
                
                # Process each product
                for product_data in products_data:
                    product_name = product_data.get('product', '').strip()
                    product_id = product_data.get('product_id')
                    quantity = float(product_data.get('quantity', 0))
                    unit = product_data.get('unit', 'pcs')
                    unit_price = float(product_data.get('unit_price', 0))
                    comment = product_data.get('comment', '')
                    
                    if not product_name or quantity <= 0 or unit_price < 0:
                        raise ValueError('–ó–∞–ø–æ–ª–Ω–∏—Ç–µ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ –≤—Å–µ –ø–æ–ª—è —Ç–æ–≤–∞—Ä–∞')
                    
                    # Initialize created variable
                    created = False
                    
                    # Get or create product
                    if product_id:
                        try:
                            product = Product.objects.get(pk=product_id)
                        except Product.DoesNotExist:
                            product, created = Product.objects.get_or_create(
                                name=product_name,
                                defaults={'unit_price': unit_price}
                            )
                    else:
                        product, created = Product.objects.get_or_create(
                            name=product_name,
                            defaults={'unit_price': unit_price}
                        )
                    
                    # If product exists but price is different, update the price
                    if not created and product.unit_price != unit_price:
                        product.unit_price = unit_price
                        product.save()
                    
                    # Create receiving item
                    ReceivingItem.objects.create(
                        receiving=receiving,
                        product=product,
                        quantity=quantity,
                        unit=unit,
                        unit_price=unit_price,
                        comment=comment
                    )
                    
                    # Update stock
                    stock, created = Stock.objects.get_or_create(
                        product=product,
                        defaults={'quantity': quantity}
                    )
                    if not created:
                        stock.quantity += quantity
                        stock.save()
                
                # Send notification to Telegram
                send_receiving_to_telegram(receiving)
                
                messages.success(request, '–ü—Ä–∏—Ö–æ–¥ —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω')
                return redirect('infrastruction:receiving_list')
                
        except Exception as e:
            messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}')
            
    # For GET request or if there are errors in POST
    return render(request, 'infrastruction/receiving_form.html', {
        'title': '–î–æ–±–∞–≤–∏—Ç—å –ø—Ä–∏—Ö–æ–¥',
        'products': Product.objects.all()
    })

# Function to send notification to Telegram about new receiving
def send_receiving_to_telegram(receiving):
    """Send receiving notification to Telegram groups"""
    telegram_groups = TelegramGroup.objects.filter(is_active=True)
    
    if not telegram_groups:
        return
    
    # Format the message
    items_text = "\n".join([
        f"- {item.product.name}: {item.quantity} {item.get_unit_display()}"
        for item in receiving.items.all()
    ])
    
    message = f"""üì¶ *–ü–û–õ–£–ß–ï–ù–ò–ï –¢–û–í–ê–†–û–í* üì¶
*–î–∞—Ç–∞:* {receiving.date.strftime('%d.%m.%Y')}
*–°–æ–∑–¥–∞–ª:* {receiving.created_by.get_full_name() or receiving.created_by.username}

*–°–æ–¥–µ—Ä–∂–∏–º–æ–µ –ø–æ—Å—Ç—É–ø–ª–µ–Ω–∏—è:*
{items_text}

*–ü—Ä–∏–º–µ—á–∞–Ω–∏—è:* {receiving.notes or '–ù–µ—Ç –ø—Ä–∏–º–µ—á–∞–Ω–∏–π'}
"""
    
    # Send to each group
    for group in telegram_groups:
        try:
            # Send message with photo if available
            if receiving.photo:
                send_telegram_photo(group.chat_id, receiving.photo.path, message)
            else:
                send_telegram_message(group.chat_id, message)
        except Exception as e:
            # Log error, but continue with other groups
            print(f"Error sending to Telegram group {group.name}: {e}")

def send_telegram_photo(chat_id, photo_path, caption):
    """Send photo with caption to Telegram chat using bot API"""
    # Get bot token from settings
    bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
    
    if not bot_token:
        # Log error and return
        print("Telegram bot token not configured")
        return
    
    url = f"https://api.telegram.org/bot{bot_token}/sendPhoto"
    
    with open(photo_path, 'rb') as photo:
        files = {'photo': photo}
        data = {
            'chat_id': chat_id,
            'caption': caption,
            'parse_mode': 'Markdown'
        }
        
        response = requests.post(url, data=data, files=files)
    
    return response.json()

@login_required
@infrastructure_required
def product_details(request, pk):
    """API endpoint to get product details"""
    try:
        product = Product.objects.get(pk=pk)
        return JsonResponse({
            'id': product.id,
            'name': product.name,
            'unit_price': product.unit_price
        })
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

# Alias for receiving_form to match URL pattern
receiving_add = receiving_form

@login_required
@infrastructure_required
def receiving_detail(request, pk):
    receiving = get_object_or_404(Receiving, pk=pk)
    return render(request, 'infrastruction/receiving_detail.html', {
        'receiving': receiving,
        'title': f'–ü—Ä–∏—Ö–æ–¥ #{receiving.pk}'
    })

@login_required
@infrastructure_required
def receiving_edit(request, pk):
    receiving = get_object_or_404(Receiving, pk=pk)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Extract data from the form
                date_str = request.POST.get('date')
                date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.now().date()
                notes = request.POST.get('notes', '')
                
                # Handle photo update
                if 'photo' in request.FILES:
                    receiving.photo = request.FILES['photo']
                
                # Update the receiving record
                receiving.date = date
                receiving.notes = notes
                receiving.save()
                
                # Get products data from hidden input
                products_data = json.loads(request.POST.get('products_data', '[]'))
                
                if not products_data:
                    raise ValueError('–ù–µ–æ–±—Ö–æ–¥–∏–º–æ –¥–æ–±–∞–≤–∏—Ç—å —Ö–æ—Ç—è –±—ã –æ–¥–∏–Ω —Ç–æ–≤–∞—Ä')
                
                # First, keep track of all items to be kept
                kept_item_ids = []
                
                # Process each product
                for product_data in products_data:
                    item_id = product_data.get('id')
                    product_id = product_data.get('product_id')
                    product_name = product_data.get('product', '').strip()
                    quantity = float(product_data.get('quantity', 0))
                    unit = product_data.get('unit', 'pcs')
                    unit_price = float(product_data.get('unit_price', 0))
                    comment = product_data.get('comment', '')
                    
                    if not product_name or quantity <= 0 or unit_price < 0:
                        raise ValueError('–ó–∞–ø–æ–ª–Ω–∏—Ç–µ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ –≤—Å–µ –ø–æ–ª—è —Ç–æ–≤–∞—Ä–∞')
                    
                    # Initialize created variable
                    created = False
                    
                    # Get or create product
                    if product_id:
                        try:
                            product = Product.objects.get(pk=product_id)
                        except Product.DoesNotExist:
                            product, created = Product.objects.get_or_create(
                                name=product_name,
                                defaults={'unit_price': unit_price}
                            )
                    else:
                        product, created = Product.objects.get_or_create(
                            name=product_name,
                            defaults={'unit_price': unit_price}
                        )
                    
                    # If product exists but price is different, update the price
                    if not created and product.unit_price != unit_price:
                        product.unit_price = unit_price
                        product.save()
                    
                    if item_id:
                        # Update existing item
                        try:
                            item = ReceivingItem.objects.get(id=item_id, receiving=receiving)
                            
                            # Update stock if product or quantity changed
                            if item.product_id != product.id or item.quantity != quantity:
                                # Revert old quantity first
                                old_stock, _ = Stock.objects.get_or_create(
                                    product=item.product,
                                    defaults={'quantity': 0}
                                )
                                old_stock.quantity -= item.quantity
                                old_stock.save()
                                
                                # Add new quantity
                                new_stock, _ = Stock.objects.get_or_create(
                                    product=product,
                                    defaults={'quantity': 0}
                                )
                                new_stock.quantity += quantity
                                new_stock.save()
                            
                            # Update item
                            item.product = product
                            item.quantity = quantity
                            item.unit = unit
                            item.unit_price = unit_price
                            item.comment = comment
                            item.save()
                            kept_item_ids.append(item.id)
                            
                        except ReceivingItem.DoesNotExist:
                            # Shouldn't happen, but just in case - create new
                            item = ReceivingItem.objects.create(
                                receiving=receiving,
                                product=product,
                                quantity=quantity,
                                unit=unit,
                                unit_price=unit_price,
                                comment=comment
                            )
                            kept_item_ids.append(item.id)
                            
                            # Update stock
                            stock, created = Stock.objects.get_or_create(
                                product=product,
                                defaults={'quantity': quantity}
                            )
                            if not created:
                                stock.quantity += quantity
                                stock.save()
                    else:
                        # Create new item
                        item = ReceivingItem.objects.create(
                            receiving=receiving,
                            product=product,
                            quantity=quantity,
                            unit=unit,
                            unit_price=unit_price,
                            comment=comment
                        )
                        kept_item_ids.append(item.id)
                        
                        # Update stock
                        stock, created = Stock.objects.get_or_create(
                            product=product,
                            defaults={'quantity': quantity}
                        )
                        if not created:
                            stock.quantity += quantity
                            stock.save()
                
                # Remove items that are no longer in the form
                items_to_delete = receiving.items.exclude(id__in=kept_item_ids)
                for item in items_to_delete:
                    # Update stock
                    try:
                        stock = Stock.objects.get(product=item.product)
                        stock.quantity -= item.quantity
                        stock.save()
                    except Stock.DoesNotExist:
                        pass
                
                # Now actually delete the items
                items_to_delete.delete()
                
                messages.success(request, '–ü—Ä–∏—Ö–æ–¥ —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω')
                return redirect('infrastruction:receiving_list')
                
        except Exception as e:
            messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}')
    
    # For GET requests, we'll need to load the existing data
    items = list(receiving.items.all())
    print(f"DEBUG: Found {len(items)} items for receiving {pk}")
    for item in items:
        print(f"DEBUG: Item {item.id}: Product={item.product.name}, Quantity={item.quantity}, Comment={item.comment}")
    
    # Send the items to the template
    return render(request, 'infrastruction/receiving_form.html', {
        'title': '–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å –ø—Ä–∏—Ö–æ–¥',
        'receiving': receiving,
        'products': Product.objects.all(),
        'edit_mode': True
    })

@login_required
@infrastructure_required
def receiving_delete(request, pk):
    receiving = get_object_or_404(Receiving, pk=pk)
    try:
        with transaction.atomic():
            # Update stock quantities for all items
            for item in receiving.items.all():
                try:
                    stock = Stock.objects.get(product=item.product)
                    stock.quantity -= item.quantity
                    stock.save()
                except Stock.DoesNotExist:
                    pass
            
            # Delete the receiving record (which will cascade delete items)
            receiving.delete()
            
            messages.success(request, '–ü—Ä–∏—Ö–æ–¥ —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω.')
    except Exception as e:
        messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —É–¥–∞–ª–µ–Ω–∏–∏: {str(e)}')
    
    return redirect('infrastruction:receiving_list')

@login_required
@infrastructure_required
def canteen_expenses_list(request):
    expenses = CanteenExpense.objects.all().order_by('-date')
    return render(request, 'infrastruction/canteen_expenses_list.html', {
        'expenses': expenses,
        'title': '–†–∞—Å—Ö–æ–¥—ã —Å—Ç–æ–ª–æ–≤–æ–π'
    })

@login_required
@infrastructure_required
def canteen_expense_add(request):
    if request.method == 'POST':
        form = CanteenExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
            messages.success(request, '–†–∞—Å—Ö–æ–¥ —Å—Ç–æ–ª–æ–≤–æ–π —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω.')
            return redirect('infrastruction:canteen_expenses_list')
    else:
        form = CanteenExpenseForm(initial={'date': timezone.now().date()})
    
    return render(request, 'infrastruction/canteen_expense_form.html', {
        'form': form,
        'title': '–î–æ–±–∞–≤–∏—Ç—å —Ä–∞—Å—Ö–æ–¥ —Å—Ç–æ–ª–æ–≤–æ–π'
    })

@login_required
@infrastructure_required
def canteen_expense_edit(request, pk):
    expense = get_object_or_404(CanteenExpense, pk=pk)
    if request.method == 'POST':
        form = CanteenExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, '–†–∞—Å—Ö–æ–¥ —Å—Ç–æ–ª–æ–≤–æ–π —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω.')
            return redirect('infrastruction:canteen_expenses_list')
    else:
        form = CanteenExpenseForm(instance=expense)
    
    return render(request, 'infrastruction/canteen_expense_form.html', {
        'form': form,
        'title': '–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å —Ä–∞—Å—Ö–æ–¥ —Å—Ç–æ–ª–æ–≤–æ–π'
    })

@login_required
@infrastructure_required
def canteen_expense_detail(request, pk):
    expense = get_object_or_404(CanteenExpense, pk=pk)
    return render(request, 'infrastruction/canteen_expense_detail.html', {
        'expense': expense,
        'title': f'–†–∞—Å—Ö–æ–¥ #{expense.pk}'
    })

@login_required
@infrastructure_required
def canteen_expense_delete(request, pk):
    expense = get_object_or_404(CanteenExpense, pk=pk)
    expense.delete()
    messages.success(request, '–†–∞—Å—Ö–æ–¥ —Å—Ç–æ–ª–æ–≤–æ–π —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω.')
    return redirect('infrastruction:canteen_expenses_list')

@login_required
@infrastructure_required
def project_list(request):
    projects = Project.objects.all().order_by('-start_date')
    return render(request, 'infrastruction/project_list.html', {
        'projects': projects,
        'title': '–ü—Ä–æ–µ–∫—Ç—ã'
    })

@login_required
@infrastructure_required
def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk)
    items = project.projectitem_set.all()
    # Get all products used in this project
    products = project.used_products.all().select_related('product')
    
    # Calculate total product cost
    total_product_cost = sum(product.total_cost for product in products)
    
    return render(request, 'infrastruction/project_detail.html', {
        'project': project,
        'items': items,
        'products': products,
        'total_product_cost': total_product_cost,
        'title': project.name
    })

@login_required
@infrastructure_required
def project_add(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = request.user
            project.save()
            messages.success(request, '–ü—Ä–æ–µ–∫—Ç —É—Å–ø–µ—à–Ω–æ —Å–æ–∑–¥–∞–Ω.')
            return redirect('infrastruction:project_list')
    else:
        form = ProjectForm(initial={
            'start_date': timezone.now().date(),
            'end_date': timezone.now().date() + timezone.timedelta(days=30)
        })
    
    return render(request, 'infrastruction/project_form.html', {
        'form': form,
        'title': '–°–æ–∑–¥–∞—Ç—å –ø—Ä–æ–µ–∫—Ç'
    })

@login_required
@infrastructure_required
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, '–ü—Ä–æ–µ–∫—Ç —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω.')
            return redirect('infrastruction:project_list')
    else:
        form = ProjectForm(instance=project)
    
    return render(request, 'infrastruction/project_form.html', {
        'form': form,
        'title': '–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å –ø—Ä–æ–µ–∫—Ç'
    })

@login_required
@infrastructure_required
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)
    project.delete()
    messages.success(request, '–ü—Ä–æ–µ–∫—Ç —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω.')
    return redirect('infrastruction:project_list')

@login_required
@infrastructure_required
def project_item_add(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk)
    if request.method == 'POST':
        form = ProjectItemForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.project = project
            item.save()
            messages.success(request, '–≠–ª–µ–º–µ–Ω—Ç –ø—Ä–æ–µ–∫—Ç–∞ —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω.')
            return redirect('infrastruction:project_detail', pk=project_pk)
    else:
        form = ProjectItemForm()
    
    return render(request, 'infrastruction/project_item_form.html', {
        'form': form,
        'project': project,
        'title': '–î–æ–±–∞–≤–∏—Ç—å —ç–ª–µ–º–µ–Ω—Ç –ø—Ä–æ–µ–∫—Ç–∞'
    })

@login_required
@infrastructure_required
def project_item_edit(request, project_pk, pk):
    project = get_object_or_404(Project, pk=project_pk)
    item = get_object_or_404(ProjectItem, pk=pk, project=project)
    if request.method == 'POST':
        form = ProjectItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, '–≠–ª–µ–º–µ–Ω—Ç –ø—Ä–æ–µ–∫—Ç–∞ —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω.')
            return redirect('infrastruction:project_detail', pk=project_pk)
    else:
        form = ProjectItemForm(instance=item)
    
    return render(request, 'infrastruction/project_item_form.html', {
        'form': form,
        'project': project,
        'item': item,
        'title': '–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å —ç–ª–µ–º–µ–Ω—Ç –ø—Ä–æ–µ–∫—Ç–∞'
    })

@login_required
@infrastructure_required
def project_item_delete(request, project_pk, pk):
    project = get_object_or_404(Project, pk=project_pk)
    item = get_object_or_404(ProjectItem, pk=pk, project=project)
    item.delete()
    messages.success(request, '–≠–ª–µ–º–µ–Ω—Ç –ø—Ä–æ–µ–∫—Ç–∞ —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω.')
    return redirect('infrastruction:project_detail', pk=project_pk)

@login_required
@infrastructure_required
def giving_list(request):
    givings = Giving.objects.all().order_by('-date')
    return render(request, 'infrastruction/giving_list.html', {
        'givings': givings,
        'title': '–°–ø–∏—Å–æ–∫ —Ä–∞—Å—Ö–æ–¥–æ–≤'
    })

@login_required
@infrastructure_required
def giving_add(request):
    if request.method == 'POST':
        # Check if this is an AJAX request with JSON data
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                date = data.get('date')
                items = data.get('items', [])
                
                if not date or not items:
                    return JsonResponse({
                        'success': False,
                        'error': '–ù–µ–ø—Ä–∞–≤–∏–ª—å–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç –¥–∞–Ω–Ω—ã—Ö'
                    })
                
                with transaction.atomic():
                    for item in items:
                        product_id = item.get('product')
                        quantity = item.get('quantity')
                        given_to = item.get('given_to')
                        comment = item.get('comment', '')
                        
                        if not all([product_id, quantity, given_to]):
                            continue  # Skip incomplete items
                        
                        # Get product and check stock
                        product = Product.objects.get(pk=product_id)
                        stock = Stock.objects.get(product=product)
                        
                        if stock.quantity < float(quantity):
                            raise ValueError(f'–ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ "{product.name}". –í –Ω–∞–ª–∏—á–∏–∏: {stock.quantity}')
                        
                        # Create giving record
                        giving = Giving.objects.create(
                            product=product,
                            quantity=quantity,
                            given_to=given_to,
                            date=date,
                            comment=comment,
                            created_by=request.user
                        )
                        
                        # Update stock
                        stock.quantity -= float(quantity)
                        stock.save()
                    
                    return JsonResponse({'success': True})
            except ValueError as e:
                return JsonResponse({'success': False, 'error': str(e)})
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}'})
        else:
            # Handle traditional form submission (fallback)
            form = GivingForm(request.POST)
            if form.is_valid():
                try:
                    with transaction.atomic():
                        # Save giving record
                        giving = form.save(commit=False)
                        giving.created_by = request.user
                        giving.save()
                        
                        # Update stock
                        product = form.cleaned_data['product']
                        quantity = form.cleaned_data['quantity']
                        
                        stock = Stock.objects.get(product=product)
                        if stock.quantity < quantity:
                            raise ValueError(f'–ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ. –í –Ω–∞–ª–∏—á–∏–∏: {stock.quantity}')
                        
                        stock.quantity -= quantity
                        stock.save()
                        
                        messages.success(request, '–†–∞—Å—Ö–æ–¥ —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω.')
                        return redirect('infrastruction:giving_list')
                except Exception as e:
                    messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}')
    else:
        form = GivingForm(initial={'date': timezone.now().date()})
    
    return render(request, 'infrastruction/giving_form.html', {
        'form': form,
        'title': '–î–æ–±–∞–≤–∏—Ç—å —Ä–∞—Å—Ö–æ–¥'
    })

@login_required
@infrastructure_required
def giving_detail(request, pk):
    giving = get_object_or_404(Giving, pk=pk)
    return render(request, 'infrastruction/giving_detail.html', {
        'giving': giving,
        'title': f'–†–∞—Å—Ö–æ–¥ #{giving.pk}'
    })

@login_required
@infrastructure_required
def giving_edit(request, pk):
    giving = get_object_or_404(Giving, pk=pk)
    if request.method == 'POST':
        form = GivingForm(request.POST, instance=giving)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Check if product or quantity changed
                    old_product = giving.product
                    old_quantity = giving.quantity
                    new_product = form.cleaned_data['product']
                    new_quantity = form.cleaned_data['quantity']
                    
                    # Update stock - first return old quantity if product changed or quantity decreased
                    if old_product != new_product:
                        # Return old product quantity to stock
                        old_stock = Stock.objects.get(product=old_product)
                        old_stock.quantity += old_quantity
                        old_stock.save()
                        
                        # Take new product quantity from stock
                        new_stock = Stock.objects.get(product=new_product)
                        if new_stock.quantity < new_quantity:
                            raise ValueError(f'–ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ. –í –Ω–∞–ª–∏—á–∏–∏: {new_stock.quantity}')
                        new_stock.quantity -= new_quantity
                        new_stock.save()
                    elif old_quantity != new_quantity:
                        # Just update the same product stock
                        stock = Stock.objects.get(product=old_product)
                        quantity_diff = new_quantity - old_quantity
                        
                        if quantity_diff > 0 and stock.quantity < quantity_diff:
                            raise ValueError(f'–ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ. –í –Ω–∞–ª–∏—á–∏–∏: {stock.quantity}')
                        
                        stock.quantity -= quantity_diff
                        stock.save()
                    
                    # Save the giving record
                    giving = form.save()
                    
                    messages.success(request, '–†–∞—Å—Ö–æ–¥ —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω.')
                    return redirect('infrastruction:giving_list')
            except Exception as e:
                messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}')
    else:
        form = GivingForm(instance=giving)
    
    return render(request, 'infrastruction/giving_form.html', {
        'form': form,
        'title': '–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å —Ä–∞—Å—Ö–æ–¥'
    })

@login_required
@infrastructure_required
def giving_delete(request, pk):
    giving = get_object_or_404(Giving, pk=pk)
    try:
        with transaction.atomic():
            # Return the quantity to stock
            stock = Stock.objects.get(product=giving.product)
            stock.quantity += giving.quantity
            stock.save()
            
            # Delete the giving record
            giving.delete()
            
            messages.success(request, '–†–∞—Å—Ö–æ–¥ —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω.')
    except Exception as e:
        messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —É–¥–∞–ª–µ–Ω–∏–∏: {str(e)}')
    
    return redirect('infrastruction:giving_list')

@login_required
@infrastructure_required
def stock_list(request):
    stocks = Stock.objects.all().select_related('product').order_by('product__name')
    total_value = sum(stock.total_value for stock in stocks)
    
    return render(request, 'infrastruction/stock_list.html', {
        'stocks': stocks,
        'total_value': total_value,
        'title': '–°–∫–ª–∞–¥'
    })

@login_required
@infrastructure_required
def canteen_monthly_report(request):
    # Get the current month and year or use GET parameters
    now = timezone.now()
    year = int(request.GET.get('year', now.year))
    month = int(request.GET.get('month', now.month))
    
    # Create date range for the selected month
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Get all expenses for the selected month
    expenses = CanteenExpense.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')
    
    # Calculate totals - total_cost is a property, not a DB field
    total_cost = sum(expense.total_cost for expense in expenses)
    
    # Group expenses by product name
    product_totals = {}
    for expense in expenses:
        product_name = expense.product  # Use product field directly
        if product_name not in product_totals:
            product_totals[product_name] = 0
        product_totals[product_name] += expense.total_cost
    
    # Convert to list for the template
    product_summary = [
        {'name': name, 'total': total} 
        for name, total in product_totals.items()
    ]
    # Sort by total cost (descending)
    product_summary.sort(key=lambda x: x['total'], reverse=True)
    
    # Get available months for the dropdown
    year_month_pairs = CanteenExpense.objects.dates('date', 'month')
    available_months = [
        (pair.year, pair.month) for pair in year_month_pairs
    ]
    
    context = {
        'expenses': expenses,
        'total_cost': total_cost,
        'product_summary': product_summary,
        'year': year,
        'month': month,
        'month_name': start_date.strftime('%B'),
        'available_months': available_months,
        'title': f'–û—Ç—á–µ—Ç –∑–∞ {start_date.strftime("%B %Y")}'
    }
    
    return render(request, 'infrastruction/canteen_monthly_report.html', context)

@login_required
@infrastructure_required
def project_product_add(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk)
    if request.method == 'POST':
        form = ProjectProductForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Get form data
                    product = form.cleaned_data['product']
                    quantity = form.cleaned_data['quantity']
                    
                    # Check if there's enough stock
                    stock = Stock.objects.get(product=product)
                    if stock.quantity < quantity:
                        raise ValueError(f'–ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ. –í –Ω–∞–ª–∏—á–∏–∏: {stock.quantity}')
                    
                    # Save project product
                    project_product = form.save(commit=False)
                    project_product.project = project
                    project_product.created_by = request.user
                    project_product.save()
                    
                    # Update stock
                    stock.quantity -= quantity
                    stock.save()
                    
                    messages.success(request, '–ü—Ä–æ–¥—É–∫—Ç —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω –∫ –ø—Ä–æ–µ–∫—Ç—É.')
                    return redirect('infrastruction:project_detail', pk=project_pk)
            except Exception as e:
                messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}')
    else:
        form = ProjectProductForm(initial={'date_used': timezone.now().date()})
    
    return render(request, 'infrastruction/project_product_form.html', {
        'form': form,
        'project': project,
        'title': '–î–æ–±–∞–≤–∏—Ç—å –ø—Ä–æ–¥—É–∫—Ç –∫ –ø—Ä–æ–µ–∫—Ç—É'
    })

@login_required
@infrastructure_required
def project_product_edit(request, project_pk, pk):
    project = get_object_or_404(Project, pk=project_pk)
    project_product = get_object_or_404(ProjectProduct, pk=pk, project=project)
    
    if request.method == 'POST':
        form = ProjectProductForm(request.POST, instance=project_product)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Get old and new data to calculate stock changes
                    old_product = project_product.product
                    old_quantity = project_product.quantity
                    new_product = form.cleaned_data['product']
                    new_quantity = form.cleaned_data['quantity']
                    
                    # Handle stock updates
                    if old_product != new_product:
                        # Return old product quantity to stock
                        old_stock = Stock.objects.get(product=old_product)
                        old_stock.quantity += old_quantity
                        old_stock.save()
                        
                        # Take new product quantity from stock
                        new_stock = Stock.objects.get(product=new_product)
                        if new_stock.quantity < new_quantity:
                            raise ValueError(f'–ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ. –í –Ω–∞–ª–∏—á–∏–∏: {new_stock.quantity}')
                        new_stock.quantity -= new_quantity
                        new_stock.save()
                    elif old_quantity != new_quantity:
                        # Update same product stock
                        stock = Stock.objects.get(product=old_product)
                        quantity_diff = new_quantity - old_quantity
                        
                        if quantity_diff > 0 and stock.quantity < quantity_diff:
                            raise ValueError(f'–ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ. –í –Ω–∞–ª–∏—á–∏–∏: {stock.quantity}')
                        
                        stock.quantity -= quantity_diff
                        stock.save()
                    
                    # Save the changes
                    form.save()
                    messages.success(request, '–ó–∞–ø–∏—Å—å —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∞.')
                    return redirect('infrastruction:project_detail', pk=project_pk)
            except Exception as e:
                messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}')
    else:
        form = ProjectProductForm(instance=project_product)
    
    return render(request, 'infrastruction/project_product_form.html', {
        'form': form,
        'project': project,
        'project_product': project_product,
        'title': '–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å –ø—Ä–æ–¥—É–∫—Ç'
    })

@login_required
@infrastructure_required
def project_product_delete(request, project_pk, pk):
    project = get_object_or_404(Project, pk=project_pk)
    project_product = get_object_or_404(ProjectProduct, pk=pk, project=project)
    
    try:
        with transaction.atomic():
            # Return quantity to stock
            stock = Stock.objects.get(product=project_product.product)
            stock.quantity += project_product.quantity
            stock.save()
            
            # Delete record
            project_product.delete()
            messages.success(request, '–ó–∞–ø–∏—Å—å —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω–∞.')
    except Exception as e:
        messages.error(request, f'–û—à–∏–±–∫–∞ –ø—Ä–∏ —É–¥–∞–ª–µ–Ω–∏–∏: {str(e)}')
    
    return redirect('infrastruction:project_detail', pk=project_pk)

@login_required
@infrastructure_required
def project_product_report(request, pk=None):
    # Determine if filtering by specific project
    project = None
    if pk:
        project = get_object_or_404(Project, pk=pk)
        products = ProjectProduct.objects.filter(project=project).select_related('product', 'project')
    else:
        # Get all project products, with optional date filtering
        products = ProjectProduct.objects.all().select_related('product', 'project')
    
    # Apply date filters if provided
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            products = products.filter(date_used__gte=start_date)
        except ValueError:
            messages.warning(request, '–ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç –Ω–∞—á–∞–ª—å–Ω–æ–π –¥–∞—Ç—ã.')
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            products = products.filter(date_used__lte=end_date)
        except ValueError:
            messages.warning(request, '–ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç –∫–æ–Ω–µ—á–Ω–æ–π –¥–∞—Ç—ã.')
    
    # Group by project and product for summary
    summary = {}
    for item in products:
        if item.project.id not in summary:
            summary[item.project.id] = {
                'project': item.project,
                'products': {},
                'total': 0
            }
        
        if item.product.id not in summary[item.project.id]['products']:
            summary[item.project.id]['products'][item.product.id] = {
                'product': item.product,
                'quantity': 0,
                'total_cost': 0
            }
        
        summary[item.project.id]['products'][item.product.id]['quantity'] += item.quantity
        summary[item.project.id]['products'][item.product.id]['total_cost'] += item.total_cost
        summary[item.project.id]['total'] += item.total_cost
    
    # Convert summary to list for template
    projects_summary = []
    for project_id, data in summary.items():
        product_list = [p for p in data['products'].values()]
        projects_summary.append({
            'project': data['project'],
            'products': sorted(product_list, key=lambda x: x['total_cost'], reverse=True),
            'total': data['total']
        })
    
    # Sort projects by total cost descending
    projects_summary.sort(key=lambda x: x['total'], reverse=True)
    
    # Calculate grand total
    grand_total = sum(item.total_cost for item in products)
    
    context = {
        'title': f'–û—Ç—á–µ—Ç –ø–æ –ø—Ä–æ–¥—É–∫—Ç–∞–º –ø—Ä–æ–µ–∫—Ç–∞{f" {project.name}" if project else ""}',
        'project': project, 
        'products': products,
        'projects_summary': projects_summary,
        'grand_total': grand_total,
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render(request, 'infrastruction/project_product_report.html', context)

@login_required
@infrastructure_required
def export_receivings_excel(request):
    """Export receiving data to Excel"""
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "–ü–æ—Å—Ç—É–ø–ª–µ–Ω–∏—è"
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # ID
    ws.column_dimensions['B'].width = 15  # Date
    ws.column_dimensions['C'].width = 15  # Items count
    ws.column_dimensions['D'].width = 20  # Total price
    
    # Create header row with styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="AAAAFF", end_color="AAAAFF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['ID', '–î–∞—Ç–∞', '–ö–æ–ª-–≤–æ —ç–ª–µ–º–µ–Ω—Ç–æ–≤', '–û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å (SUM)']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all receivings ordered by date
    receivings = Receiving.objects.all().order_by('-date')
    
    # Add data rows
    for row_num, receiving in enumerate(receivings, 2):
        ws.cell(row=row_num, column=1, value=receiving.id)
        ws.cell(row=row_num, column=2, value=receiving.date.strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=3, value=receiving.items.count())
        ws.cell(row=row_num, column=4, value=receiving.total_price)
    
    # Create summary sheet for total statistics
    ws_summary = wb.create_sheet(title="–°–≤–æ–¥–∫–∞")
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Add summary headers
    summary_headers = ['–ü–æ–∫–∞–∑–∞—Ç–µ–ª—å', '–ó–Ω–∞—á–µ–Ω–∏–µ']
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add summary data
    total_receivings = receivings.count()
    total_value = sum(receiving.total_price for receiving in receivings)
    
    summary_data = [
        ['–û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø–æ—Å—Ç—É–ø–ª–µ–Ω–∏–π', total_receivings],
        ['–û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å –≤—Å–µ—Ö –ø–æ—Å—Ç—É–ø–ª–µ–Ω–∏–π', f"{total_value} SUM"],
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_num, column=1, value=label)
        ws_summary.cell(row=row_num, column=2, value=value)
    
    # Create detailed sheet with all items
    ws_details = wb.create_sheet(title="–î–µ—Ç–∞–ª–∏")
    ws_details.column_dimensions['A'].width = 10  # Receiving ID
    ws_details.column_dimensions['B'].width = 15  # Date
    ws_details.column_dimensions['C'].width = 30  # Product name
    ws_details.column_dimensions['D'].width = 15  # Quantity
    ws_details.column_dimensions['E'].width = 15  # Unit price
    ws_details.column_dimensions['F'].width = 20  # Total price
    ws_details.column_dimensions['G'].width = 30  # Comment
    
    # Add details headers
    details_headers = ['ID –ø–æ—Å—Ç—É–ø–ª.', '–î–∞—Ç–∞', '–ü—Ä–æ–¥—É–∫—Ç', '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ', '–¶–µ–Ω–∞ –µ–¥.', '–û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å', '–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π']
    for col_num, header in enumerate(details_headers, 1):
        cell = ws_details.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add details data
    row_num = 2
    for receiving in receivings:
        for item in receiving.items.all():
            ws_details.cell(row=row_num, column=1, value=receiving.id)
            ws_details.cell(row=row_num, column=2, value=receiving.date.strftime('%d.%m.%Y'))
            ws_details.cell(row=row_num, column=3, value=item.product.name)
            ws_details.cell(row=row_num, column=4, value=item.quantity)
            ws_details.cell(row=row_num, column=5, value=item.unit_price)
            ws_details.cell(row=row_num, column=6, value=item.total_price)
            ws_details.cell(row=row_num, column=7, value=item.comment if item.comment else "")
            row_num += 1
    
    # Set response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=receivings_report.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response

@login_required
@infrastructure_required
def export_canteen_monthly_report_excel(request):
    """Export canteen monthly report data to Excel"""
    # Get the current month and year or use GET parameters
    now = timezone.now()
    year = int(request.GET.get('year', now.year))
    month = int(request.GET.get('month', now.month))
    
    # Create date range for the selected month
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Get all expenses for the selected month
    expenses = CanteenExpense.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')
    
    # Calculate totals
    total_cost = sum(expense.total_cost for expense in expenses)
    
    # Group expenses by product name
    product_totals = {}
    for expense in expenses:
        product_name = expense.product
        if product_name not in product_totals:
            product_totals[product_name] = 0
        product_totals[product_name] += expense.total_cost
    
    # Convert to list for processing
    product_summary = [
        {'name': name, 'total': total} 
        for name, total in product_totals.items()
    ]
    # Sort by total cost (descending)
    product_summary.sort(key=lambda x: x['total'], reverse=True)
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "–û—Ç—á–µ—Ç —Å—Ç–æ–ª–æ–≤–æ–π"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Date
    ws.column_dimensions['B'].width = 35  # Product
    ws.column_dimensions['C'].width = 15  # Quantity
    ws.column_dimensions['D'].width = 15  # Unit Price
    ws.column_dimensions['E'].width = 20  # Total Cost
    
    # Add title
    month_name = start_date.strftime('%B')
    title = f"–û—Ç—á–µ—Ç —Ä–∞—Å—Ö–æ–¥–æ–≤ —Å—Ç–æ–ª–æ–≤–æ–π: {month_name} {year}"
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add summary data
    ws.merge_cells('A3:B3')
    summary_label = ws.cell(row=3, column=1, value="–û–±—â–∏–µ —Ä–∞—Å—Ö–æ–¥—ã –∑–∞ –ø–µ—Ä–∏–æ–¥:")
    summary_label.font = Font(bold=True)
    
    ws.merge_cells('C3:E3')
    summary_value = ws.cell(row=3, column=3, value=f"{total_cost} SUM")
    summary_value.font = Font(bold=True)
    summary_value.alignment = Alignment(horizontal='right')
    
    # Create header row for detailed expenses with styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="AAAAFF", end_color="AAAAFF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add a little space
    row_num = 5
    
    # Add detailed expenses table
    ws.cell(row=row_num, column=1, value="–î–µ—Ç–∞–ª–∏–∑–∞—Ü–∏—è —Ä–∞—Å—Ö–æ–¥–æ–≤").font = Font(bold=True, size=12)
    row_num += 2
    
    headers = ['–î–∞—Ç–∞', '–ü—Ä–æ–¥—É–∫—Ç/–£—Å–ª—É–≥–∞', '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ', '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É', '–û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    row_num += 1
    
    # Add data rows
    for expense in expenses:
        ws.cell(row=row_num, column=1, value=expense.date.strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=2, value=expense.product)
        ws.cell(row=row_num, column=3, value=expense.quantity)
        ws.cell(row=row_num, column=4, value=expense.unit_price)
        ws.cell(row=row_num, column=5, value=expense.total_cost)
        row_num += 1
    
    # Add total row
    ws.cell(row=row_num, column=1, value="–ò—Ç–æ–≥–æ:").font = Font(bold=True)
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='right')
    
    total_cell = ws.cell(row=row_num, column=5, value=total_cost)
    total_cell.font = Font(bold=True)
    
    # Create summary by product sheet
    ws_summary = wb.create_sheet(title="–ü–æ –ø—Ä–æ–¥—É–∫—Ç–∞–º")
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    
    # Add title to summary sheet
    ws_summary.merge_cells('A1:C1')
    summary_title = ws_summary.cell(row=1, column=1, value=f"–†–∞—Å—Ö–æ–¥—ã –ø–æ –ø—Ä–æ–¥—É–∫—Ç–∞–º: {month_name} {year}")
    summary_title.font = Font(bold=True, size=14)
    summary_title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers to summary sheet
    summary_headers = ['–ü—Ä–æ–¥—É–∫—Ç/–£—Å–ª—É–≥–∞', '–û–±—â–∞—è —Å—É–º–º–∞', '% –æ—Ç –æ–±—â–µ–≥–æ']
    
    row_num = 3
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    row_num += 1
    
    # Add product summary data
    for item in product_summary:
        ws_summary.cell(row=row_num, column=1, value=item['name'])
        ws_summary.cell(row=row_num, column=2, value=item['total'])
        percentage = (item['total'] / total_cost * 100) if total_cost > 0 else 0
        ws_summary.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
        row_num += 1
    
    # Add total row to summary
    ws_summary.cell(row=row_num, column=1, value="–ò—Ç–æ–≥–æ:").font = Font(bold=True)
    ws_summary.cell(row=row_num, column=2, value=total_cost).font = Font(bold=True)
    ws_summary.cell(row=row_num, column=3, value="100%").font = Font(bold=True)
    
    # Set response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=canteen_report_{year}_{month}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response

@login_required
@infrastructure_required
def export_project_excel(request, pk):
    """Export project data to Excel"""
    # Get the project and its related products
    project = get_object_or_404(Project, pk=pk)
    products = project.used_products.all().select_related('product')
    
    # Calculate total product cost
    total_product_cost = sum(product.total_cost for product in products)
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "–î–µ—Ç–∞–ª–∏ –ø—Ä–æ–µ–∫—Ç–∞"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    
    # Add title
    title = f"–ü—Ä–æ–µ–∫—Ç: {project.name}"
    ws.merge_cells('A1:B1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add project information
    row_num = 3
    
    # Styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Project Info section header
    ws.merge_cells('A3:B3')
    info_header = ws.cell(row=row_num, column=1, value="–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ –ø—Ä–æ–µ–∫—Ç–µ")
    info_header.font = Font(bold=True, size=12)
    info_header.fill = header_fill
    row_num += 1
    
    # Project details
    project_details = [
        ("–°—Ç–∞—Ç—É—Å", {
            'planned': '–ó–∞–ø–ª–∞–Ω–∏—Ä–æ–≤–∞–Ω',
            'in_progress': '–í –ø—Ä–æ—Ü–µ—Å—Å–µ',
            'completed': '–ó–∞–≤–µ—Ä—à–µ–Ω',
            'cancelled': '–û—Ç–º–µ–Ω–µ–Ω'
        }.get(project.status, project.status)),
        ("–î–∞—Ç–∞ –Ω–∞—á–∞–ª–∞", project.start_date.strftime('%d.%m.%Y')),
        ("–î–∞—Ç–∞ –æ–∫–æ–Ω—á–∞–Ω–∏—è", project.end_date.strftime('%d.%m.%Y') if project.end_date else "–ù–µ —É–∫–∞–∑–∞–Ω–∞"),
        ("–û–ø–∏—Å–∞–Ω–∏–µ", project.description or "–ù–µ —É–∫–∞–∑–∞–Ω–æ"),
        ("–û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å –ø—Ä–æ–¥—É–∫—Ç–æ–≤", f"{total_product_cost} SUM")
    ]
    
    for label, value in project_details:
        ws.cell(row=row_num, column=1, value=label).font = header_font
        ws.cell(row=row_num, column=2, value=value)
        row_num += 1
    
    # Add space before products section
    row_num += 2
    
    # Products section header
    ws.merge_cells(f'A{row_num}:G{row_num}')
    products_header = ws.cell(row=row_num, column=1, value="–ò—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–Ω—ã–µ –ø—Ä–æ–¥—É–∫—Ç—ã")
    products_header.font = Font(bold=True, size=12)
    products_header.fill = header_fill
    row_num += 1
    
    # Products table headers
    headers = ['–ü—Ä–æ–¥—É–∫—Ç', '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É', '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ', '–°—Ç–æ–∏–º–æ—Å—Ç—å', '–î–∞—Ç–∞ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è', '–ü—Ä–∏–º–µ—á–∞–Ω–∏—è']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        # Adjust column width based on header
        ws.column_dimensions[chr(64 + col_num)].width = max(15, len(header) * 1.2)
    
    row_num += 1
    
    # Add products data
    for item in products:
        ws.cell(row=row_num, column=1, value=item.product.name)
        ws.cell(row=row_num, column=2, value=f"{item.product.unit_price} SUM")
        ws.cell(row=row_num, column=3, value=item.quantity)
        ws.cell(row=row_num, column=4, value=f"{item.total_cost} SUM")
        ws.cell(row=row_num, column=5, value=item.date_used.strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=6, value=item.notes if item.notes else "-")
        row_num += 1
    
    # Add total row
    ws.cell(row=row_num, column=1, value="–ò—Ç–æ–≥–æ:").font = Font(bold=True)
    ws.merge_cells(f'A{row_num}:C{row_num}')
    total_cell = ws.cell(row=row_num, column=4, value=f"{total_product_cost} SUM")
    total_cell.font = Font(bold=True)
    
    # Set response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=project_{project.id}_{project.name}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response

@login_required
@infrastructure_required
def export_products_excel(request):
    """Export products data to Excel"""
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "–¢–æ–≤–∞—Ä—ã"
    
    # Set column widths
    ws.column_dimensions['A'].width = 40  # –ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ
    ws.column_dimensions['B'].width = 20  # –ï–¥–∏–Ω–∏—Ü–∞ –∏–∑–º–µ—Ä–µ–Ω–∏—è
    ws.column_dimensions['C'].width = 15  # –¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É
    ws.column_dimensions['D'].width = 15  # –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –Ω–∞ —Å–∫–ª–∞–¥–µ
    ws.column_dimensions['E'].width = 20  # –û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å
    ws.column_dimensions['F'].width = 30  # –ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–∏
    
    # Create header row with styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="AAAAFF", end_color="AAAAFF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ', '–ï–¥–∏–Ω–∏—Ü–∞ –∏–∑–º–µ—Ä–µ–Ω–∏—è', '–¶–µ–Ω–∞ –∑–∞ –µ–¥–∏–Ω–∏—Ü—É', '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –Ω–∞ —Å–∫–ª–∞–¥–µ', '–û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å', '–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–∏']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all products with their stock information
    products = Product.objects.all().prefetch_related('stock')
    
    # Create dictionaries for easier access to stock information and totals
    stock_dict = {}
    total_value_dict = {}
    comments_dict = {}
    
    for stock in Stock.objects.select_related('product').all():
        stock_dict[stock.product_id] = stock.quantity
        total_value_dict[stock.product_id] = stock.quantity * stock.product.unit_price
    
    # Get the most recent comments for each product from receiving transactions
    for product in products:
        # Get the most recent receiving items with comments for this product
        recent_items = ReceivingItem.objects.filter(
            product_id=product.id, 
            comment__isnull=False
        ).exclude(
            comment__exact=''
        ).order_by('-receiving__date', '-created_at')[:1]
        
        if recent_items:
            comments_dict[product.id] = recent_items[0].comment
    
    # Add data rows
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.name)
        ws.cell(row=row_num, column=2, value="—à—Ç.")  # Default unit - can be extended if needed
        ws.cell(row=row_num, column=3, value=product.unit_price)
        
        quantity = stock_dict.get(product.id, 0)
        ws.cell(row=row_num, column=4, value=quantity)
        
        total_value = total_value_dict.get(product.id, 0)
        ws.cell(row=row_num, column=5, value=total_value)
        
        comment = comments_dict.get(product.id, "")
        ws.cell(row=row_num, column=6, value=comment)
    
    # Calculate and add grand total row
    grand_total = sum(total_value_dict.values())
    
    row_num = len(products) + 2
    ws.cell(row=row_num, column=1, value="–ò—Ç–æ–≥–æ:").font = Font(bold=True)
    ws.merge_cells(f'A{row_num}:D{row_num}')
    
    total_cell = ws.cell(row=row_num, column=5, value=grand_total)
    total_cell.font = Font(bold=True)
    
    # Set response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products_report.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response

# Order views
@login_required
def order_list(request):
    """View for listing orders"""
    orders = Order.objects.filter(created_by=request.user).order_by('-date')
    
    context = {
        'orders': orders,
    }
    return render(request, 'infrastruction/order_list.html', context)

@login_required
def order_add(request):
    """View for adding a new order"""
    if request.method == 'POST':
        form = OrderForm(request.POST)
        formset = OrderItemFormSet(request.POST, prefix='items')
        
        if form.is_valid() and formset.is_valid():
            order = form.save(commit=False)
            order.created_by = request.user
            order.save()
            
            formset.instance = order
            formset.save()
            
            # Save products to OrderProduct model for future use
            for item_form in formset.forms:
                if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                    product_name = item_form.cleaned_data.get('product')
                    if product_name:
                        OrderProduct.objects.get_or_create(name=product_name)
            
            # Send notification to Telegram
            send_order_to_telegram(order)
            
            messages.success(request, '–ó–∞–∫–∞–∑ —É—Å–ø–µ—à–Ω–æ —Å–æ–∑–¥–∞–Ω')
            return redirect('infrastruction:order_list')
    else:
        form = OrderForm()
        formset = OrderItemFormSet(prefix='items')
    
    context = {
        'form': form,
        'formset': formset,
    }
    return render(request, 'infrastruction/order_form.html', context)

@login_required
def order_edit(request, pk):
    """View for editing an order"""
    order = get_object_or_404(Order, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        formset = OrderItemFormSet(request.POST, instance=order, prefix='items')
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            
            # Save products to OrderProduct model for future use
            for item_form in formset.forms:
                if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                    product_name = item_form.cleaned_data.get('product')
                    if product_name:
                        OrderProduct.objects.get_or_create(name=product_name)
            
            messages.success(request, '–ó–∞–∫–∞–∑ —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω')
            return redirect('infrastruction:order_list')
    else:
        form = OrderForm(instance=order)
        formset = OrderItemFormSet(instance=order, prefix='items')
    
    context = {
        'form': form,
        'formset': formset,
        'order': order,
    }
    return render(request, 'infrastruction/order_form.html', context)

@login_required
def order_detail(request, pk):
    """View for viewing order details"""
    order = get_object_or_404(Order, pk=pk, created_by=request.user)
    
    context = {
        'order': order,
    }
    return render(request, 'infrastruction/order_detail.html', context)

@login_required
def order_delete(request, pk):
    """View for deleting an order"""
    order = get_object_or_404(Order, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        order.delete()
        messages.success(request, '–ó–∞–∫–∞–∑ —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω')
        return redirect('infrastruction:order_list')
    
    context = {
        'order': order,
    }
    return render(request, 'infrastruction/order_confirm_delete.html', context)

@login_required
@require_POST
def change_order_status(request):
    """AJAX view for changing order status"""
    order_id = request.POST.get('order_id')
    new_status = request.POST.get('status')
    
    try:
        order = Order.objects.get(pk=order_id, created_by=request.user)
        order.status = new_status
        order.save()
        return JsonResponse({'success': True})
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Order not found'}, status=404)

@login_required
def get_order_products(request):
    """AJAX view to get products for order form autocomplete"""
    if request.method == 'POST':
        try:
            # Handle adding a new product
            data = json.loads(request.body)
            product_name = data.get('name')
            
            if product_name:
                OrderProduct.objects.get_or_create(name=product_name)
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'Product name is required'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET request - return all products or filter by query
    query = request.GET.get('q', '')
    
    if query:
        # First, get products from OrderProduct model
        order_products = list(OrderProduct.objects.filter(
            name__icontains=query
        ).values_list('name', flat=True).distinct()[:10])
        
        # If we need more results, also look at previous order items
        if len(order_products) < 10:
            previous_products = OrderItem.objects.filter(
                product__icontains=query
            ).values_list('product', flat=True).distinct()[:10 - len(order_products)]
            
            # Add these to the result list if not already present
            for product in previous_products:
                if product not in order_products:
                    order_products.append(product)
    else:
        # Get all products (limited to 50 to avoid performance issues)
        order_products = list(OrderProduct.objects.values_list('name', flat=True).distinct()[:50])
        
        # Add common products from previous orders if we have fewer than 50 products
        if len(order_products) < 50:
            previous_products = OrderItem.objects.values_list('product', flat=True).distinct()[:50 - len(order_products)]
            
            # Add these to the result list if not already present
            for product in previous_products:
                if product not in order_products:
                    order_products.append(product)
    
    return JsonResponse({'products': order_products})

# Telegram integration
def send_order_to_telegram(order):
    """Send order notification to Telegram groups"""
    telegram_groups = TelegramGroup.objects.filter(is_active=True)
    
    if not telegram_groups:
        return
    
    # Format the message
    items_text = "\n".join([
        f"- {item.product}: {item.quantity} {item.get_unit_display()}"
        for item in order.items.all()
    ])
    
    message = f"""üÜï *–ù–û–í–´–ô –ó–ê–ö–ê–ó* üÜï
*–ù–æ–º–µ—Ä –∑–∞–∫–∞–∑–∞:* {order.order_number}
*–î–∞—Ç–∞:* {order.date.strftime('%d.%m.%Y')}
*–°–æ–∑–¥–∞–ª:* {order.created_by.get_full_name() or order.created_by.username}

*–°–æ–¥–µ—Ä–∂–∏–º–æ–µ –∑–∞–∫–∞–∑–∞:*
{items_text}

*–ü—Ä–∏–º–µ—á–∞–Ω–∏—è:* {order.notes or '–ù–µ—Ç –ø—Ä–∏–º–µ—á–∞–Ω–∏–π'}
"""
    
    # Send to each group
    for group in telegram_groups:
        try:
            send_telegram_message(group.chat_id, message)
        except Exception as e:
            # Log error, but continue with other groups
            print(f"Error sending to Telegram group {group.name}: {e}")

def send_telegram_message(chat_id, message):
    """Send message to Telegram chat using bot API"""
    # Get bot token from settings
    bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
    
    if not bot_token:
        # Log error and return
        print("Telegram bot token not configured")
        return
    
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {
        'chat_id': chat_id,
        'text': message,
        'parse_mode': 'Markdown'
    }
    
    response = requests.post(url, json=payload)
    return response.json()

# ... rest of the views ... 